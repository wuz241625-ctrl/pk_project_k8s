import json
import logging
import time
from datetime import datetime, timedelta
import pandas as pd
import xlrd
import re
import csv
import io
from openpyxl import load_workbook
import random
from io import BytesIO

from decimal import Decimal
import tornado
from aiomysql import DictCursor

from application.base import BaseHandler
from application.message import msg
import hashlib
import requests
import math
import traceback


from .query_third_order_status import query_lucky_order, query_apay_order, query_kingpay_order, query_wepay_order, query_kingpay2_order, query_777pay_order, query_swiftpay_order, query_quickpay_order, query_snakepay_order, query_hkpay_order, query_skpay_order, \
query_ospay_order, query_ospay_upi_order, query_TataPay_order, query_789pay_upi_order, query_789pay_order, query_TataPay_t100037_order,query_Vibrapay_order,query_qqpay_order, query_gamepayer_order, query_easypay_order


def build_third_duplicate_lookup_payload(third_party_name, utr, query_result=None):
    if third_party_name == 'easypay':
        trans_id = str((query_result or {}).get('transactionId') or '').strip()
        if not trans_id:
            return None
        return {'field': 'trans_id', 'value': trans_id, 'message_key': 10320}
    return {'field': 'utr', 'value': utr, 'message_key': 10229}

class BaseOrderHandler(BaseHandler):
    # 确认
    async def order_success(self, code, flag = 0):
        # 查询订单
        sql_select_order = """select * from orders_df where code=%s and status in (1,2) order by id desc limit 1"""
        # 商户代理费率
        sql_select_rates = """select id,rate_df from (select @orgId id, (select rate_df from merchant where id=@orgId) rate_df,
                                    (select @orgId:=pid from merchant where id=@orgId) pid from 
                                    (select @orgId:=%s) vars,merchant) t where id is not null order by pid desc"""
        # 更新系统余额
        sql_update_payment = """update payment set sys_balance=sys_balance+%s where id=%s"""
        # 更新订单
        sql_update = """update orders_df set earn_merchant=%s,time_success=%s,status=%s where code=%s and status in (-1,1,2) limit 1"""
        # --- 新增 SQL 查询：用于检查子订单完成状态 ---
        # 需要 orders_df 表有 parent_id 字段来表示父订单的 code
        # 还需要 is_split 字段（比如 1 表示子订单，0 表示母订单或独立订单）来区分类型
        sql_select_sub_orders = """SELECT COUNT(*) AS total_subs,
                                    SUM(CASE WHEN status = 4 THEN 1 ELSE 0 END) AS successful_subs
                                FROM orders_df
                                WHERE parent_id = %s AND is_del=0""" # parent_id 存储的是父订单的 code
        # --- 用于查询母订单下所有未成功的子订单 ---
        sql_select_child_orders_for_success = """
            SELECT code FROM orders_df
            WHERE parent_id = %s AND status IN (1, 2)
        """
        
        time_now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        earn_merchant = Decimal(0)
        newStatus = 0
        
        # -----------------------------------------------
        async with self.application.db.acquire() as conn:
            async with conn.cursor(DictCursor) as cur:
                try:
                    # 查询订单
                    if not await cur.execute(sql_select_order, code):
                        if flag == 0:
                            return await self.json_response(msg[10007])
                        else:
                            return msg[10007]
                    
                    order = (await cur.fetchall())[0]
                    amount = order['amount']
                    
                    # 先给-1， 1=常规订单，2=拆单主单，3=拆单子单
                    order_type = -1
                    if not order.get('is_split') and not order.get('parent_id'):
                        order_type = 1
                        newStatus = 3
                    if order.get('is_split') and not order.get('parent_id'):
                        order_type = 2
                        newStatus = 3
                    if order.get('parent_id'):
                        order_type = 3
                        newStatus = 4
                    
                    # 需求确定为通过“人为操作规范”来限制，程序不限制
                    # 如【母单】下并非所有【子单】【皆为成功】的情况下执行，线下返还失败金额
                    
                    # 需求调整，母子单状态无关联，注释相关代码
                    # region【重要新增】子订单处理逻辑从这里开始
                    # if order_type == 3
                    #     newStatus = 4
                        
                    #     # 中文日志：识别到子订单，开始处理其自身状态
                    #     self.logger.info(f"识别到子订单 {code}，父订单为 {order['parent_id']}。")

                    #     # 1. 更新当前子订单的状态为成功 (status=4)
                    #     # 子订单的 earn_merchant 通常设为 0，因为佣金在母订单上计算
                    #     # 这里 status in (-1,1,2) 确保只更新待处理的子订单
                    #     if not await cur.execute(sql_update, (earn_merchant, time_now, newStatus, code)):
                    #         await conn.rollback() # 子订单状态更新失败，立即回滚
                    #         self.logger.warning(f"子订单 {code} 状态更新失败，已回滚。")
                    #         return msg[10007]

                    #     # 【关键】单独提交子订单的状态更新
                    #     # 确保子订单的状态变化能立即持久化，不依赖于父订单的最终处理
                    #     await conn.commit()
                    #     # 中文日志：子订单状态已成功更新并提交
                    #     self.logger.info(f"子订单 {code} 已成功更新状态为 4，并已提交其事务。")

                    #     # 2. 查询父订单下的所有子订单状态
                    #     parent_code = order['parent_id']
                    #     if not await cur.execute(sql_select_sub_orders, parent_code):
                    #         # 中文日志：未找到父订单的子订单信息 (异常情况)
                    #         self.logger.warning(f"未找到父订单 {parent_code} 的子订单信息。")
                    #         # 这种情况通常是数据异常，需要特殊处理
                    #         return msg[10007]

                    #     sub_order_stats = (await cur.fetchall())[0]
                    #     # 检查所有子订单是否都已成功
                    #     if sub_order_stats['total_subs'] > 0 and \
                    #     sub_order_stats['total_subs'] == sub_order_stats['successful_subs']:
                    #         # 所有子订单都已成功，开始处理父订单
                    #         self.logger.info(f"父订单 {parent_code} 的所有子订单 ({sub_order_stats['total_subs']} 个) 都已成功，开始处理父订单。")
                    #         # 递归调用 order_success 来处理父订单
                    #         await self.order_success(parent_code, flag=1)
                    #         if flag == 0:
                    #             return await self.json_response(dict(code=20000, msg='操作成功'))
                    #         else:
                    #             return dict(code=20000, msg='操作成功')
                    #     else:
                    #         # 子订单处理成功，但仍有其他子订单待完成
                    #         self.logger.info(f"子订单 {code} 处理成功，但父订单 {parent_code} 仍有 {sub_order_stats['total_subs'] - sub_order_stats['successful_subs']} 个子订单待完成。")
                    #         if flag == 0:
                    #             return await self.json_response(dict(code=20000, msg='操作成功'))
                    #         else:
                    #             return dict(code=20000, msg='操作成功')
                    # endregion 子订单处理逻辑到此结束 ---
                    
                    # region 码商余额处理
                    if order_type in [1, 3]:
                        partner_id = order['partner_id']
                        if partner_id:
                            self.logger.info(f"开始处理码商{partner_id}的余额")

                            # 更新码商余额
                            self.logger.info(f"开始更新码商{partner_id}余额，金额：{amount}")
                            if not await self.change_balance(conn, cur, 'partner', partner_id, amount, code, 1):
                                await conn.rollback()
                                self.logger.error(f"更新码商{partner_id}余额失败，金额：{amount}")
                                if flag == 0:
                                    return await self.json_response(msg[10007])
                                else:
                                    return msg[10007]
                    # endregion 码商余额处理

                    # region 码商佣金
                    if order_type in [1]:
                        partner_id = order['partner_id']
                        if partner_id:
                            self.logger.info(f"开始更新码商{partner_id}的佣金，金额：{order['earn_partner_self']}")
                            if not await self.change_balance(conn, cur, 'partner', partner_id, order['earn_partner_self'], code, 3):
                                await conn.rollback()
                                self.logger.error(f"更新码商{partner_id}佣金失败，金额：{order['earn_partner_self']}")
                                if flag == 0:
                                    return await self.json_response(msg[10007])
                                else:
                                    return msg[10007]
                        
                            self.logger.info(f"码商{partner_id}佣金更新成功，金额：{order['earn_partner_self']}")

                            # 优惠代付处理
                            disprice = Decimal(0)
                            range_df = (await self.get_cache_result('sys_info', ['range_df']))['range_df']
                            
                            if range_df:
                                self.logger.info(f"开始检查代付优惠，当前金额：{amount}，商户ID：{order['merchant_id']}")
                                range_df = json.loads(range_df)
                                for i in range(1, 7):
                                    if range_df['isOpen' + str(i)] == 1:
                                        if Decimal(range_df['rangemin' + str(i)]) <= amount <= Decimal(
                                                range_df['rangemax' + str(i)]):
                                            disprice = Decimal(range_df['disprice' + str(i)])
                                            self.logger.info(
                                                '代付优惠 disprice:{disprice} rangemin:{rangemin} rangemax:{rangemax} amount:{amount} merchant_id:{merchant_id}'.format(
                                                    disprice=disprice, rangemin=range_df['rangemin' + str(i)],
                                                    rangemax=range_df['rangemax' + str(i)], amount=amount,
                                                    merchant_id=order['merchant_id']))
                                            self.logger.info(f"代付优惠应用：优惠金额{disprice}，范围：{range_df['rangemin' + str(i)]}-{range_df['rangemax' + str(i)]}，订单金额：{amount}，商户ID：{order['merchant_id']}")
                                            break

                            # 代付优惠入库
                            if disprice > 0:
                                self.logger.info(f"代付优惠入库，码商{partner_id}，优惠金额：{disprice}")
                                if not await self.change_balance(conn, cur, 'partner', partner_id, disprice, code, 10):
                                    await conn.rollback()
                                    self.logger.error(f"代付优惠入库失败，码商{partner_id}，优惠金额：{disprice}")
                                    return dict(code=99, msg='Failed to add partner balance')
                            self.logger.info(f"代付优惠入库成功，码商{partner_id}，优惠金额：{disprice}")
                    # endregion 码商佣金
                    
                    # region 修改卡系统余额
                    if  order_type in [1, 3]:
                        if order['payment_id']:
                            self.logger.info(f"开始修改卡系统余额，支付ID：{order['payment_id']}，修改金额：{-amount}")
                            if not await cur.execute(sql_update_payment, (-amount, order['payment_id'])):
                                await conn.rollback()
                                self.logger.error(f"修改卡系统余额失败，支付ID：{order['payment_id']}，金额：{-amount}")
                                if flag == 0:
                                    return await self.json_response(msg[10007])
                                else:
                                    return msg[10007]
                            self.logger.info(f"卡系统余额修改成功，支付ID：{order['payment_id']}，金额：{-amount}")
                    # endregion 修改卡系统余额
                    
                    # region 常规单 或 母单
                    if order_type in [1, 2]:
                        # region 扣商户(过期订单)
                        if order['status'] == -1:
                            self.logger.info(f"开始处理订单：{order['merchant_id']}，支付金额：{-order['realpay']}，支付代码：{code}")
                            if not await self.change_balance(conn, cur, 'merchant', order['merchant_id'], -order['realpay'], code, 0):
                                self.logger.error(f"余额更新失败，商户ID：{order['merchant_id']}，支付金额：{-order['realpay']}")
                                if flag == 0:
                                    self.logger.info(f"返回错误信息: {msg[10016]}")
                                    return await self.json_response(msg[10016])
                                else:
                                    self.logger.info(f"返回错误信息: {msg[10016]}")
                                    return msg[10016]
                            self.logger.info(f"余额更新成功，商户ID：{order['merchant_id']}，支付金额：{-order['realpay']}")
                        # endregion 扣商户(过期订单)
                        # region 商户代理费用
                        if order['earn_merchant'] > Decimal(0):
                            self.logger.info(f"开始处理商户{order['merchant_id']} = {order['earn_merchant']}的代付订单，订单金额：{order['amount']}")
                            # 查询商户费率
                            if not await cur.execute(sql_select_rates, order['merchant_id']):
                                await conn.rollback()
                                self.logger.error(f"查询商户{order['merchant_id']}的费率失败")
                                if flag == 0:
                                    return await self.json_response(msg[10007])
                                else:
                                    return msg[10007]
                            
                            merchant_prates = await cur.fetchall()
                            
                            # 处理商户的代付费用差
                            for k, v in enumerate(merchant_prates):
                                if not k == 0 and v['rate_df']:
                                    _amount = order['amount'] * (merchant_prates[k - 1]['rate_df'] - v['rate_df'])
                                    self.logger.info(f"计算代付费用差，上一商户费率：{merchant_prates[k - 1]['rate_df']}，当前商户费率：{v['rate_df']}，差额：{_amount}")
                                    
                                    if _amount == 0:
                                        self.logger.info('代付订单{code}没有代付费用差,上级商户{id}费率{rate_df} ,本级商户{id2}费率{rate_df2}'.format(code=code, id=merchant_prates[k - 1]['id'],rate_df=merchant_prates[k - 1]['rate_df'], id2=v['id'],rate_df2=v['rate_df']))
                                        self.logger.info(f"代付订单{code}没有代付费用差, 上级商户{merchant_prates[k - 1]['id']}费率{merchant_prates[k - 1]['rate_df']}，本级商户{v['id']}费率{v['rate_df']}")
                                        continue
                                    
                                    if _amount < 0:
                                        await conn.rollback()
                                        self.logger.error(f"代付费用差为负值，商户{v['id']}余额不足，差额：{_amount}")
                                        if flag == 0:
                                            return await self.json_response(msg[10013])
                                        else:
                                            return msg[10013]

                                    # 更新商户余额
                                    self.logger.info(f"开始更新商户{v['id']}的余额，金额：{_amount}")
                                    if not await self.change_balance(conn, cur, 'merchant', v['id'], _amount, code, 3):
                                        await conn.rollback()
                                        self.logger.error(f"更新商户{v['id']}余额失败，金额：{_amount}")
                                        if flag == 0:
                                            return await self.json_response(msg[10007])
                                        else:
                                            return msg[10007]
                                    
                                    earn_merchant += _amount
                                    # 码商余额
                                    self.logger.info(f"商户{v['id']}余额更新成功，金额：{_amount}")
                        # endregion 商户代理费用
                    # endregion 常规单 或 母单
                    
                    # 需求调整，母子单状态无关联，注释相关代码
                    # region【新增逻辑】如果当前订单是母订单，级联更新其所有活跃子订单为成功
                    # if order.get('is_split') == 1: # 假设 is_split=1 表示是母订单
                    #     self.logger.info(f"检测到订单 {code} 是母订单，正在尝试级联完成其所有活跃子订单。")
                    #     self._last_sql = sql_select_child_orders_for_success % code
                    #     await cur.execute(sql_select_child_orders_for_success, code)
                    #     child_orders_to_complete_data = await cur.fetchall()

                    #     if child_orders_to_complete_data:
                    #         child_codes_to_complete = [co['code'] for co in child_orders_to_complete_data]
                    #         in_placeholders = ', '.join(['%s'] * len(child_codes_to_complete))
                            

                    #         time_success_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    #         update_params_for_execute = [time_success_str] + child_codes_to_complete

                    #         self.logger.info(f"正在批量更新子订单状态为成功: {child_codes_to_complete}")
                                                
                    #         dynamic_sql_update_child_order = f"""
                    #             UPDATE orders_df SET status=4, earn_merchant=0, time_success=%s
                    #             WHERE code IN ({in_placeholders}) AND status IN (1, 2)
                    #         """
                    #         if not await cur.execute(dynamic_sql_update_child_order, tuple(update_params_for_execute)):
                    #             self.logger.error(f"批量更新子订单 {child_codes_to_complete} 状态为成功失败。")
                    #             await conn.rollback()
                    #             if flag == 0:
                    #                 return self.json_response(self.msg[10007])
                    #             else:
                    #                 return {'code': 1, 'msg': self.msg[10007]}
                    #         self.logger.info(f"母订单 {code} 的 {len(child_codes_to_complete)} 个子订单已成功级联完成。")
                    #     else:
                    #         self.logger.info(f"母订单 {code} 没有找到活跃子订单需要级联完成。")
                    # endregion

                    # 修改订单状态
                    if not await cur.execute(sql_update, (earn_merchant, time_now, newStatus, code)):
                        await conn.rollback()
                        self.logger.warning(cur._last_executed)
                        if flag == 0:
                            return await self.json_response(msg[10007])
                        else:
                            return msg[10007]
                    self.logger.info('更新订单状态%s' % cur._last_executed)
                    # 统计操作
                    if not await self.create_result('operate', {'type': 12, 'admin_id': self.current_user['id'], 'ip': await self.get_ip()}):
                        await conn.rollback()
                        self.logger.warning(cur._last_executed)
                        if flag == 0:
                            return await self.json_response(msg[10007])
                        else:
                            return msg[10007]
                except Exception as e:
                    self.logger.warning('确认订单失败,code={code},异常={e}'.format(code=code, e=e))
                    await conn.rollback()
                    if flag == 0:
                        return await self.json_response(msg[10007])
                    else:
                        return msg[10007]
                else:
                    await conn.commit()
                    if order['payment_id']:
                        if await self.redis.sismember('payment_online_df', order['payment_id']):
                            await self.redis.lrem('payment_active_df', 0, order['payment_id'])
                            await self.redis.rpush('payment_active_df', order['payment_id'])                    
                    # 代付回调-确认
                    if order_type in [1, 2]:
                        await self.redis.publish('order_df_notify', code)                        
                    if flag == 0:
                        return await self.json_response(dict(code=20000, msg='操作成功'))
                    else:
                        return dict(code=20000, msg='操作成功')
    async def acquire_spin_lock(self, code):
        """使用自旋锁，5s使用自旋锁，防止取消的同时回调"""
        count_circle = 0
        while True:
            busy_key = 'grab_df_{code}'.format(code=code)
            if await self.redis.setnx(busy_key, 1):
                await self.redis.expire(busy_key, 60)  # 修改为60秒，足够处理数据库操作
                return True  # 锁获取成功
            if count_circle >= 25:
                self.logger.warning('code:{} 有其他进程正在处理中'.format(code))
                return False  # 锁获取失败
            time.sleep(0.2)  # 睡眠 200 毫秒
            count_circle += 1
# 获取代收订单
class getOrderDs(BaseHandler):
    @tornado.web.authenticated
    async def post(self):

        # 查询角色的所有权限name
        role_permissions_sql = f"select name from permissions where status = 1 and id in ({self.current_user['permissions']})"
        role_permission_names_json = await self.query(role_permissions_sql)
        # 字典对象 转 集合
        role_permission_names = set(value for dictionary in role_permission_names_json for value in dictionary.values())

        data = json.loads(self.request.body)

        condition, time_create_between = await self.split_between_condition(data['serchData'], 'time_create')
        condition, time_success_between = await self.split_between_condition(condition, 'time_success')
        between = None
        if time_create_between:
            between = time_create_between
        if not between:
            between = time_success_between

        if not condition or not condition['code'] and not between:
            between = {'key': 'time_create', 'start': datetime.today().date(), 'end': datetime.now()}
        keys_count = ['amount', 'status','realpay','earn_merchant','earn_partner','earn_system']
        # data_r, total, count = await self.get_result('orders_ds', ['*'], keys_count, condition, between,
        #                                              data['size'], data['page'])

        # 获取顶级id下的所有订单
        top_partner_sql_part = ''
        partner_ids = []
        if 'top_partner_id' in condition.keys():
            partner_ids = await self.get_partners(condition['top_partner_id'])
            del condition['top_partner_id']
            if partner_ids:
                _partner_ids = ','.join(['%s'] * len(partner_ids))
                top_partner_sql_part = ' partner_id in ({partner_ids})'.format(partner_ids=_partner_ids)

        sql_part = ''
        values = []

        # 如果有金额范围查找
        amount_range = ''
        if 'amount_range' in condition.keys():
            if condition['amount_range'] == 1: # <=500
                amount_range = ' amount <= 500'
            elif condition['amount_range'] == 2: # 500-1000
                amount_range = ' amount <= 1000 and amount > 500'
            elif condition['amount_range'] == 3: # 1000-2000
                amount_range = ' amount <= 2000 and amount > 1000'
            elif condition['amount_range'] == 4: # 2000-5000
                amount_range = ' amount <= 5000 and amount > 2000'
            elif condition['amount_range'] == 5: # 5000-20000
                amount_range = ' amount <= 20000 and amount > 5000'
            elif condition['amount_range'] == 6: # 20000-50000
                amount_range = ' amount <= 50000 and amount > 20000'
            elif condition['amount_range'] == 7: # >50000
                amount_range = ' amount > 50000'
            del condition['amount_range']

        # 用户输入的金额范围
        # amount_range_input = condition['amount_range_new']  # 用户输入的金额范围

        # 用户输入的值可能是类似 "500-1000", "1000-2000" 这样的格式
        if 'amount_range_new' in condition.keys():
            # 去掉输入的首尾空格
            amount_range_input = condition['amount_range_new'].strip()

            # 解析用户输入
            if '-' in amount_range_input:  # 如果用户输入了一个范围（如 500-1000）
                amount_parts = amount_range_input.split('-')
                if len(amount_parts) == 2:
                    lower_limit = amount_parts[0].strip()  # 下限
                    upper_limit = amount_parts[1].strip()  # 上限
                    amount_range = f"amount <= {upper_limit} and amount > {lower_limit}"
                    del condition['amount_range_new']

        # 删除原始条件中的 'amount_range' 键
        # print(amount_range)

        if condition:
            for k in list(condition.keys()):
                if not condition[k] and condition[k] != 0:
                    condition.pop(k)
        if condition or between:
            sql_part += ' where '
        if condition:
            if 'status' in condition and condition['status'] == 5:
                sql_part += ' status in (1,2) and'
                del condition['status']
            where_key, where_val = await self.dict_to_and(condition)
            sql_part += ' {keys} '.format(keys=where_key)
            values += where_val
        if between:
            bt_key, bt_start, bt_end = await self.dict_to_between(between)
            if await self.calculate_date_diff(bt_start,bt_end):
                return await self.json_response(msg[10301])

            if condition:
                sql_part += " and " + bt_key
            else:
                sql_part += bt_key
            values += [bt_start, bt_end]

        if sql_part == '':
            if top_partner_sql_part:
                sql_part = ' where ' + top_partner_sql_part
                values = partner_ids
        else:
            if top_partner_sql_part:
                sql_part += ' and ' + top_partner_sql_part
                values.extend(partner_ids)

        # 如果有金额范围查找
        if sql_part == '':
            if amount_range:
                sql_part = ' where ' + amount_range
        else:
            if amount_range:
                sql_part += ' and ' + amount_range

        # 如果是推广账号，则过滤出推广账号下的码商订单
        if str(self.current_user['role_id']) == '19':
            tg_partners_ids = await self.get_partners_by_parent_id(self.current_user['parent_id'])
            if sql_part == '':
                sql_part = ' where partner_id in ({})'.format(tg_partners_ids)
            else:
                sql_part += ' and partner_id in ({})'.format(tg_partners_ids)

        # 获取所有数据总数
        sql = "select count(id) from orders_ds"
        sql += sql_part
        total = await self.query(sql, *values)
        if total:
            total = total[0]['count(id)']
        else:
            total = 0
        # 获取所有数据的指定key数据
        count = []
        # 直接查询数据库导致sql执行过长，这里进行优化，改成聚合查询，所以下面代码作废
        # if keys_count:
        #     sql = "select {keys} from orders_ds ".format(keys=await self.list_keys(keys_count))
        #     sql += sql_part
        #     count = await self.query(sql, *values)

        sql = """
              SELECT COALESCE(SUM(CASE WHEN status >= 3 THEN 1 ELSE 0 END), 0)                            AS successOrder, \
                     COALESCE(SUM(CASE WHEN status >= 3 THEN amount ELSE 0 END), 0)                       AS amount, \
                     COALESCE(SUM(CASE WHEN status >= 3 THEN realpay ELSE 0 END), 0)                      AS realpay, \
                     COALESCE(SUM(CASE WHEN status >= 3 THEN earn_merchant ELSE 0 END), 0)                AS earn_merchant, \
                     COALESCE(SUM(CASE WHEN status >= 3 THEN earn_partner ELSE 0 END), 0)                 AS earn_partner, \
                     COALESCE(SUM(CASE WHEN status >= 3 THEN earn_system ELSE 0 END), 0)                  AS earn_system, \
                     COALESCE(SUM(CASE WHEN status = -1 THEN 1 ELSE 0 END), 0)                            AS failOrder, \
                     COALESCE(SUM(CASE WHEN status NOT IN (-1) AND status < 3 THEN 1 ELSE 0 END), 0)      AS processing, \
                     COALESCE(SUM(CASE WHEN status NOT IN (-1) AND status < 3 THEN amount ELSE 0 END), \
                              0)                                                                          AS processing_amount, \
                     COALESCE(SUM(CASE WHEN channel_code = '1003' AND (payment_id IS NOT NULL AND payment_id != 0) AND (count_statics IS NOT NULL AND count_statics != '') AND count_statics NOT LIKE '%%PT-RTP-CPS-2002%%' AND status = -1 THEN 1 ELSE 0 END), 0) AS send_failed_err,
                     COALESCE(SUM(CASE WHEN channel_code = '1003' AND (payment_id IS NOT NULL AND payment_id != 0) AND count_statics LIKE '%%PT-RTP-CPS-2002%%' AND status = -1 THEN 1 ELSE 0 END), 0) AS send_success_unpaid,
                     COALESCE(SUM(CASE WHEN channel_code = '1003' AND (payment_id IS NOT NULL AND payment_id != 0) AND (count_statics IS NULL OR count_statics = '') THEN 1 ELSE 0 END), 0) AS not_sent
              FROM orders_ds
              """ + sql_part
        count = await self.query(sql, *values)
        # 2. 格式化统计数据为 字符串 "报错失败|成功未付|未发"
        default_stat = "失败A:0|失败B:0|未发:0"
        if count and isinstance(count, list) and len(count) > 0 and isinstance(count[0], dict):
            count_r = self.convert_decimal(count[0])
            raw = count[0]
            # 【拼装核心】：按照“标签:数值”格式拼接
            count_r['stat_1003'] = "失败A:{err}|失败B:{unpaid}|未发:{not_sent}".format(
                err=int(raw.get('send_failed_err', 0)),
                unpaid=int(raw.get('send_success_unpaid', 0)),
                not_sent=int(raw.get('not_sent', 0))
            )
        else:
            count_r = {
                'successOrder': 0,
                'amount': 0,
                'realpay': 0,
                'earn_merchant': 0,
                'earn_partner': 0,
                'earn_system': 0,
                'failOrder': 0,
                'processing': 0,
                'processing_amount': 0,
                'stat_1003': default_stat,
            }
        # 获取权限值
        allowAdminProfit = await self.redis.get('allowAdminProfit')
        if allowAdminProfit:
            allowAdminProfit = allowAdminProfit.split(',')
        # 权限控制
        userPermissions = False
        # 获取登录管理员id
        user_id = str(self.get_secure_cookie("user"),'utf-8')
        if allowAdminProfit:
            # 查看权限是否存在
            if user_id in allowAdminProfit:
                userPermissions = True

        # 获取分页数据
        # sql = "select * from {table}  where id in (select id from {table} ".format(table='orders_ds')
        # order_by = ') order by id desc '
        sql = "select * from orders_ds "
        order_by = ' order by id desc '
        sql += sql_part + order_by
        if data['size'] and data['page'] > -1:
            sql += 'limit %s offset %s'
            values += [data['size'], (data['page'] - 1) * data['size']]
        data_r = await self.query(sql, *values)
        # print('sql===', sql)

        for i in range(len(data_r)):
            if ("禁止查看代收订单结算金额" in role_permission_names):
                data_r[i]['realpay'] = '******'
            if ("禁止查看代收订单商户费率" in role_permission_names):
                data_r[i]['merchant_rate'] = '******'
            if ("禁止查看代收订单手续费" in role_permission_names):
                data_r[i]['poundage'] = '******'
            if ("禁止查看代收订单商代盈利" in role_permission_names):
                data_r[i]['earn_merchant'] = '******'
            if ("禁止查看代收订单码商盈利" in role_permission_names):
                data_r[i]['earn_partner'] = '******'
            if ("禁止查看代收订单平台利润" in role_permission_names):
                data_r[i]['earn_system'] = '******'

            count_statics_str = data_r[i].get('count_statics')

            if count_statics_str:
                try:
                    # 1. 尝试解析成 JSON
                    count_statics_data = json.loads(count_statics_str)
                    
                    # 2. 检查是否包含预期的计数键名
                    if isinstance(count_statics_data, dict) and ('download_count' in count_statics_data or 'next_count' in count_statics_data):
                        download_count = count_statics_data.get('download_count', 0)
                        next_count = count_statics_data.get('next_count', 0)
                        # 只有符合格式的数据，才进行格式化组合
                        data_r[i]['count_statics'] = f'下载次数: {download_count}, Next次数: {next_count}'
                    else:
                        data_r[i]['count_statics'] = count_statics_str
                        
                except (json.JSONDecodeError, TypeError):
                    # 如果解析 JSON 失败（说明存的是纯文本信息），直接原封不动显示原始数据
                    data_r[i]['count_statics'] = count_statics_str
            else:
                # 字段本身为空的情况
                data_r[i]['count_statics'] = '暂无数据'

        # count_r = {'failOrder': 0, 'successOrder': 0, 'processing': 0, 'amount': Decimal(0),
        #            'processing_amount': Decimal(0), 'realpay': Decimal(0), 'earn_merchant': Decimal(0),'earn_partner': Decimal(0), 'earn_system': Decimal(0)}
        # for i in count:
        #     if i['status'] >= 3:
        #         count_r['successOrder'] += 1
        #         count_r['amount'] += i['amount']
        #         if userPermissions is True:
        #             count_r['realpay'] += i['realpay']
        #             count_r['earn_merchant'] += i['earn_merchant']
        #             count_r['earn_partner'] += i['earn_partner']
        #             count_r['earn_system'] += i['earn_system']
        #
        #     elif i['status'] == -1:
        #         count_r['failOrder'] += 1
        #     else:
        #         count_r['processing'] += 1
        #         count_r['processing_amount'] += i['amount']



                # 获取所有数据的指定key数据
        # 没有权限赋值为******
        # if userPermissions is False:
        #     count_r['realpay'] = "******"
        #     count_r['earn_merchant'] = "******"
        #     count_r['earn_partner'] = "******"
        #     count_r['earn_system'] = "******"
        if ("禁止查看代收订单结算金额" in role_permission_names):
            count_r['realpay'] = "******"
        if ("禁止查看代收订单商代盈利" in role_permission_names):
            count_r['earn_merchant'] = "******"
        if ("禁止查看代收订单码商盈利" in role_permission_names):
            count_r['earn_partner'] = "******"
        if ("禁止查看代收订单平台利润" in role_permission_names):
            count_r['earn_system'] = "******"
        # 商户处理中的订单数
        # cdt = ""
        # value = []
        # if condition and "channel_code" in condition.keys():
        #     cdt = ' and channel_code = ' + str(condition['channel_code'])
        # if between:
        #     cdt += ' and {key} between %s and %s'.format(key=between['key'])
        #     value = [between['start'], between['end']]
        # if not between:
        #     between = dict()
        #     between['key'] = 'time_create'
        #     between['start'] = datetime.today().date()
        #     between['end'] = datetime.now()
        #     cdt += ' and {key} between %s and %s'.format(key=between['key'])
        #     value = [between['start'], between['end']]

        merchant_processing = []

        # sql_p = """
        #     select merchant_id,count(1) as cnt,sum(amount) as total
        #     from orders_ds where status < 3 and status > -1  {cdt}
        #     group by merchant_id order by cnt desc;
        # """.format(cdt=cdt)
        # merchant_processing = await self.query(sql_p, *value)

        merchant_finish = []

        # sql_p = """
        #     select merchant_id,count(1) as cnt,sum(amount) as total
        #     from orders_ds where status in (3, 4)  {cdt}
        #     group by merchant_id order by cnt desc;
        # """.format(cdt=cdt)
        # merchant_finish = await self.query(sql_p, *value)

        result = dict(code=20000, data=data_r, total=total, count=count_r, merchant_processing=merchant_processing, merchant_finish=merchant_finish, msg='获取成功')
        return await self.json_response(result)

    # 增加一个方法，把字符串改为数字类型
    def convert_decimal(self, data):
        if isinstance(data, dict):
            return {k: self.convert_decimal(v) for k, v in data.items()}
        elif isinstance(data, list):
            return [self.convert_decimal(i) for i in data]
        elif isinstance(data, Decimal):
            # 如果有小数就float，没有小数就int
            return float(data) if data % 1 != 0 else int(data)
        return data

# 获取代收订单处理中或成交额数据
class getDSMerchantFinishOrProcessing(BaseHandler):
    @tornado.web.authenticated
    async def post(self):

        data = json.loads(self.request.body)

        merchant_finish_ = False
        if 'merchant_finish' in data and data['merchant_finish']:
            merchant_finish_ = True

        processing_ = False
        if 'processing' in data and data['processing']:
            processing_ = True

        condition, between = await self.split_between_condition(data['serchData'], 'time_create')
        if not between:
            condition, between = await self.split_between_condition(condition, 'time_success')

        # 商户处理中的订单数
        cdt = ""
        value = []
        if condition and "channel_code" in condition.keys() and condition['channel_code']:
            cdt = ' and channel_code = ' + str(condition['channel_code'])
        if between:
            cdt += ' and {key} between %s and %s'.format(key=between['key'])
            value = [between['start'], between['end']]
        if not between:
            between = dict()
            between['key'] = 'time_create'
            between['start'] = datetime.today().date()
            between['end'] = datetime.now()
            cdt += ' and {key} between %s and %s'.format(key=between['key'])
            value = [between['start'], between['end']]
        merchant_processing = []
        if processing_:
            sql_p = """
                select merchant_id,count(1) as cnt,sum(amount) as total 
                from orders_ds where status < 3 and status > -1  {cdt}
                group by merchant_id order by cnt desc; 
            """.format(cdt=cdt)
            merchant_processing = await self.query(sql_p, *value)

        merchant_finish = []
        if merchant_finish_:
            sql_p = """
                select merchant_id,count(1) as cnt,sum(amount) as total 
                from orders_ds where status in (3, 4)  {cdt}
                group by merchant_id order by cnt desc; 
            """.format(cdt=cdt)
            merchant_finish = await self.query(sql_p, *value)
        result = dict(code=20000, merchant_processing=merchant_processing,merchant_finish=merchant_finish, msg='获取成功')
        return await self.json_response(result)

# 获取代收CD订单
class getOrderDsCd(BaseHandler):
    @tornado.web.authenticated
    # 获取 partner_id 的顶级父节点的函数
    async def get_top_level_parent(self, partner_id):
        sql = """
        SELECT parent
        FROM partner_tree
        WHERE child = %s
        ORDER BY distance DESC
        LIMIT 1
        """
        values = (partner_id,)
        result = await self.query(sql, *values)
        if result:
            return result[0]['parent']  # 返回查询结果中的 parent（顶级父节点）
        else:
            return None  # 如果没有找到，返回 None
    # 获取 partner_name 的函数
    async def get_partner_name(self, partner_id):
        sql = "SELECT name FROM partner WHERE id = %s"
        values = (partner_id,)
        result = await self.query(sql, *values)
        if result:
            return result[0]['name']  # 返回查询结果中的 partner_name
        else:
            return None  # 如果没有找到，返回 None
    async def post(self):
        data = json.loads(self.request.body)
        condition, between = await self.split_between_condition(data['serchData'], 'time_create')

        condition, time_create_between = await self.split_between_condition(data['serchData'], 'time_create')
        condition, time_success_between = await self.split_between_condition(condition, 'time_success')
        condition, created_at_between = await self.split_between_condition(condition, 'created_at')
        between = None
        if time_create_between:
            between = time_create_between
        if not between:
            between = time_success_between
        if not between:
            between = created_at_between

        if not condition or not condition['code'] and not between:
            between = {'key': 'time_create', 'start': datetime.today().date(), 'end': datetime.now()}
        keys_count = ['amount', 'status','realpay','earn_merchant','earn_partner','earn_system']
        # data_r, total, count = await self.get_result('orders_ds', ['*'], keys_count, condition, between,
        #                                              data['size'], data['page'])

        # 获取顶级id下的所有订单
        top_partner_sql_part = ''
        partner_ids = []
        if 'top_partner_id' in condition.keys():
            partner_ids = await self.get_partners(condition['top_partner_id'])
            del condition['top_partner_id']
            if partner_ids:
                _partner_ids = ','.join(['%s'] * len(partner_ids))
                top_partner_sql_part = ' partner_id in ({partner_ids})'.format(partner_ids=_partner_ids)

        sql_part = ''
        values = []

        if condition:
            for k in list(condition.keys()):
                if not condition[k] and condition[k] != 0:
                    condition.pop(k)
        if condition or between:
            sql_part += ' where '
        if condition:
            if 'status' in condition and condition['status'] == 5:
                sql_part += ' status in (1,2) and'
                del condition['status']
            where_key, where_val = await self.dict_to_and(condition)
            sql_part += ' {keys} '.format(keys=where_key)
            values += where_val
        if between:
            bt_key, bt_start, bt_end = await self.dict_to_between(between)
            if await self.calculate_date_diff(bt_start,bt_end):
                return await self.json_response(msg[10301])

            if condition:
                sql_part += " and " + bt_key
            else:
                sql_part += bt_key
            values += [bt_start, bt_end]

        if sql_part == '':
            if top_partner_sql_part:
                sql_part = ' where ' + top_partner_sql_part
                values = partner_ids
        else:
            if top_partner_sql_part:
                sql_part += ' and ' + top_partner_sql_part
                values.extend(partner_ids)

        # 如果是推广账号，则过滤出推广账号下的码商订单
        if str(self.current_user['role_id']) == '19':
            tg_partners_ids = await self.get_partners_by_parent_id(self.current_user['parent_id'])
            if sql_part == '':
                sql_part = ' where partner_id in ({})'.format(tg_partners_ids)
            else:
                sql_part += ' and partner_id in ({})'.format(tg_partners_ids)

        # 获取所有数据总数
        sql = "select count(id) from orders_cd"
        sql += sql_part
        total = await self.query(sql, *values)
        if total:
            total = total[0]['count(id)']
        else:
            total = 0
        # 获取所有数据的指定key数据
        count = []
        if keys_count:
            sql = "select {keys} from orders_cd ".format(keys=await self.list_keys(keys_count))
            sql += sql_part
            count = await self.query(sql, *values)
        # 获取权限值  取缔allowAdminProfit识别权限，改为代收代付权限
        # allowAdminProfit = await self.redis.get('allowAdminProfit')
        # if allowAdminProfit:
        #     allowAdminProfit = allowAdminProfit.split(',')
        # # 权限控制
        # userPermissions = False
        # # 获取登录管理员id
        # user_id = str(self.get_secure_cookie("user"),'utf-8')
        # if allowAdminProfit:
        #     # 查看权限是否存在
        #     if user_id in allowAdminProfit:
        #         userPermissions = True

        # 获取分页数据
        # sql = "select * from {table}  where id in (select id from {table} ".format(table='orders_cd')
        # order_by = ') order by id desc '
        sql = "select * from orders_cd "
        order_by = ' order by id desc '
        sql += sql_part + order_by
        if data['size'] and data['page'] > -1:
            sql += 'limit %s offset %s'
            values += [data['size'], (data['page'] - 1) * data['size']]
        data_r = await self.query(sql, *values)
        # 为每个记录添加 partner_name 字段
        for record in data_r:
            # 先获取最大的 parent_id（顶级父节点）
            top_level_partner_id = await self.get_top_level_parent(record['partner_id'])
            
            # 如果获取到了顶级父节点，则用该节点的 ID 查询 partner_name
            if top_level_partner_id:
                record['top_level_partner_id'] = top_level_partner_id
                record['top_level_partner_id_name'] = await self.get_partner_name(top_level_partner_id)
            else:
                record['top_level_partner_id'] = None
                record['top_level_partner_id_name'] = None
        count_r = {'failOrder': 0, 'successOrder': 0, 'processing': 0, 'amount': Decimal(0),
                   'processing_amount': Decimal(0), 'realpay': Decimal(0), 'earn_merchant': Decimal(0),'earn_partner': Decimal(0), 'earn_system': Decimal(0)}

        for i in count:
            if i['status'] >= 3:
                count_r['successOrder'] += 1
                count_r['amount'] += i['amount']
                count_r['realpay'] += i['realpay']
                count_r['earn_merchant'] += i['earn_merchant']
                count_r['earn_partner'] += i['earn_partner']
                count_r['earn_system'] += i['earn_system']

            elif i['status'] == -1:
                count_r['failOrder'] += 1
            else:
                count_r['processing'] += 1
                count_r['processing_amount'] += i['amount']
                # 获取所有数据的指定key数据
        # 没有权限赋值为******
        # if userPermissions is False:
        #     count_r['realpay'] = "******"
        #     count_r['earn_merchant'] = "******"
        #     count_r['earn_partner'] = "******"
        #     count_r['earn_system'] = "******"
        # 查询角色的所有权限name
        role_permissions_sql = f"select name from permissions where status = 1 and id in ({self.current_user['permissions']})"
        role_permission_names_json = await self.query(role_permissions_sql)
        # 字典对象 转 集合
        role_permission_names = set(value for dictionary in role_permission_names_json for value in dictionary.values())
        if ("禁止查看代收订单结算金额" in role_permission_names):
            count_r['realpay'] = "******"
        if ("禁止查看代收订单商代盈利" in role_permission_names):
            count_r['earn_merchant'] = "******"
        if ("禁止查看代收订单码商盈利" in role_permission_names):
            count_r['earn_partner'] = "******"
        if ("禁止查看代收订单平台利润" in role_permission_names):
            count_r['earn_system'] = "******"
        # 商户处理中的订单数
        # cdt = ""
        # value = []
        # if condition and "channel_code" in condition.keys():
        #     cdt = ' and channel_code = ' + str(condition['channel_code'])
        # if between:
        #     cdt += 'and {key} between %s and %s'.format(key=between['key'])
        #     value = [between['start'], between['end']]
        # if not between:
        #     between = dict()
        #     between['key'] = 'time_create'
        #     between['start'] = datetime.today().date()
        #     between['end'] = datetime.now()
        #     cdt += 'and {key} between %s and %s'.format(key=between['key'])
        #     value = [between['start'], between['end']]
        # sql_p = """
        #     select merchant_id,count(1) as cnt,sum(amount) as total
        #     from orders_cd where status < 3 and status > -1  {cdt}
        #     group by merchant_id order by cnt desc;
        # """.format(cdt=cdt)
        # merchant_processing = await self.query(sql_p, *value)
        merchant_processing = []
        result = dict(code=20000, data=data_r, total=total, count=count_r, merchant_processing=merchant_processing, msg='获取成功')
        return await self.json_response(result)

class getDSCDProcessing(BaseHandler):
    async def post(self):
        data = json.loads(self.request.body)
        condition, between = await self.split_between_condition(data['serchData'], 'time_create')
        if not between:
            condition, between = await self.split_between_condition(condition, 'time_success')
        if not between:  # 这里先检查 between
            condition, between = await self.split_between_condition(condition, 'created_at')  # 添加对 created_at 的处理
        # 商户处理中的订单数
        cdt = ""
        value = []
        if condition and "channel_code" in condition.keys() and condition['channel_code']:
            cdt = ' and channel_code = ' + str(condition['channel_code'])
        if between:
            cdt += 'and {key} between %s and %s'.format(key=between['key'])
            value = [between['start'], between['end']]
        if not between:
            between = dict()
            between['key'] = 'time_create'
            between['start'] = datetime.today().date()
            between['end'] = datetime.now()
            cdt += 'and {key} between %s and %s'.format(key=between['key'])
            value = [between['start'], between['end']]
        sql_p = """
            select merchant_id,count(1) as cnt,sum(amount) as total 
            from orders_cd where status < 3 and status > -1  {cdt}
            group by merchant_id order by cnt desc; 
        """.format(cdt=cdt)
        merchant_processing = await self.query(sql_p, *value)
        result = dict(code=20000, merchant_processing=merchant_processing,msg='获取成功')
        return await self.json_response(result)

# 代收补单
class handleOrder(BaseHandler):
    @tornado.web.authenticated
    async def post(self):
        data = json.loads(self.request.body)
        # 查找码商
        sql_select_partner = """select partner_id,upi from payment where id=%s"""
        # 查找订单
        sql_select_order = """select * from orders_ds where code=%s and status in (-1,1,2) order by id desc limit 1"""
        # 查询银行记录
        sql_select_bank_record = """select * from bank_record where utr=%s and amount=%s and callback=0 and trade_type=1 and invalid=0 order by id desc limit 1"""
        # 修改银行记录
        sql_update_bank_record = """update bank_record set callback=1,order_code=%s where id=%s and callback=0"""
        # 商户代理费率
        sql_select_rates_merchant = """select mid as id,rate from (select @orgId mid, (select @orgId:=pid from merchant 
                                            where id=@orgId) pid from (select @orgId:=%s) vars,merchant) t inner join 
                                            merchant_channel m on m.merchant_id=mid and m.code=%s where m.merchant_id is not null  order by m.merchant_id desc"""
        # 码商代理费率
        sql_select_rates_partner = """select rates from channel where code=%s"""
        # 更新系统余额
        sql_update_payment = """update payment set sys_balance=sys_balance+%s where id=%s"""
        # 更新订单
        sql_update_order = """update orders_ds set earn_merchant=%s,earn_partner=%s,earn_system=%s,partner_id=%s,
                                payment_id=%s,utr=%s,time_success=%s,status=3,upi=%s,trans_id=%s where code=%s and status in (-1,1,2) limit 1"""

        # 使用锁，5s使用自旋锁, 防止取消的同时回调
        count_circle = 0
        while True:
            busy_key = 'order_success_busy_{code}'.format(code=data['code'])
            if await self.redis.setnx(busy_key, 1):
                await self.redis.expire(busy_key, 10)
                break
            if count_circle >= 25:
                self.logger.warning('code:{}有其他进程正在处理中'.format(data['code']))
                return await self.json_response(msg[10032])
            time.sleep(0.2)
            count_circle = count_circle + 1

        async with self.application.db.acquire() as conn:
            async with conn.cursor(DictCursor) as cur:
                try:
                    # 查询订单
                    if not await cur.execute(sql_select_order, data['code']):
                        return await self.json_response(msg[10007])
                    order = (await cur.fetchall())[0]
                    code = order['code']

                    # ==================== 变更开始：新增 UTR 并发/频率锁====================
                    # 定义 UTR 锁的键名和过期时间
                    UTR_LOCK_PREFIX = "utr_submission_lock:"
                    UTR_LOCK_EXPIRY_SECONDS = 10 # 锁的有效期，10秒
                    utr_lock_key = f'{UTR_LOCK_PREFIX}{data['utr']}:{code}'
                    # 先使用 setnx 尝试获取锁，如果成功，再使用 expire 设置过期时间
                    got_utr_lock = await self.redis.setnx(utr_lock_key, 1)
                    
                    if got_utr_lock: # 只有当成功获取锁时，才设置过期时间
                        await self.redis.expire(utr_lock_key, UTR_LOCK_EXPIRY_SECONDS)
                        self.logger.info(f'订单：{code}，上传的卡密信息：{data['utr']} 提交频率锁获取成功并设置过期时间。')
                    else: # 未能获取锁 (键已存在且未过期)
                        self.logger.warning(f'UTR {data['utr']} 提交过于频繁或正在被其他请求处理，放弃操作。')
                        self.logger.info(f"订单：{code}，上传的卡密信息：{data['utr']} UTR submitted too frequently or already processing.")
                        return await self.json_response(msg[10032]) # UTR 提交频率过高/处理中
                    
                    trans_id_param = data['trans_id']
                    # 1. 检查是否为非空字符串
                    if not trans_id_param or not isinstance(trans_id_param, str):
                        self.logger.info("错误：交易ID不能为空或非字符串类型。")
                    else:
                        # 2. 检查长度
                        if len(trans_id_param) > 50:
                            self.logger.info(f"错误：交易ID长度超过50个字符。当前长度为：{len(trans_id_param)}")
                            return await self.json_response(msg[10330])
                        else:
                            # 3. 检查是否包含特殊字符
                            # 正则表达式：只允许字母、数字、下划线和连字符
                            pattern = re.compile(r'^[a-zA-Z0-9_-]+$')
                            if not pattern.match(trans_id_param):
                                self.logger.info(f"错误：交易ID包含特殊或非法字符。无效的ID为: {trans_id_param}")
                                return await self.json_response(msg[10330])
                            else:
                                # --- 验证通过，执行业务逻辑 ---
                                self.logger.info(f"交易ID '{trans_id_param}' 格式有效，开始处理补单请求...")
                                
                    if 'trans_id' in data.keys() and data['trans_id']:
                        count_circle = 0
                        while True:
                            busy_key = 'success_busy_{trans_id}'.format(trans_id=data['trans_id'])
                            if await self.redis.setnx(busy_key, 1):
                                await self.redis.expire(busy_key, 10)
                                break
                            if count_circle >= 10:
                                self.logger.warning(
                                    'trans_id:{trans_id}Do not operate frequently'.format(trans_id=data['trans_id']))
                                res = dict(code=99, msg='Do not operate frequently')
                                return await self.json_response(res)
                            time.sleep(0.2)
                            count_circle = count_circle + 1


                    # ----------------- 交易ID重复校验逻辑 -----------------
                    if 'trans_id' in data.keys() and data['trans_id']:
                        sql_check_trans_id = """
                            SELECT code FROM orders_ds WHERE trans_id=%s AND id != %s LIMIT 1
                        """
                        # 这里的 _order[0]['id'] 是当前找到的订单的ID
                        existing_order = await self.query(sql_check_trans_id, data['trans_id'], order['id'])
                        
                        # 打印即将执行的查询日志，包含SQL和参数
                        self.logger.info(
                            f"新增：交易ID重复校验查询 | SQL: {sql_check_trans_id.strip()} | 参数: ({data['trans_id']}, {order['id']})"
                        )
                        
                        # 如果查询结果不为空，则说明有其他订单已使用此交易ID
                        if existing_order:
                            self.logger.warning(f"交易ID {data['trans_id']} 已被其他订单使用。冲突订单号: {existing_order[0]['code']}")
                            # 返回一个错误提示
                            return await self.json_response(msg[10330])
                        
                    # ==================== 变更结束 ====================

                    amount = order['amount']
                    # 查询银行记录
                    if not await cur.execute(sql_select_bank_record, (data['utr'], amount)):
                        return await self.json_response(msg[10014])
                    bank_record = (await cur.fetchall())[0]
                    payment_id = bank_record['payment_id']
                    trans_id = bank_record['trans_id']
                    self.logger.info(f"交易ID trans_id_param: {trans_id_param} , trans_id: {trans_id}")
                    if trans_id_param and trans_id and trans_id_param != trans_id:
                        self.logger.warning(f"交易ID trans_id_param: {trans_id_param} , trans_id: {trans_id}")
                        # 返回一个错误提示
                        return await self.json_response(msg[10330])
                    
                    # ----------------- 交易ID重复校验逻辑 -----------------
                    if trans_id:
                        sql_check_trans_id = """
                            SELECT code FROM orders_ds WHERE trans_id=%s AND id != %s LIMIT 1
                        """
                        # 这里的 _order[0]['id'] 是当前找到的订单的ID
                        existing_order = await self.query(sql_check_trans_id, trans_id, order['id'])
                        
                        # 打印即将执行的查询日志，包含SQL和参数
                        self.logger.info(
                            f"新增：交易ID重复校验查询 | SQL: {sql_check_trans_id.strip()} | 参数: ({trans_id}, {order['id']})"
                        )
                        
                        # 如果查询结果不为空，则说明有其他订单已使用此交易ID
                        if existing_order:
                            self.logger.warning(f"交易ID {trans_id} 已被其他订单使用。冲突订单号: {existing_order[0]['code']}")
                            # 返回一个错误提示
                            return await self.json_response(msg[10330])
                    # ==================== 变更结束 ====================

                    # 修改银行记录
                    if not await cur.execute(sql_update_bank_record, (code, bank_record['id'])):
                        await conn.rollback()
                        return await self.json_response(msg[10014])
                    # 码商查询
                    if not await cur.execute(sql_select_partner, payment_id):
                        await conn.rollback()
                        return await self.json_response(msg[10007])
                    _payment = (await cur.fetchall())[0]
                    partner_id = _payment['partner_id']
                    # 退掉额外扣款
                    if bank_record['ew_code']:
                        if not await self.change_balance(conn, cur, 'partner', partner_id, amount, bank_record['ew_code'], 0):
                            return await self.json_response(msg[10015])
                    # 补扣码商(非自身订单、过期订单)
                    if not order['partner_id'] == partner_id or order['status'] == -1:
                        if not await self.change_balance(conn, cur, 'partner', partner_id, -amount, code, 0):
                            return await self.json_response(msg[10015])
                    # 非自身订单并且未过期退款给旧码商
                    if not order['partner_id'] == partner_id and not order['status'] == -1:
                        if not await self.change_balance(conn, cur, 'partner', order['partner_id'], amount,
                                                         code, 0):
                            return await self.json_response(msg[10007])
                    # 增加商户余额
                    if not await self.change_balance(conn, cur, 'merchant', order['merchant_id'], order['realpay'],
                                                     code, 0):
                        return await self.json_response(msg[10007])
                    # 商户代理费用
                    earn_merchant = Decimal(0)
                    if order['earn_merchant'] > 0:
                        if not await cur.execute(sql_select_rates_merchant,
                                                 (order['merchant_id'], order['channel_code'])):
                            await conn.rollback()
                            return await self.json_response(msg[10007])
                        merchant_rates = (await cur.fetchall())
                        self.logger.info('订单号{code}商户汇率{merchant_rates}'.format(code=code, merchant_rates=merchant_rates))
                        for k, v in enumerate(merchant_rates):
                            if not k == 0 and v['rate']:
                                _amount = amount * (merchant_rates[k - 1]['rate'] - v['rate'])
                                self.logger.info('订单号{code}手续费{_amount}'.format(code=code, _amount=_amount))
                                if _amount < 0:
                                    await conn.rollback()
                                    return await self.json_response(msg[10013])
                                if not await self.change_balance(conn, cur, 'merchant', v['id'], _amount, code, 3):
                                    return await self.json_response(msg[10007])
                                earn_merchant += _amount
                    # 增加码商佣金
                    if not await self.change_balance(conn, cur, 'partner', partner_id, order['earn_partner_self'], code,
                                                     3):
                        return await self.json_response(msg[10007])
                    # 增加码商代理佣金
                    earn_partner = order['earn_partner_self']
                    if not await cur.execute(sql_select_rates_partner, order['channel_code']):
                        return await self.json_response(msg[10007])
                    rates = (await cur.fetchall())[0]['rates'].split(',')
                    _partner_id = partner_id
                    for i in range(len(rates)):
                        partner = await self.get_result_by_condition('partner', ['pid'], {'id': _partner_id})
                        if not partner['pid']:
                            break
                        _partner_id = partner['pid']
                        _amount = amount * Decimal(rates[i])
                        if not await self.change_balance(conn, cur, 'partner', _partner_id, _amount, code, 3):
                            return await self.json_response(msg[10007])
                        earn_partner += _amount
                    # 系统盈利
                    earn_system = order['poundage'] - earn_merchant - earn_partner
                    self.logger.info('订单号{code}系统盈利{earn_system}'.format(code=code, earn_system=earn_system))
                    if earn_system < 0:
                        await conn.rollback()
                        return await self.json_response(msg[10013])
                    # 修改卡系统余额
                    if not await cur.execute(sql_update_payment, (amount, payment_id)):
                        await conn.rollback()
                        return await self.json_response(msg[10007])
                    # 修改订单状态
                    time_now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    if not await cur.execute(sql_update_order, (earn_merchant, earn_partner, earn_system, partner_id,
                                                                payment_id, data['utr'], time_now, _payment['upi'], trans_id, code)):
                        await conn.rollback()
                        return await self.json_response(msg[10007])
                    self.logger.info('更新订单状态%s' % cur._last_executed)
                    if not await self.create_result('operate', {'type': 11, 'admin_id': self.current_user['id'], 'ip': await self.get_ip()}):
                        await conn.rollback()
                        self.logger.warning(cur._last_executed)
                        return await self.json_response(msg[10007])
                except Exception as e:
                    self.logger.warning('确认订单失败,code={code},异常={e}'.format(code=data['code'], e=e))
                    await conn.rollback()
                    return await self.json_response(msg[10007])
                else:
                    await conn.commit()
                    await self.redis.publish('order_notify', code)
        result = dict(code=20000, msg='补单成功')
        return await self.json_response(result)

# 代收补单-from-third
class  handleOrderFromThird(BaseHandler):
    @tornado.web.authenticated
    async def post(self):
        data = json.loads(self.request.body)
        code = data['code']
        utr = data.get('utr', '')
        self.logger.info(f"Starting order_success_ds_third with code: {code}, UTR: {utr}")
        # 查找订单
        sql_select_order = """select * from orders_ds where code=%s and status in (0,-1,1,2) order by id desc limit 1"""
        # 商户代理费率
        sql_select_rates_merchant = """select mid as id,rate from (select @orgId mid, (select @orgId:=pid from merchant 
                                    where id=@orgId) pid from (select @orgId:=%s) vars,merchant) t inner join 
                                    merchant_channel m on m.merchant_id=mid and m.code=%s where m.merchant_id is not null  order by m.merchant_id desc"""
        # 更新订单utr
        sql_update_order_utr = """update orders_ds set utr=%s where id=%s and status in (0,-1,1,2) limit 1"""
        # 更新订单
        sql_update_order = """update orders_ds set earn_merchant=%s, earn_partner=%s, earn_system=%s, partner_id=%s,
                                        payment_id=%s, utr=%s, time_success=%s, status=3, upi=%s, trans_id=%s where code=%s and status in (0,-1,1,2) limit 1"""

        # 使用锁，5s使用自旋锁, 防止取消的同时回调
        count_circle = 0
        while True:
            busy_key = 'order_success_busy_{code}'.format(code=code)
            if await self.redis.setnx(busy_key, 1):
                await self.redis.expire(busy_key, 10)
                self.logger.warning(f"Lock acquired for code: {code}")
                break
            if count_circle >= 25:
                self.logger.warning(f"Do not operate frequently {code}")
                return dict(code=99, msg='Do not operate frequently')
            time.sleep(0.2)
            count_circle += 1

        async with self.application.db.acquire() as conn:
            async with conn.cursor(DictCursor) as cur:
                try:
                    # 查询订单
                    self.logger.info(f"Executing SQL: {sql_select_order} with code: {code}")
                    if not await cur.execute(sql_select_order, code):
                        self.logger.warning(f"No order found or timeout for code: {code}")
                        await conn.rollback()
                        return await self.json_response(msg[10007])
                    order = (await cur.fetchall())[0]
                    self.logger.info(f"Order found: {order}")
                    amount = order['amount']
                    third_party_name = order['third_party_name']
                    third_party_order_number = order['third_party_order_number']
                    partner_id = None
                    payment_id = None

                    sql_t = 'SELECT merchant_id, `key`, `key2`, `key3`, pay_url, name, channel_code, query_url, notify_ip FROM otherpay WHERE name = %s'
                    r_t = await self.query(sql_t, third_party_name)

                    mc_key = r_t[0]['key']
                    mc_key2 = r_t[0]['key2']
                    private_key = r_t[0]['key3']
                    mer_id = r_t[0]['merchant_id']
                    query_url = r_t[0]['query_url']
                    # 校验utr格式
                    if not bool(re.match(r'^\d{10,11}$', utr)):
                        self.logger.warning(f"补单手机号格式错误: {code} {utr}")
                        await conn.rollback()
                        return await self.json_response(msg[10231])

                    duplicate_lookup = None
                    if third_party_name != 'easypay':
                        duplicate_lookup = build_third_duplicate_lookup_payload(
                            third_party_name=third_party_name,
                            utr=utr,
                        )
                        duplicate_sql = (
                            f"select id from orders_ds where {duplicate_lookup['field']}=%s "
                            "and status in (3, 4)"
                        )
                        if duplicate_lookup['value'] and await self.query(duplicate_sql, duplicate_lookup['value']):
                            self.logger.warning(
                                f"此{duplicate_lookup['field']}在代收订单里已存在且已回调: "
                                f"{code}, value={duplicate_lookup['value']}"
                            )
                            await conn.rollback()
                            return await self.json_response(msg[duplicate_lookup['message_key']])

                    # 更新utr
                    if order["utr"] != utr:
                         if not await cur.execute(sql_update_order_utr, (utr, order["id"])):
                            self.logger.warning(f"Failed to update UTR for code: {code}")
                            await conn.rollback()
                            return await self.json_response(msg[10007])
                         self.logger.info(f"update utr, order code {code}, utr: old {order['utr']} new {utr}")

                    # 把 admin 输入的 utr 挂到 self 上，供需要用它做响应校验的 query 函数读取
                    # （handler 事务里的 UPDATE 尚未 commit，query 函数里用 self.query 新连接读不到最新值）
                    self._easypay_admin_utr = utr

                    # 根据 third_party_name 动态生成查询函数的名称
                    query_func_name = f'query_{third_party_name}_order'

                    # 使用 globals() 获取当前作用域中的函数，检查是否存在对应的函数
                    query_third_func = globals().get(query_func_name)
                    self.logger.info(f"函数名称: {query_third_func}")
                    # 如果对应的查询函数不存在，则记录错误并返回相应的错误信息
                    if not query_third_func:
                        await conn.rollback()
                        self.logger.error('未定义对应第三方支付的查询方法： {}'.format(third_party_name))
                        return await self.json_response(msg[10007])
                         
                    # 判断需要填写utr的三方
                    if third_party_name in ['snakepay', 'hkpay', 'skpay']:
                        if not order['utr']:
                            self.logger.warning(f"需要填写utr的三方: {third_party_name}")
                            await conn.rollback()
                            return await self.json_response(msg[10226])

                    # 如果查询函数存在，调用该函数进行查询
                    response = await query_third_func(self, mer_id, code, mc_key, mc_key2, query_url, third_party_name, private_key, third_party_order_number)

                    # 如果查询没有返回有效响应，则返回错误信息
                    # if not response:
                    # 更严格地判断查询结果
                    if response is not True:
                        self.logger.warning(f"[{third_party_name}] 查询未成功或状态异常，停止补单流程")
                        await conn.rollback()
                        return await self.json_response(msg[10007])

                    query_result = None
                    trans_id = order.get('trans_id')
                    if third_party_name == 'easypay':
                        query_result = getattr(self, '_easypay_query_result', None) or {}
                        duplicate_lookup = build_third_duplicate_lookup_payload(
                            third_party_name=third_party_name,
                            utr=utr,
                            query_result=query_result,
                        )
                        if not duplicate_lookup:
                            self.logger.warning(
                                f"[{third_party_name}] 账单详情缺少 transactionId，停止补单流程, code={code}"
                            )
                            await conn.rollback()
                            return await self.json_response(msg[10007])
                        duplicate_sql = (
                            f"select id from orders_ds where {duplicate_lookup['field']}=%s "
                            "and status in (3, 4)"
                        )
                        if await self.query(duplicate_sql, duplicate_lookup['value']):
                            self.logger.warning(
                                f"[{third_party_name}] 交易ID已存在且已回调: "
                                f"{code}, trans_id={duplicate_lookup['value']}"
                            )
                            await conn.rollback()
                            return await self.json_response(msg[duplicate_lookup['message_key']])
                        trans_id = duplicate_lookup['value']

                    # 增加商户余额
                    self.logger.info(f"Increasing merchant balance for merchant_id: {order['merchant_id']}, amount: {order['realpay']}")
                    if not await self.change_balance(conn, cur, 'merchant', order['merchant_id'], order['realpay'], code, 0):
                        self.logger.warning("Failed to increase merchant balance.")
                        # await conn.rollback()
                        return await self.json_response(msg[10007])

                    # 商户代理费用
                    earn_merchant = Decimal(0)
                    if order['earn_merchant'] > 0:
                        self.logger.info("Calculating merchant agent fees.")
                        if not await cur.execute(sql_select_rates_merchant, (order['merchant_id'], order['channel_code'])):
                            self.logger.warning("Failed to fetch merchant rates.")
                            await conn.rollback()
                            return await self.json_response(msg[10007])
                        merchant_rates = await cur.fetchall()
                        self.logger.info(f"Merchant rates fetched: {merchant_rates}")
                        for k, v in enumerate(merchant_rates):
                            if k != 0 and v['rate']:
                                _amount = amount * (merchant_rates[k - 1]['rate'] - v['rate'])
                                if _amount < 0:
                                    self.logger.warning(f"Invalid amount calculated: {_amount}")
                                    await conn.rollback()
                                    return await self.json_response(msg[10007])
                                if not await self.change_balance(conn, cur, 'merchant', v['id'], _amount, code, 3):
                                    self.logger.warning(f"Failed to update balance for merchant {v['id']} with amount {_amount}.")
                                    # await conn.rollback()
                                    return await self.json_response(msg[10007])
                                earn_merchant += _amount
                        self.logger.info(f"Total merchant earnings: {earn_merchant}")

                    # 系统盈利
                    earn_system = order['poundage'] - earn_merchant
                    if earn_system < 0:
                        self.logger.warning(f"System earnings calculation error, earnings: {earn_system}")
                        await conn.rollback()
                        return await self.json_response(msg[10007])

                    # 修改订单状态
                    time_now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    self.logger.info(f"Updating order status with earn_merchant: {earn_merchant}, earn_system: {earn_system}, code: {code}")
                    if not await cur.execute(sql_update_order, (earn_merchant, 0, earn_system, partner_id, payment_id, utr, time_now, order['upi'], trans_id, code)):
                        self.logger.warning("Failed to update order status.")
                        await conn.rollback()
                        return await self.json_response(msg[10007])
                    self.logger.info(f"Order status updated successfully for code: {code}")

                except Exception as e:
                    # self.logger.info(f"Exception occurred: {e}")
                    self.logger.info("Exception occurred:\n" + traceback.format_exc())
                    self.logger.warning(f"确认订单失败, code={code}, 异常={e}")
                    await conn.rollback()
                    return await self.json_response(msg[10007])
                else:
                    await conn.commit()
                    self.logger.info(f"通知订单, code={code}")
                    await self.redis.publish('order_notify', code)
                    result = dict(code=20000, msg='补单成功')
                    return await self.json_response(result)


class updateCd(BaseHandler):
    @tornado.web.authenticated
    async def post(self):
        data = json.loads(self.request.body)
        code = data.get('code')  # 获取传递的参数 code
        cd_memo = data.get('cd_memo')  # 获取查单备注
        status = data.get('cd_status')  # 获取状态
        utr = data.get('utr') 
        operator = self.current_user['id']
        ip_address = await self.get_ip()  # 获取操作的IP地址

        # 查找订单
        sql_select_order = """select * from orders_cd where code=%s order by id desc limit 1"""
        
        if status == 1:
            # 更新订单的 SQL 语句，增加对 `cd_memo` 和 `cd_status` 的更新
            sql_update_order = """update orders_cd set is_cd=1, admin_id=%s, cd_memo=%s, cd_status=%s, time_updated=%s , utr=%s
                                where code=%s limit 1"""
        else:
            # 更新订单的 SQL 语句，增加对 `cd_memo` 和 `cd_status` 的更新
            sql_update_order = """update orders_cd set is_cd=1, cd_memo=%s, cd_status=%s, time_updated=%s , utr=%s
                                where code=%s limit 1"""
        
        # 插入操作日志的 SQL 语句
        sql_insert_log = """INSERT INTO operation_logs (operator, ip_address, operation_time, operation_button, menu, operation_content) 
                            VALUES (%s, %s, NOW(), %s, %s, %s)"""

        async with self.application.db.acquire() as conn:
            async with conn.cursor(DictCursor) as cur:
                try:
                    # 查询订单
                    if not await cur.execute(sql_select_order, code):
                        return await self.json_response(msg[10007])
                    order = (await cur.fetchall())[0]
                    
                    # 获取当前时间
                    time_now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    
                    # 更新订单状态，增加 cd_memo 和 cd_status
                    if status == 1:
                        if not await cur.execute(sql_update_order, (operator, cd_memo, status, time_now, utr, code)):
                            await conn.rollback()
                            return await self.json_response(msg[10007])
                    else:
                        if not await cur.execute(sql_update_order, (cd_memo, status, time_now, utr, code)):
                            await conn.rollback()
                            return await self.json_response(msg[10007])
                    
                    # 记录操作日志
                    # 根据 cd_status 设置操作按钮名称
                    if status == 1:
                        operation_button = '审核'
                    elif status == 2:
                        operation_button = '确认审核'
                    elif status == 3:
                        operation_button = '反审核'
                    else:
                        operation_button = '其他操作'  # 可根据需要修改
                    menu = '代收查单'  # 根据需要修改菜单名称
                    operation_content = f'更新代收查单: {code}，备注: {cd_memo}，请求数据: {data}'

                    await cur.execute(sql_insert_log, (operator, ip_address, operation_button, menu, operation_content))

                    self.logger.info('更新代收查单状态%s' % cur._last_executed)
                    await conn.commit()

                except Exception as e:
                    self.logger.warning('代收查单处理失败,code={code},异常={e}'.format(code=code, e=e))
                    await conn.rollback()
                    return await self.json_response(msg[10007])

        result = dict(code=20000, msg='处理成功')
        return await self.json_response(result)

# 代收订单 to 查单
class addDsToCd(BaseHandler):
    @tornado.web.authenticated
    async def post(self):
        data = json.loads(self.request.body)
        code = data.get('code')  # 获取传递的参数 code
        utr = data.get('utr')
        operator = self.current_user['id']

        # 查找订单是否已存在于 orders_cd
        sql_check_cd = """SELECT 1 FROM orders_cd WHERE code=%s LIMIT 1"""
        
        # 查找订单
        sql_select_order = """SELECT * FROM orders_ds WHERE code=%s ORDER BY id DESC LIMIT 1"""
        
        # 插入数据到 orders_cd 的 SQL 语句
        sql_insert_order_cd = """
            INSERT INTO orders_cd (code, amount, realpay, poundage, channel_code, status, callback, 
                                   notice_api, notify, player_ip, remark, pay_url, time_create, time_accept, 
                                   time_payed, time_success, merchant_id, merchant_code, merchant_rate, 
                                   earn_merchant, partner_id, earn_partner_self, earn_partner, payment_id, 
                                   upi, utr, auth_code, realname, player_provence, otherpay, earn_system, 
                                   time_updated, cd_memo, cd_status, is_cd, cd_admin_id)
            SELECT code, amount, realpay, poundage, channel_code, status, callback, 
                   notice_api, notify, player_ip, remark, pay_url, time_create, time_accept, 
                   time_payed, time_success, merchant_id, merchant_code, merchant_rate, 
                   earn_merchant, partner_id, earn_partner_self, earn_partner, payment_id, 
                   upi, %s, auth_code, realname, player_provence, otherpay, earn_system, 
                   time_updated, '', 0, 0, %s
            FROM orders_ds WHERE code=%s LIMIT 1
        """

        # 锁定非内部码商(码商类型[type] 0内部 1外部，状态 [status] 1正常 0封禁)
        sql_lock_partner_with_type1 = """
            update partner 
            inner join orders_ds on orders_ds.partner_id = partner.id 
            set partner.status = 0 
            where partner.type = 1 and orders_ds.code = %s
        """

        async with self.application.db.acquire() as conn:
            async with conn.cursor(DictCursor) as cur:
                try:
                    # 检查 orders_cd 表中是否已存在记录
                    if await cur.execute(sql_check_cd, code):
                        return await self.json_response(msg[10008])
                    # 查询 orders_ds 中的订单
                    if not await cur.execute(sql_select_order, code):
                        return await self.json_response(msg[10007])
                    # 插入数据到 orders_cd
                    if not await cur.execute(sql_insert_order_cd, (utr, operator, code)):
                        await conn.rollback()
                        return await self.json_response(msg[10007])

                    # 锁定外部码商
                    await cur.execute(sql_lock_partner_with_type1, (code))

                    self.logger.info('数据插入 orders_cd 成功: %s' % cur._last_executed)
                    await conn.commit()

                except Exception as e:
                    self.logger.warning('插入订单到 orders_cd 失败,code={code},异常={e}'.format(code=code, e=e))
                    await conn.rollback()
                    return await self.json_response(msg[10007])

        result = dict(code=20000, msg='处理成功')
        return await self.json_response(result)


class getCdType(BaseHandler):
    @tornado.web.authenticated
    async def post(self):
        # SQL 查询语句，根据需要调整查询的表和字段
        sql_select_cd_types = """SELECT * FROM cd_types ORDER BY id"""

        async with self.application.db.acquire() as conn:
            async with conn.cursor(DictCursor) as cur:
                try:
                    # 执行查询
                    await cur.execute(sql_select_cd_types)
                    cd_types = await cur.fetchall()  # 获取所有结果

                    result = dict(code=20000, data=cd_types, msg='获取成功')
                    return await self.json_response(result)
    
                except Exception as e:
                    self.logger.warning('获取 cd_type 失败, 异常={e}'.format(e=e))
                    return await self.json_response(msg[10007])

        return await self.json_response(result)


# 代收手动回调
class handleNotifyds(BaseHandler):
    @tornado.web.authenticated
    async def post(self):
        data = json.loads(self.request.body)
        order = await self.get_result_by_condition('orders_ds', ['status'], {'code': data['code']})
        if order['status'] not in [-1, 3]:
            return await self.json_response(msg[10007])
        await self.redis.publish('order_notify', data['code'])
        result = dict(code=20000, msg='操作成功')
        return await self.json_response(result)


# 拆分代付订单
# ==========================================================
# confirmSplitOrder 类文件内容
# ==========================================================
class confirmSplitOrder(BaseHandler):
    @tornado.web.authenticated
    async def post(self):
        data = json.loads(self.request.body)
        
		# 查询角色的所有权限name
        # role_permissions_sql = f"select name from permissions where status = 1 and id in ({self.current_user['permissions']})"
        # role_permission_names_json = await self.query(role_permissions_sql)
        # # 字典对象 转 集合
        # role_permission_names = set(value for dictionary in role_permission_names_json for value in dictionary.values())
        # # print(role_permission_names)
        # if ("代付拆单" not in role_permission_names):
        #         return await self.json_response(msg[10257])
				
        self.logger.info(f"[confirmSplitOrder] 收到拆分请求，数据: {data}")
        required_params = ['code', 'amount']
        for param in required_params:
            if param not in data:
                self.logger.warning(f"[confirmSplitOrder] 缺少必要参数: {param}。请求数据: {data}")
                return await self.json_response(msg[10247])
        parent_order_code = data['code']
        split_amount = Decimal(str(data['amount'])) # 确保金额是Decimal类型
        if split_amount <= 0:
            self.logger.warning(f"[confirmSplitOrder] 拆分金额 {split_amount} 必须大于0。母订单号: {parent_order_code}")
            return await self.json_response(msg[10248])
        # 使用 Redis 锁防止并发拆分
        busy_key = f'split_order_{parent_order_code}'
        self.logger.info(f"[confirmSplitOrder] 尝试获取Redis锁: {busy_key}")
        count_circle = 0
        while True:
            if await self.redis.setnx(busy_key, 1):
                await self.redis.expire(busy_key, 10)
                self.logger.info(f"[confirmSplitOrder] 成功获取Redis锁: {busy_key}")
                break
            if count_circle >= 25: # 25 * 0.2s = 5秒
                self.logger.warning(f"[confirmSplitOrder] 获取Redis锁超时，订单 {parent_order_code} 正在被其他进程操作或操作频繁。")
                return await self.json_response(msg[10249])
            time.sleep(0.2)
            count_circle += 1
        conn = None
        try:
            async with self.application.db.acquire() as conn:
                async with conn.cursor(DictCursor) as cur:
                    # 1. 查询母订单详情
                    self.logger.info(f"[confirmSplitOrder] 事务内查询母订单 {parent_order_code} 详情。")
                    sql_select_parent = "SELECT * FROM orders_df WHERE code = %s LIMIT 1"
                    await cur.execute(sql_select_parent, (parent_order_code,))
                    parent_order_data = await cur.fetchone()
                    if not parent_order_data:
                        await conn.rollback()
                        self.logger.warning(f"[confirmSplitOrder] 母订单 {parent_order_code} 未找到。")
                        return await self.json_response(msg[10250])
                    # 2. 验证母订单状态
                    if  parent_order_data['status'] not in (0, 1) or parent_order_data['parent_id'] != '':
                        await conn.rollback()
                        self.logger.warning(f"[confirmSplitOrder] 母订单 {parent_order_code} 状态为 {parent_order_data['status']}，不允许拆分。只允许派单中(0)状态的订单进行拆分。")
                        return await self.json_response(msg[10251])
                    # 3. 计算已拆分金额和验证拆分总金额
                    sql_sum_split_amount = """
                        SELECT SUM(amount) as total_split_amount
                        FROM orders_df
                        WHERE status!=-1 AND parent_id = %s AND code != %s AND is_del=0
                    """
                    await cur.execute(sql_sum_split_amount, (parent_order_code, parent_order_code))
                    sum_res = await cur.fetchone()
                    existing_split_amount = Decimal(str(sum_res['total_split_amount'])) if sum_res and sum_res['total_split_amount'] else Decimal('0')
                    total_after_this_split = existing_split_amount + split_amount
                    self.logger.info(f"[confirmSplitOrder] 母订单 {parent_order_code} 原始金额: {parent_order_data['amount']}，已拆分金额: {existing_split_amount}，本次拆分金额: {split_amount}，拆分后总金额: {total_after_this_split}。")
                    if total_after_this_split > parent_order_data['amount']:
                        await conn.rollback() # 金额超限，回滚
                        self.logger.warning(f"[confirmSplitOrder] 拆分金额 {split_amount} 导致子订单总金额 ({total_after_this_split}) 超过母订单原始金额 ({parent_order_data['amount']})。")
                        return await self.json_response(msg[10252]) # 直接传递字典
                    # 4. 更新母订单为已拆分状态 (is_split = 1)
                    if parent_order_data.get('is_split') != 1:
                        sql_update_parent = "UPDATE orders_df SET is_split = 1, status=1 WHERE code = %s"
                        await cur.execute(sql_update_parent, (parent_order_code,))
                        if cur.rowcount > 0:
                            self.logger.info(f"[confirmSplitOrder] 成功更新母订单 {parent_order_code} 的 is_split 为 1。")
                        else:
                            await conn.rollback() # 更新失败，回滚
                            self.logger.error(f"[confirmSplitOrder] 更新母订单 {parent_order_code} 的 is_split 状态失败，rowcount为0。")
                            return await self.json_response(msg[10253]) # 直接传递字典
                    # 5. 生成新的子订单 code
                    sql_count_children = "SELECT COUNT(*) as count_children FROM orders_df WHERE parent_id = %s AND code != %s"
                    await cur.execute(sql_count_children, (parent_order_code, parent_order_code))
                    children_count = (await cur.fetchone())['count_children']
                    new_child_code = f"{parent_order_code}{"z"}{str(children_count + 1).zfill(2)}" # 例如 z01, z02
                    self.logger.info(f"[confirmSplitOrder] 为母订单 {parent_order_code} 生成新的子订单号: {new_child_code}。")
                    # 6. 插入新的子订单记录
                    child_order_data = parent_order_data.copy()
                    child_order_data['code'] = new_child_code
                    child_order_data['parent_id'] = parent_order_code
                    child_order_data['amount'] = split_amount
                    child_order_data['realpay'] = split_amount
                    child_order_data['status'] = 0 
                    child_order_data['is_split'] = 0
                    child_order_data['id'] = None
                    child_order_data['time_create'] = datetime.now() # 创建时间
                    child_order_data['time_accept'] = datetime.now() # 接受时间
                    child_order_data['time_success'] = None # 成功时间
                    child_order_data['poundage'] = Decimal('0')  # 手续费重置
                    child_order_data['earn_merchant'] = Decimal('0') # 佣金重置
                    child_order_data['earn_partner_self'] = Decimal('0')
                    child_order_data['earn_system'] = Decimal('0')
                    child_order_data['utr'] = None
                    child_order_data['is_del'] = 0 #追加0701
                    child_order_data['debit_account'] = None
                    columns = []
                    placeholders = []
                    insert_values = []
                    for k, v in child_order_data.items():
                        if k != 'id':
                            columns.append(k)
                            placeholders.append('%s')
                            insert_values.append(v)
                    sql_insert_child = f"INSERT INTO orders_df ({', '.join(columns)}) VALUES ({', '.join(placeholders)})"
                    await cur.execute(sql_insert_child, tuple(insert_values))
                    if cur.rowcount > 0:
                        self.logger.info(f"[confirmSplitOrder] 成功插入子订单 {new_child_code}，金额 {split_amount}。")
                    else:
                        await conn.rollback() # 插入失败，回滚
                        self.logger.error(f"[confirmSplitOrder] 插入子订单 {new_child_code} 失败，rowcount为0。")
                        return await self.json_response(msg[10254]) # 直接传递字典
                # 所有数据库操作成功后，在这里提交事务
                await conn.commit()
                self.logger.info(f"[confirmSplitOrder] 母订单 {parent_order_code} 拆分成功，事务已提交。")
            result = dict(code=20000, msg='操作成功')
            return await self.json_response(result)
        except Exception as e:
            self.logger.error(f"[confirmSplitOrder] 订单 {parent_order_code} 拆分过程中发生未预期异常: {e}", exc_info=True) # 打印详细 traceback
            try:
                if conn:
                    await conn.rollback()
                    self.logger.info(f"[confirmSplitOrder] 事务因异常而回滚。")
            except Exception as rollback_e:
                self.logger.error(f"[confirmSplitOrder] 异常回滚时再次发生错误: {rollback_e}", exc_info=True)
            # 返回通用错误消息，包含异常信息
            return await self.json_response(msg[10256])
        finally:
            self.logger.info(f"[confirmSplitOrder] 释放Redis锁: {busy_key}")
            await self.redis.delete(busy_key)
# 获取拆单明细
class getOrderDfSplit(BaseHandler):
    @tornado.web.authenticated
    async def post(self):
        data = json.loads(self.request.body)
		# 查询角色的所有权限name
        # role_permissions_sql = f"select name from permissions where status = 1 and id in ({self.current_user['permissions']})"
        # role_permission_names_json = await self.query(role_permissions_sql)
        # # 字典对象 转 集合
        # role_permission_names = set(value for dictionary in role_permission_names_json for value in dictionary.values())
        # # print(role_permission_names)
        # if ("代付子单查看" not in role_permission_names):
        #         return await self.json_response(msg[10257])
        # 变更后: 只需要传入一个 'code' 参数，作为父订单号
        if 'code' not in data:
            self.logger.warning(f"请求参数缺少 'code'，无法查询拆单明细。请求数据: {data}")
            return await self.json_response(dict(code=99, msg='缺少订单号参数'))
        parent_order_code = data['code']
        
        # 考虑到 parent_id 字段可能是 varchar，且存储的是父订单的 code
        sql = """
            SELECT
                id,
                code,
                merchant_code,
                amount,
                status,
                merchant_id,
                partner_id,
                payment_id,
                ifsc,
                time_create,
                time_accept,
                time_payed,
                time_success,
                time_updated,
                otherpay_id,
                otherpay,
                payment_bank,
                payment_account,
                payment_name,
                realpay,
                merchant_rate,
                poundage,
                sys_remark,
                utr,
                debit_account
            FROM
                orders_df
            WHERE
                parent_id = %s
                AND code != %s -- 确保不包含父订单自身
            ORDER BY time_create ASC -- 按照创建时间排序，便于查看
        """
        values = [parent_order_code, parent_order_code]

        # 变更后: 直接执行查询获取子订单数据
        self.logger.info(f"查询拆单明细：根据父订单号 '{parent_order_code}' 获取子订单。SQL: {sql}，参数: {values}")
        data_r = await self.query(sql, *values)

        # 变更后: 返回查询到的子订单数据
        result = dict(code=20000, data=data_r, msg='获取拆单明细成功')
        self.logger.info(f"成功获取父订单 '{parent_order_code}' 的拆单明细，共 {len(data_r)} 条子订单。")
        return await self.json_response(result)
# 获取代付订单
class getOrderDf(BaseHandler):
    @tornado.web.authenticated
    async def post(self):

        # 查询角色的所有权限name
        role_permissions_sql = f"select name from permissions where status = 1 and id in ({self.current_user['permissions']})"
        role_permission_names_json = await self.query(role_permissions_sql)
        # 字典对象 转 集合
        role_permission_names = set(value for dictionary in role_permission_names_json for value in dictionary.values())

        data = json.loads(self.request.body)
        
        isExport = data.get('size') == 0 and data.get('page') == 0

        condition, between = await self.split_between_condition(data['serchData'], 'time_create')
        if not between:
            condition, between = await self.split_between_condition(condition, 'time_success')
        if not between:
            condition, between = await self.split_between_condition(condition, 'time_updated')

        if not condition or not condition.get('code') and not between:
            between = {'key': 'time_create', 'start': datetime.today().date(), 'end': datetime.now()}
        keys_count = ['amount', 'status','realpay','earn_merchant','earn_partner_self','earn_system','otherpay_id','otherpay','debit_account', 'utr', 'is_split', 'parent_id']
        sql_part = ' where 1 = 1 '
        sql_part_ot = ''
        values = []

        # 获取顶级id下的所有订单
        top_partner_sql_part = ''
        partner_ids = []
        if 'top_partner_id' in condition.keys():
            partner_ids = await self.get_partners(condition['top_partner_id'])
            del condition['top_partner_id']
            if partner_ids:
                _partner_ids = ','.join(['%s'] * len(partner_ids))
                top_partner_sql_part = ' partner_id in ({partner_ids})'.format(partner_ids=_partner_ids)

        # 如果有金额范围查找
        amount_range = ''
        if 'amount_range' in condition.keys():
            if condition['amount_range'] == 1: # <=500
                amount_range = ' amount <= 500'
            elif condition['amount_range'] == 2: # 500-1000
                amount_range = ' amount <= 1000 and amount > 500'
            elif condition['amount_range'] == 3: # 1000-2000
                amount_range = ' amount <= 2000 and amount > 1000'
            elif condition['amount_range'] == 4: # 2000-5000
                amount_range = ' amount <= 5000 and amount > 2000'
            # elif condition['amount_range'] == 5: # >5000
            #     amount_range = ' amount > 5000'
            elif condition['amount_range'] == 5: # 5000-20000
                amount_range = ' amount <= 20000 and amount > 5000'
            elif condition['amount_range'] == 6: # 20000-50000
                amount_range = ' amount <= 50000 and amount > 20000'
            elif condition['amount_range'] == 7: # >50000
                amount_range = ' amount > 50000'
            del condition['amount_range']

        # 用户输入的金额范围
        # amount_range_input = condition['amount_range_new']  # 用户输入的金额范围

        # 用户输入的值可能是类似 "500-1000", "1000-2000" 这样的格式
        if 'amount_range_new' in condition.keys():
            # 去掉输入的首尾空格
            amount_range_input = condition['amount_range_new'].strip()

            # 解析用户输入
            if '-' in amount_range_input:  # 如果用户输入了一个范围（如 500-1000）
                amount_parts = amount_range_input.split('-')
                if len(amount_parts) == 2:
                    lower_limit = amount_parts[0].strip()  # 下限
                    upper_limit = amount_parts[1].strip()  # 上限
                    amount_range = f"amount <= {upper_limit} and amount > {lower_limit}"
                    del condition['amount_range_new']

        # 删除原始条件中的 'amount_range' 键
        # print(amount_range)

        bank_type_id = condition.pop('bank_type_id', None)
        if bank_type_id:
            sql_part += ' and payment_id in (select id from payment where bank_type_id = %s)'
            values.append(bank_type_id)

        if condition:
            for k in list(condition.keys()):
                if not condition[k] and condition[k] != 0:
                    condition.pop(k)
        payment_id = condition.get('payment_id')
        # 根据传参读取订单类型
        order_type = condition.get('order_type')
        if order_type:
            cond_arr = []
            if 'no' in order_type:
                cond_arr.append('(is_split = 0 and parent_id = \'\')')
            if 'mo' in order_type:
                cond_arr.append('(is_split = 1 and parent_id = \'\')')
            if 'so' in order_type:
                cond_arr.append('(parent_id <> \'\')')
            sql_part_ot = f' and ({" or ".join(cond_arr)})'
            sql_part += sql_part_ot
            condition.pop('order_type')
        
        if condition:
            if 'status' in condition and condition['status'] == 5:
                sql_part += ' status in (1,2) and'
                del condition['status']
            where_key, where_val = await self.dict_to_and(condition)
            sql_part += ' and {keys} '.format(keys=where_key)
            values += where_val
        if between:
            bt_key, bt_start, bt_end = await self.dict_to_between(between)
            if await self.calculate_date_diff(bt_start, bt_end):
                return await self.json_response(msg[10301])
            sql_part += " and " + bt_key
            values += [bt_start, bt_end]
        
        if top_partner_sql_part:
            sql_part += ' and ' + top_partner_sql_part
            values.extend(partner_ids)
                
        # --- 新增的母/子订单筛选逻辑 ---
        new_condition = None
        # order_type_filter = data['serchData']['is_split']
        # if order_type_filter == 1:
        #     new_condition = " (parent_id IS NULL OR parent_id = '') "
        # elif order_type_filter == 0:
        #     new_condition = " (is_split = 0 AND parent_id IS NOT NULL AND parent_id != '') "

        if new_condition:
            sql_part += ' and ' + new_condition
                
        # 如果有金额范围查找        
        if amount_range:
            sql_part += ' and ' + amount_range

        # 如果是推广账号，则过滤出推广账号下的码商订单
        if str(self.current_user['role_id']) == '19':
            tg_partners_ids = await self.get_partners_by_parent_id(self.current_user['parent_id'])
            sql_part += ' and partner_id in ({})'.format(tg_partners_ids)

        # 获取所有数据总数
        sql = "select count(id) from orders_df"
        sql += sql_part
        total = await self.query(sql, *values)
        if total:
            total = total[0]['count(id)']
        else:
            total = 0
            
        # 获取所有数据的指定key数据
        count = []
        if keys_count:
            sql = "select {keys} from orders_df ".format(keys=await self.list_keys(keys_count))
            # 统计数据，强制过滤子单的
            if not payment_id:
                if sql_part_ot:
                    sql += sql_part.replace(sql_part_ot, " and parent_id = \'\'")
                else:
                    sql += sql_part + " and parent_id = \'\'"
            else:
                sql += sql_part
            count = await self.query(sql, *values)
            
        # region 获取权限值 取缔allowAdminProfit识别权限，改为代收代付权限
        # allowAdminProfit = await self.redis.get('allowAdminProfit')
        # if allowAdminProfit:
        #     allowAdminProfit = allowAdminProfit.split(',')
        # # 权限控制
        # userPermissions = False
        # # 获取登录管理员id
        # user_id = str(self.get_secure_cookie("user"), 'utf-8')
        # if allowAdminProfit:
        #     # 查看权限是否存在
        #     if user_id in allowAdminProfit:
        #         userPermissions = True
        # 获取分页数据
        # sql = "select * from {table}  where id in (select id from {table} ".format(table='orders_df')
        # order_by = ') order by {order_field} {sort} '.format(order_field=data['order_field'], sort=data['sort'])
        # endregion
        
        sql = "select * from orders_df "
        order_by = ' order by {order_field} {sort} '.format(order_field=data['order_field'], sort=data['sort'])
        sql += sql_part + order_by
        if data['size'] and data['page'] > -1:
            sql += 'limit %s offset %s'
            values += [data['size'], (data['page'] - 1) * data['size']]
        data_r = await self.query(sql, *values)
        # print('sql==', sql)
        
        if isExport and (not order_type or set(['mo', 'so']).issubset(order_type)):
            data_r = self.orderExportSort(data_r)
        
        parent_codes = []
        for item in data_r:
            if item.get('is_split') == 1:
                parent_codes.append(item['code'])

        child_amounts_sum = {}
        if parent_codes:
            # sql_child_sum = """
            #     SELECT parent_id, SUM(amount) as total_child_amount
            #     FROM orders_df
            #     WHERE parent_id IN ({}) AND status != -1 AND is_del=0
            #     GROUP BY parent_id
            # """.format(','.join(['%s'] * len(parent_codes)))
            sql_child_sum = """
                SELECT parent_id, SUM(amount) as total_child_amount
                FROM orders_df
                WHERE parent_id IN ({}) AND is_del=0
                GROUP BY parent_id
            """.format(','.join(['%s'] * len(parent_codes)))
            child_sums_raw = await self.query(sql_child_sum, *parent_codes)
            for row in child_sums_raw:
                child_amounts_sum[row['parent_id']] = Decimal(str(row['total_child_amount']))
                
        child_status = {}
        if parent_codes:
            sql_child_status = """SELECT parent_id, code, status FROM orders_df WHERE parent_id IN ({}) AND is_del=0""".format(','.join(['%s'] * len(parent_codes)))
            child_status_rows = await self.query(sql_child_status, *parent_codes)
            for child_status_row in child_status_rows:
                key = child_status_row["parent_id"]
                if key not in child_status:
                    child_status[key] = []
                child_status[key].append(child_status_row["status"])
     
        split_amount = Decimal('0.00')
        
        for i in range(len(data_r)):
            item_code = data_r[i]['code']
            # region 计算 amount_remains
            parent_total_amount = Decimal(str(data_r[i]['amount'])) # 母订单的原始金额
            
            current_child_sum = child_amounts_sum.get(item_code, Decimal('0'))
            
            data_r[i]['amount_remains'] = parent_total_amount - current_child_sum
            
            # 确保 amount_remains 不会是负数（如果数据异常）
            if data_r[i]['amount_remains'] < 0:
                data_r[i]['amount_remains'] = Decimal('0.00')

            if ("禁止查看代付订单结算金额" in role_permission_names):
                data_r[i]['realpay'] = '******'
            if ("禁止查看代付订单商户费率" in role_permission_names):
                data_r[i]['merchant_rate'] = '******'
            if ("禁止查看代付订单手续费" in role_permission_names):
                data_r[i]['poundage'] = '******'
            if ("禁止查看代付订单商代盈利" in role_permission_names):
                data_r[i]['earn_merchant'] = '******'
            if ("禁止查看代付订单码商盈利" in role_permission_names):
                data_r[i]['earn_partner_self'] = '******'
            if ("禁止查看代付订单平台利润" in role_permission_names):
                data_r[i]['earn_system'] = '******'
            # if data_r[i]['is_split'] == 0 and data_r[i]['parent_id'] != '':
            if payment_id:
                split_amount += data_r[i]['amount']
            else:
                split_amount += data_r[i]['amount']-data_r[i]['amount_remains']
                 # print(f'is_split: {data_r[i]['is_split']}, parent_id: {data_r[i]['parent_id']}, amount: {data_r[i]['amount']}, split_amount: {data_r[i]['split_amount']}')
            #endregion
            #region 获取子订单状态集
            data_r[i]['childset'] = child_status.get(item_code, [])
            #endregion
        
        count_r = {'failOrder': 0, 'split_amount': split_amount, 'successOrder': 0, 'processing': 0, 'amount': Decimal(0),
                   'processing_amount': Decimal(0),'realpay': Decimal(0),'earn_merchant': Decimal(0),'earn_partner': Decimal(0),'earn_system': Decimal(0)}
        for i in count:
            if i['status'] == 4:
                count_r['successOrder'] += 1
                count_r['amount'] += i['amount']
                # if userPermissions is True:
                count_r['realpay'] += i['realpay']
                count_r['earn_merchant'] += i['earn_merchant']
                if i['earn_partner_self']:
                    count_r['earn_partner'] += i['earn_partner_self']
                if i['earn_system']:
                    count_r['earn_system'] += i['earn_system']
            elif i['status'] == -1:
                count_r['failOrder'] += 1
            else:
                count_r['processing'] += 1
                count_r['processing_amount'] += i['amount']
            
        # 没有权限赋值为******
        # if userPermissions is False:
        #     count_r['realpay'] = "******"
        #     count_r['earn_merchant'] = "******"
        #     count_r['earn_partner'] = "******"
        #     count_r['earn_system'] = "******"
        if ("禁止查看代收订单结算金额" in role_permission_names):
            count_r['realpay'] = "******"
        if ("禁止查看代收订单商代盈利" in role_permission_names):
            count_r['earn_merchant'] = "******"
        if ("禁止查看代收订单码商盈利" in role_permission_names):
            count_r['earn_partner'] = "******"
        if ("禁止查看代收订单平台利润" in role_permission_names):
            count_r['earn_system'] = "******"
        # 商户处理中的订单数
        # cdt = ""
        # value = []
        # if condition and "channel_code" in condition.keys():
        #     cdt = ' and channel_code = ' + str(condition['channel_code'])
        # if between_create:
        #     cdt += ' and {key} between %s and %s'.format(key=between_create['key'])
        #     value.extend([between_create['start'], between_create['end']])
        #
        # # 对于 time_success
        # if between_success:
        #     cdt += ' and {key} between %s and %s'.format(key=between_success['key'])
        #     value.extend([between_success['start'], between_success['end']])
        #
        # # 对于 time_updated
        # if between_updated:
        #     cdt += ' and {key} between %s and %s'.format(key=between_updated['key'])
        #     value.extend([between_updated['start'], between_updated['end']])
        #
        # if not between:
        #     between = dict()
        #     between['key'] = 'time_create'
        #     between['start'] = datetime.today().date()
        #     between['end'] = datetime.now()
        #     cdt += 'and {key} between %s and %s'.format(key=between['key'])
        #     value = [between['start'], between['end']]

        merchant_processing = []

        # sql_p = """
        #     select merchant_id,count(1) as cnt,sum(amount) as total ,
        #      SUM(CASE WHEN status = 0 THEN 1 ELSE 0 END) AS status_0_count,
        #      SUM(CASE WHEN status = 0 THEN amount ELSE 0 END) AS status_0_total,
        #      SUM(CASE WHEN status = 1 THEN 1 ELSE 0 END) AS status_1_count,
        #      SUM(CASE WHEN status = 1 THEN amount ELSE 0 END) AS status_1_total
        #     from orders_df where status in (0,1)  {cdt}
        #     group by merchant_id order by cnt desc;
        # """.format(cdt=cdt)
        # merchant_processing = await self.query(sql_p, *value)

        merchant_finish = []

        # sql_p = """
        #     select merchant_id,count(1) as cnt,sum(amount) as total
        #     from orders_df where status in (3,4)  {cdt}
        #     group by merchant_id order by cnt desc;
        # """.format(cdt=cdt)
        # merchant_finish = await self.query(sql_p, *value)

        result = dict(code=20000, data=data_r, total=total, count=count_r, merchant_processing=merchant_processing, merchant_finish=merchant_finish, msg='获取成功')
        return await self.json_response(result)
    
    def orderExportSort(self, orders):
        child_orders_map = {}
        child_order_codes = set()

        def get_parent_code(code: str) -> str:
            match = re.match(r"(.+?)(?:_|z)\d+$", code)
            return match.group(1) if match else code

        # 构建子订单分组
        child_orders_map = {}
        child_order_codes = set()

        for order in orders:
            code = order["code"]
            parent_code = get_parent_code(code)
            if parent_code != code:
                child_orders_map.setdefault(parent_code, []).append(order)
                child_order_codes.add(code)

        # 子订单统一按时间升序排序
        for v in child_orders_map.values():
            v.sort(key=lambda x: x["time_create"])

        # 构建最终顺序（主单 → 所有子单（升序））
        result = []
        seen_codes = set()

        for order in orders:
            code = order["code"]
            if code in seen_codes:
                continue
            parent_code = get_parent_code(code)
            if code == parent_code:  # 主订单
                result.append(order)
                seen_codes.add(code)
                # 插入所有子订单（无论是 _ 还是 z 的）
                for child in child_orders_map.pop(code, []):
                    if child["code"] not in seen_codes:
                        result.append(child)
                        seen_codes.add(child["code"])
        for v in child_orders_map.values():
            for child_ in v:
                if child_["code"] not in seen_codes:
                    result.append(child_)
        return result

# 获取代付订单处理中或成交额数据
class getDFMerchantFinishOrProcessing(BaseHandler):
    @tornado.web.authenticated
    async def post(self):
        data = json.loads(self.request.body)

        merchant_finish_ = False
        if 'merchant_finish' in data and data['merchant_finish']:
            merchant_finish_ = True

        processing_ = False
        if 'processing' in data and data['processing']:
            processing_ = True

        condition, between = await self.split_between_condition(data['serchData'], 'time_create')
        if not between:
            condition, between = await self.split_between_condition(condition, 'time_success')
        if not between:
            condition, between = await self.split_between_condition(condition, 'time_updated')

        # 商户处理中的订单数
        cdt = ""
        value = []
        if condition and "channel_code" in condition.keys() and condition['channel_code']:
            cdt = ' and channel_code = ' + str(condition['channel_code'])
        if between:
            cdt += 'and {key} between %s and %s'.format(key=between['key'])
            value = [between['start'], between['end']]
        if not between:
            between = dict()
            between['key'] = 'time_create'
            between['start'] = datetime.today().date()
            between['end'] = datetime.now()
            cdt += 'and {key} between %s and %s'.format(key=between['key'])
            value = [between['start'], between['end']]
        merchant_processing = []
        if processing_:
            sql_p = """
                select merchant_id,count(1) as cnt,sum(amount) as total ,   
                 SUM(CASE WHEN status = 0 THEN 1 ELSE 0 END) AS status_0_count,
                 SUM(CASE WHEN status = 0 THEN amount ELSE 0 END) AS status_0_total,
                 SUM(CASE WHEN status = 1 THEN 1 ELSE 0 END) AS status_1_count,
                 SUM(CASE WHEN status = 1 THEN amount ELSE 0 END) AS status_1_total
                from orders_df where status in (0,1)  {cdt}
                group by merchant_id order by cnt desc; 
            """.format(cdt=cdt)
            merchant_processing = await self.query(sql_p, *value)

        merchant_finish = []
        if merchant_finish_:
            sql_p = """
                select merchant_id,count(1) as cnt,sum(amount) as total
                from orders_df where status in (3,4)  {cdt}
                group by merchant_id order by cnt desc;
            """.format(cdt=cdt)
            merchant_finish = await self.query(sql_p, *value)
        result = dict(code=20000,merchant_processing=merchant_processing,merchant_finish=merchant_finish, msg='获取成功')
        return await self.json_response(result)

# region
# 处理代付 代付改派权限分离
# 确认，上传凭证，改派，指派 全部分开   20240827
# class handleOrderdf(BaseHandler):
#     @tornado.web.authenticated
#     async def post(self):
#         data = json.loads(self.request.body)
#         # 使用锁，5s使用自旋锁, 防止取消的同时回调
#         count_circle = 0
#         while True:
#             busy_key = 'grab_df_{code}'.format(code=data['code'])
#             if await self.redis.setnx(busy_key, 1):
#                 await self.redis.expire(busy_key, 10)
#                 break
#             if count_circle >= 25:
#                 self.logger.warning('code:{}有其他进程正在处理中'.format(data['code']))
#                 return await self.json_response(msg[10032])
#             time.sleep(0.2)
#             count_circle = count_circle + 1

#         if data['type'] == 1:# 确认
#             return await BaseOrderHandler.order_success(self, data['code'], 0)
#         # 上传凭证
#         elif data['type'] == 2:
#             async with self.application.db.acquire() as conn:
#                 async with conn.cursor(DictCursor) as cur:
#                     try:
#                         sql_select = """select status from orders_df where code=%s limit 1"""
#                         if not await cur.execute(sql_select, data['code']):
#                             return await self.json_response(msg[10007])
#                         order = (await cur.fetchall())[0]
#                         # if order['status'] == 0:
#                         #     return await self.json_response(msg[10007])
#                         sql_update = """update orders_df set payment_img=1{status} where code=%s and status=%s limit 1""".format(
#                             status=',status=2' if order['status'] == 1 else '')
#                         if not await cur.execute(sql_update, (data['code'], order['status'])):
#                             return await self.json_response(msg[10007])
#                     except Exception:
#                         return await self.json_response(msg[10007])
#                     else:
#                         await conn.commit()
#         # 改派
#         elif data['type'] == 3:
#             self.logger.info('改派,code={code},管理员={e}'.format(code=data['code'], e=self.current_user['id']))
#             # 指派
#             if data['payment_id']:
#                 keys = ['certified', 'status', 'partner_id']
#                 payment = await self.get_result_by_condition('payment', keys, {'id': data['payment_id']})
#                 if not payment:
#                     return await self.json_response(msg[10033])
#                 # if not payment['certified']:
#                 #     return await self.json_response(msg[10019])
#                 if not payment['status']:
#                     return await self.json_response(msg[10020])
#                 keys = ['id', 'certified', 'status']
#                 partner = await self.get_result_by_condition('partner', keys, {'id': payment['partner_id']})
#                 if not partner['certified']:
#                     return await self.json_response(msg[10021])
#                 if not partner['status']:
#                     return await self.json_response(msg[10022])
#                 # 有三方代付的不能改派给其他码商
#                 sql_update = """update orders_df set partner_id=%s,payment_id=%s,time_accept=%s,status=1 where code=%s 
#                                     and status=%s and otherpay_id is null and otherpay is null limit 1"""
#                 if not await self.execute(sql_update,
#                                           *(partner['id'], data['payment_id'], datetime.now(), data['code'], data['status'])):
#                     return await self.json_response(msg[10007])
#             # 重派
#             else:
#                 return await self.json_response(msg[10007])
#                 # order = await self.get_result_by_condition('orders_df', ['amount', 'payment_id'],
#                 #                                            {'code': data['code']})
#                 # if not order:
#                 #     return await self.json_response(msg[10007])
#                 # sql_update = """update orders_df set partner_id=null,payment_id=null,status=0 where code=%s and status=%s limit 1"""
#                 # if not await self.execute(sql_update, *(data['code'], data['status'])):
#                 #     return await self.json_response(msg[10007])
#                 # # 重新派单
#                 # await self.redis.publish('order_df_push', '{code}_{amount}'.format(code=data['code'], amount=order['amount']))
#                 # # 码继续接单
#                 # if await self.redis.sismember('payment_online_df', order['payment_id']):
#                 #     await self.redis.lrem('payment_active_df', 0, order['payment_id'])
#                 #     await self.redis.rpush('payment_active_df', order['payment_id'])
#         elif data['type'] == 4:
#             self.logger.info('获取,code={code},管理员={e}'.format(code=data['code'], e=self.current_user['id']))
#             # 指派
#             if not data['payment_id']:
#                 return await self.json_response(msg[10007])
#             keys = ['certified', 'status', 'partner_id']
#             payment = await self.get_result_by_condition('payment', keys, {'id': data['payment_id']})
#             # if not payment['certified']:
#             #     return await self.json_response(msg[10019])
#             if not payment['status']:
#                 return await self.json_response(msg[10020])
#             keys = ['id', 'certified', 'status']
#             partner = await self.get_result_by_condition('partner', keys, {'id': payment['partner_id']})
#             if not partner['certified']:
#                 return await self.json_response(msg[10021])
#             if not partner['status']:
#                 return await self.json_response(msg[10022])
#             sql_update = """update orders_df set partner_id=%s,payment_id=%s,time_accept=%s,status=1 where code=%s 
#                                 and status=0 limit 1"""
#             if not await self.execute(sql_update,
#                                       *(partner['id'], data['payment_id'], datetime.now(), data['code'])):
#                 return await self.json_response(msg[10007])
#         return await self.json_response(msg[20000])
# endregion

# 代付确认
class HandleOrderdfType1(BaseOrderHandler):
    @tornado.web.authenticated
    async def post(self):
        data = json.loads(self.request.body)

        # #328-需求变化：只要有【确认】权限，同时也能进入【子订单查看】，也就可以操作子单
        # --- 子订单特殊处理逻辑 ---
        # if '_' in data['code']:
        #     # 查询角色的所有权限name
        #     role_permissions_sql = f"select name from permissions where status = 1 and id in ({self.current_user['permissions']})"
        #     role_permission_names_json = await self.query(role_permissions_sql)
        #     # 字典对象 转 集合
        #     role_permission_names = set(value for dictionary in role_permission_names_json for value in dictionary.values())
        #     if ("代付拆单" not in role_permission_names):
        #             return await self.json_response(msg[10257])
                
        if not await self.acquire_spin_lock(data['code']):
            return await self.json_response(msg[10032])  # 处理其他进程正在处理中

        return await self.order_success(data['code'], 0)  # 确认订单

class HandleOrderdfType2(BaseHandler):
    @tornado.web.authenticated
    async def post(self):
        data = json.loads(self.request.body)

        if not await BaseOrderHandler.acquire_spin_lock(self, data['code']):
            return await self.json_response(msg[10032])  # 处理其他进程正在处理中

        async with self.application.db.acquire() as conn:
            async with conn.cursor(DictCursor) as cur:
                try:
                    sql_select = """SELECT status FROM orders_df WHERE code=%s LIMIT 1"""
                    if not await cur.execute(sql_select, data['code']):
                        return await self.json_response(msg[10007])
                    order = (await cur.fetchall())[0]

                    sql_update = """UPDATE orders_df SET payment_img=1{status} WHERE code=%s AND status=%s LIMIT 1""".format(
                        status=',status=2' if order['status'] == 1 else '')
                    if not await cur.execute(sql_update, (data['code'], order['status'])):
                        return await self.json_response(msg[10007])
                except Exception:
                    return await self.json_response(msg[10007])
                else:
                    await conn.commit()
        return await self.json_response(msg[20000])  # 上传凭证成功

class HandleOrderdfType3(BaseHandler):
    @tornado.web.authenticated
    async def post(self):
        data = json.loads(self.request.body)
                
        if await self.is_null(data, ['status', 'code', 'payment_id']):
            return await self.json_response(data=msg[10007])
        
        status = data['status']
        code = data['code']
        payment_id = data['payment_id']
        # parent_id = data.get('parent_id')

        if not await BaseOrderHandler.acquire_spin_lock(self, code):
            return await self.json_response(msg[10032])  # 处理其他进程正在处理中

        self.logger.info('改派, code={code}, 管理员={e}'.format(code=code, e=self.current_user['id']))
        
        if payment_id:
            keys = ['certified', 'status', 'partner_id']
            payment = await self.get_result_by_condition('payment', keys, {'id': payment_id})
            print('HandleOrderdfType3--------------1-----------------------')
            if not payment:
                return await self.json_response(msg[10033])
            if not payment['status']:
                return await self.json_response(msg[10020])
            
            keys = ['id', 'certified', 'status', 'type']
            partner = await self.get_result_by_condition('partner', keys, {'id': payment['partner_id']})
            if not partner:
                return await self.json_response(msg[10034])
            if not partner['certified']:
                return await self.json_response(msg[10021])
            if not partner['status']:
                return await self.json_response(msg[10022])
            # if parent_id and partner['type'] == 1:
            #     return await self.json_response(msg[10023])
            
            # 有三方代付的不能改派给其他码商
            sql_update = """update orders_df set partner_id=%s,payment_id=%s,time_accept=%s,status=1 where code=%s 
								and status=%s and otherpay_id is null and otherpay is null limit 1"""
            if not await self.execute(sql_update, *(partner['id'], payment_id, datetime.now(), code, status)):
               return await self.json_response(msg[10007])
        else:
            return await self.json_response(msg[10007])  # 处理错误情况
                # order = await self.get_result_by_condition('orders_df', ['amount', 'payment_id'],
                #                                            {'code':code})
                # if not order:
                #     return await self.json_response(msg[10007])
                # sql_update = """update orders_df set partner_id=null,payment_id=null,status=0 where code=%s and status=%s limit 1"""
                # if not await self.execute(sql_update, *(code, status)):
                #     return await self.json_response(msg[10007])
                # # 重新派单
                # await self.redis.publish('order_df_push', '{code}_{amount}'.format(code=code, amount=order['amount']))
                # # 码继续接单
                # if await self.redis.sismember('payment_online_df', order['payment_id']):
                #     await self.redis.lrem('payment_active_df', 0, order['payment_id'])
                #     await self.redis.rpush('payment_active_df', order['payment_id'])
                
        return await self.json_response(msg[20000])  # 改派成功

class HandleOrderdfType4(BaseHandler):
    @tornado.web.authenticated
    async def post(self):
        data = json.loads(self.request.body)
                
        if await self.is_null(data, ['code', 'payment_id']):
            return await self.json_response(data=msg[10007])
        
        code = data['code']
        payment_id = data['payment_id']
        # parent_id = data.get('parent_id')

        if not await BaseOrderHandler.acquire_spin_lock(self, code):
            return await self.json_response(msg[10032])  # 处理其他进程正在处理中

        self.logger.info('获取, code={code}, 管理员={e}'.format(code=code, e=self.current_user['id']))
                
        keys = ['certified', 'status', 'partner_id']
        payment = await self.get_result_by_condition('payment', keys, {'id': payment_id})
        if not payment:
            return await self.json_response(msg[10033])
        if not payment['status']:
            return await self.json_response(msg[10020])
        
        keys = ['id', 'certified', 'status', 'type']
        partner = await self.get_result_by_condition('partner', keys, {'id': payment['partner_id']})
        if not partner:
            return await self.json_response(msg[10034])
        if not partner['certified']:
            return await self.json_response(msg[10021])
        if not partner['status']:
            return await self.json_response(msg[10022])
        # if parent_id and partner['type'] == 1:
        #     return await self.json_response(msg[10023])
        
        # 先查询订单，检查是否已分配payment_id
        order = await self.get_result_by_condition('orders_df', 
            ['status', 'payment_id'], 
            {'code': code})
        
        if not order:
            return await self.json_response(msg[10036])  # 订单不存在
        
        if order['status'] != 0:
            return await self.json_response(msg[10043])  # 订单状态不是派单中，无法分配
        
        if order['payment_id'] and order['payment_id'] != 0:
            self.logger.warning(f"订单{code}已分配payment_id={order['payment_id']}，拒绝重复分配")
            return await self.json_response(msg[10044])  # 订单已分配账号，不能重复分配
        
        sql_update = """update orders_df set partner_id=%s,payment_id=%s,time_accept=%s,status=1 where code=%s and status=0 limit 1"""
        if not await self.execute(sql_update, *(partner['id'], payment_id, datetime.now(), code)):
            return await self.json_response(msg[10007])
        
        return await self.json_response(msg[20000])

class HandleOrderdfRevert(BaseOrderHandler):
    @tornado.web.authenticated
    async def post(self):
        data = json.loads(self.request.body)
        
        if await self.is_null(data, ['code']):
            return await self.json_response(data=msg[10007])
        
        sql_child_sum = """SELECT SUM(amount) as total_child_amount FROM orders_df WHERE parent_id = %s AND is_del=0 GROUP BY parent_id """
        child_sums = await self.query(sql_child_sum, data['code'])
        usingAmount = 0
        if child_sums:
            usingAmount = child_sums[0]['total_child_amount']
        if usingAmount != 0:
            return await self.json_response(msg[10024])
        
        async with self.application.db.acquire() as conn:
            async with conn.cursor(DictCursor) as cur:
                try:
                    # 固定为【派单中】
                    sql_update = """UPDATE orders_df SET is_split=0, status=0 WHERE code=%s LIMIT 1"""
                    if not await cur.execute(sql_update, data['code']):
                        return await self.json_response(msg[10007])
                except Exception:
                    return await self.json_response(msg[10007])
                else:
                    await conn.commit()
                    
        return await self.json_response(msg[20000])


# 处理代付
class handleBatchOrderdf(BaseOrderHandler):
    @tornado.web.authenticated
    async def post(self):
        data = json.loads(self.request.body)
        for i in data['codes']:
            count_circle = 0
            while True:
                busy_key = 'grab_df_{code}'.format(code=i)
                if await self.redis.setnx(busy_key, 1):
                    await self.redis.expire(busy_key, 10)
                    break
                if count_circle >= 25:
                    self.logger.warning('handleBatchOrderdf-code:{}有其他进程正在处理中'.format(i))
                    return await self.json_response(msg[10032])
                time.sleep(0.2)
                count_circle = count_circle + 1
            if data['type'] == 1:# 确认
                await self.order_success(i, 1)
            elif data['type'] == 4:
                self.logger.info('handleBatchOrderdf-获取,code={code},管理员={e}'.format(code=i, e=self.current_user['id']))
                # 指派
                if not data['payment_id']:
                    continue
                keys = ['certified', 'status', 'partner_id']
                payment = await self.get_result_by_condition('payment', keys, {'id': data['payment_id']})
                if not payment['status']:
                    continue
                keys = ['id', 'certified', 'status']
                partner = await self.get_result_by_condition('partner', keys, {'id': payment['partner_id']})
                if not partner:
                    continue
                if not partner['certified']:
                    continue
                if not partner['status']:
                    continue

                sql_update = """update orders_df set partner_id=%s,payment_id=%s,time_accept=%s,status=1 where code=%s 
                                    and status=0 limit 1"""
                if not await self.execute(sql_update,
                                        *(partner['id'], data['payment_id'], datetime.now(), i)):
                    continue
        return await self.json_response(msg[20000])
    
    
# 处理代付-保存回执信息
class saveHuizhi(BaseHandler):
    @tornado.web.authenticated
    async def post(self):
        data = json.loads(self.request.body)
                
        if await self.is_null(data, ['id']):
            return await self.json_response(data=msg[10007])
        
        utr = data.get('huizhi_utr', '')
        da = data.get('huizhi_debitAccount', '')
        id = data.get('id', '')
        
        sql_update = """update orders_df set utr=%s,debit_account=%s where id=%s limit 1"""
        if not await self.execute(sql_update, *(utr, da, id)):
            return await self.json_response(msg[10007])
        return await self.json_response(msg[20000])
    
# 代付驳回
class cancelOrderdf(BaseHandler):
    @tornado.web.authenticated
    async def post(self):
        data = json.loads(self.request.body)

        if await self.is_null(data, ['code', 'sys_remark']):
            return await self.json_response(msg[10007])
        
        code = data['code']
        sys_remark = data['sys_remark']
        
        # 使用锁，5s使用自旋锁, 防止取消的同时回调
        count_circle = 0
        while True:
            busy_key = 'grab_df_{code}'.format(code=code)
            if await self.redis.setnx(busy_key, 1):
                await self.redis.expire(busy_key, 10)
                break
            if count_circle >= 25:
                self.logger.warning('code:{}有其他进程正在处理中'.format(code))
                return await self.json_response(msg[10032])
            time.sleep(0.2)
            count_circle = count_circle + 1

        # 查找订单
        sql_select_order = """select * from orders_df where code=%s and status != -1 order by id desc limit 1"""
        # 查询有效子单
        sql_select_orders_child = """SELECT code, amount, merchant_id FROM orders_df WHERE parent_id = %s AND status NOT IN (-1, -2, 3)"""
        # 查找流水
        sql_select_record = """select amount,user_type,user_id,record_type from balance_record where code=%s"""
        # 更新系统余额
        sql_update_payment = """update payment set sys_balance=sys_balance+%s where id=%s"""
        # 更新订单-驳回调
        sql_update_cancel = """UPDATE orders_df SET status=-2, sys_remark=%s WHERE code=%s AND status NOT IN (-1, -2) limit 1"""
        # 更新订单-取消&删除
        sql_update_cancel_n_del = """UPDATE orders_df SET status=-1, sys_remark=%s, is_del = 1 WHERE code=%s AND status NOT IN (-1, -2) limit 1"""
        
        async with self.application.db.acquire() as conn:
            async with conn.cursor(DictCursor) as cur:
                try:
                    # 查询订单
                    if not await cur.execute(sql_select_order, code):
                        return await self.json_response(msg[10007])
                    
                    order = (await cur.fetchall())[0]
                    
                    status = int(order.get('status'))
                    is_split = order.get('is_split')
                    parent_id = order.get('parent_id')
                    
                    # 先给-1， 1=常规订单，2=拆单主单，3=拆单子单
                    order_type = -1
                    if not is_split and not parent_id:
                        order_type = 1
                    if is_split and not parent_id:
                        order_type = 2
                    if parent_id:
                        order_type = 3
                        
                    self.logger.info(f"订单 {code} 类型: {order_type}")
                    
                    # 前置逻辑, 代付订单一旦产生, 商户即扣款
                    # 订单状态 0派单中, 1待支付, 2待确认, 3回调中, 4已完成, 5异常按成功处理, -1已取消, -2驳回调
                    
                    # #328-需求变化, 只要有【驳回】权限, 同时也能进入【子订单查看】, 也就可以操作子单
                    # 子单 0|1 的情况进行驳回: 金额退回母订单余额, 母订单可继续拆分
                    # 子单 3|4 的情况进行驳回：子单状态变为 -2, 不通知商家流水, 不产生流水, 驳回金额手动退回商家
                                        
                    # 取消/驳回 常规单
                    # 如果状态是 0|1, 变更为 -2, 回退商家（金额&流水）
                    # 如果状态是 3|4|5, 变更为 -2, 回退商家&码商（金额&流水）
                    
                    # 取消/驳回 母单
                    # 如果状态是 0|1, 变更为 -2, 回退商家（金额&流水）
                    # 如果状态是 3|4|5, 变更为 -2, 回退商家（金额&流水）
                    
                    # 取消/驳回 子单
                    # 如果状态是 0|1, 变更为 -1, 金额归集回母单
                    # 如果状态是 3|4|5, 变更为 -2, 回退码商（金额&流水）

                    # #328-取消逻辑
                    # region 子单权限验证，查询角色的所有权限name
                    # role_permissions_sql = f"select name from permissions where status = 1 and id in ({self.current_user['permissions']})"
                    # role_permission_names_json = await self.query(role_permissions_sql)
                    # # 字典对象 转 集合
                    # role_permission_names = set(value for dictionary in role_permission_names_json for value in dictionary.values())
                    # if ("代付拆单" not in role_permission_names):
                    #         return await self.json_response(msg[10257])
                    # endregion
                    
                    # region 查流水
                    await cur.execute(sql_select_record, code)
                    all_records = await cur.fetchall()
                    mer_records = [record for record in all_records if record['user_type'] == 1]
                    prn_records = [record for record in all_records if record['user_type'] == 0]
                    # endregion
                    
                    # #328-取消逻辑
                    # region 子单退款给商户
                    # merchant_id_for_refund = order.get('merchant_id')
                    # if not merchant_id_for_refund:
                    #     self.logger.warning(f"子订单 {code} 找不到对应的商户ID，无法退款。")
                    #     await conn.rollback()
                    #     return await self.json_response(msg[10017])
                    # # 执行退款给商户
                    # refund_amount = order['amount']
                    # if not await self.change_balance(conn, cur, 'merchant', merchant_id_for_refund, refund_amount, code, 9,sys_remark):
                    #     self.logger.warning(f"子订单 {code} 退款给商户 {merchant_id_for_refund} 失败。")
                    #     await conn.rollback()
                    #     return await self.json_response(msg[10017]) # 假设这是“余额变更失败”消息
                    # self.logger.info(f"子订单 {code} 金额 {refund_amount} 已退款给商户 {merchant_id_for_refund}。")
                    # 【关键】单独提交子订单的状态更新和退款操作
                    # endregion
                    
                    # region 资金流水处理
                    if order_type == 1:
                        if status in [0, 1, 2, 5]:  # 0,1,2,5状态码商都没有实际付款，只退商户余额
                            for record in mer_records:
                                if not await self.change_balance(conn, cur, 'merchant', record['user_id'], -record['amount'], code, 9, sys_remark):
                                    return await self.json_response(msg[10017])
                        if status in [3, 4]:  # 只有真正成功的订单才退还码商余额
                            for record in all_records:
                                tabel_name = 'merchant' if record['user_type'] else 'partner'
                                if not await self.change_balance(conn, cur, tabel_name, record['user_id'], -record['amount'], code, 9, sys_remark):
                                    return await self.json_response(msg[10017])
                            if order['payment_id']:
                                if not await cur.execute(sql_update_payment, (order['amount'], order['payment_id'])):
                                    await conn.rollback()
                                    return await self.json_response(msg[10017])
                    if order_type == 2:
                        if status in [0, 1, 2, 3, 4, 5]:  # 添加状态2,5：状态2待确认、状态5异常按成功处理的订单
                            for record in mer_records:
                                if not await self.change_balance(conn, cur, 'merchant', record['user_id'], -record['amount'], code, 9, sys_remark):
                                    return await self.json_response(msg[10017])
                    if order_type == 3:
                        if status in [3, 4]:  # 只有真正成功的子单才退还码商余额，状态5码商没有实际付款
                            for record in prn_records:
                                if not await self.change_balance(conn, cur, 'partner', record['user_id'], -record['amount'], code, 9, sys_remark):
                                    return await self.json_response(msg[10017])
                            if order['payment_id']:
                                if not await cur.execute(sql_update_payment, (order['amount'], order['payment_id'])):
                                    await conn.rollback()
                                    return await self.json_response(msg[10017])
                    # endregion
                    
                    # region 母子联动逻辑
                    if order_type == 3:
                        # region 此处主要是统计
                        self.logger.info(f"开始检查母订单 {parent_id} 的活跃子订单数量...")
                        sql_count_active_children = """
                            SELECT COUNT(1) COUNT1 FROM orders_df
                            WHERE parent_id = %s AND status NOT IN (-2)
                        """
                        try:
                            self.logger.info("执行 SQL（统计活跃子订单）：%s", sql_count_active_children.strip())
                            self.logger.info("SQL 参数：(%s,)", parent_id)

                            await cur.execute(sql_count_active_children, (parent_id,))

                            # 这里务必要 await
                            result = await cur.fetchone()
                            self.logger.info("统计结果：%s", result)

                            active_children_count = result['COUNT1']
                            self.logger.info(f"母订单 {parent_id} 剩余活跃子订单数量: {active_children_count}")
                            
                        except Exception as e:
                            self.logger.exception(f"统计母订单 {parent_id} 子订单数量时发生异常: {e}")
                            await conn.rollback()
                            return self.json_response(self.msg[10017])
                        # endregion

                        # #328-取消自动驳回，母单需要自己手动驳回
                        # region 子单影响母单
                        # if active_children_count == 0:
                        #     self.logger.info(f"检测到母订单 {parent_id} 没有活跃子订单，将自动驳回母订单。")

                        #     parent_remark = f"Parent order {parent_id} automatically rejected as last active sub-order {code} was rejected at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                        #     self.logger.info("生成自动驳回母订单备注：%s", parent_remark)

                        #     self.logger.info("执行 SQL（查询母订单）：%s", sql_select_order.strip())
                        #     self.logger.info("SQL 参数：(%s,)", parent_id)
                        #     await cur.execute(sql_select_order, (parent_id,))
                        #     parent_order_data = await cur.fetchone()

                        #     self.logger.info("母订单数据查询结果：%s", parent_order_data)

                        #     if parent_order_data and parent_order_data['status'] not in (-1, -2):
                        #         self.logger.info(f"母订单 {parent_id} 当前状态为 {parent_order_data['status']}，满足自动驳回条件。")

                        #         self.logger.info("执行 SQL（更新母订单状态）：%s", sql_update_order_status_remark.strip())
                        #         self.logger.info("SQL 参数：(%s, %s)", parent_remark, parent_id)

                        #         if not await cur.execute(sql_update_order_status_remark, (parent_remark, parent_id)):
                        #             self.logger.error(f"自动驳回母订单 {parent_id} 失败，准备回滚。")
                        #             await conn.rollback()
                        #             return self.json_response(self.msg[10007])

                        #         self.logger.info(f"母订单 {parent_id} 已自动驳回。提交事务中...")
                        #         # 按流水退款--仅仅是母订单的处理逻辑    子订单先标记  如果都驳回了产生一个总的驳回流水 20250705调整不返回金额
                        #         if not await cur.execute(sql_select_record, parent_id):
                        #             return await self.json_response(msg[10007])
                        #         record = await cur.fetchall()
                        #         for i in record:
                        #             self.logger.info(f"数据明细 {i}...")
                        #             tabel_name = 'merchant' if i['user_type'] else 'partner'
                        #             if not await self.change_balance(conn, cur, tabel_name, i['user_id'], -i['amount'], parent_id, 9, sys_remark):
                        #                 return await self.json_response(msg[10017])
                        #         # 修改卡系统余额
                        #         if order['payment_id']:
                        #             if status in [3, 4]:
                        #                 if not await cur.execute(sql_update_payment, (order['amount'], order['payment_id'])):
                        #                     await conn.rollback()
                        #                     return await self.json_response(msg[10017])
                        #         await conn.commit()
                        #         return await self.json_response(dict(code=20000, msg='驳回成功'))
                        #     else:
                        #         self.logger.info(f"母订单 {parent_id} 不存在或已处于非活跃状态（status={parent_order_data['status'] if parent_order_data else 'N/A'}），无需自动驳回。")
                        # endregion
                        
                        # #328-取消联动驳回
                        # region 母单影响子单
                        # if order.get('is_split') == 1:
                        #     self.logger.info(f"检测到订单 {code} 是母订单，正在尝试级联取消其所有活跃子订单。")
                        #     self._last_sql = sql_select_orders_child % code  # 如果是字符串模板，这里 OK
                        #     await cur.execute(sql_select_orders_child, code)
                        #     child_orders = await cur.fetchall()

                        #     if child_orders:
                        #         child_codes = [co['code'] for co in child_orders]
                        #         placeholders = ', '.join(['%s'] * len(child_codes))  # 构造 SQL IN 占位符字符串

                        #         sys_remark = data.get('sys_remark', f"母订单 {code} 取消导致")
                        #         params = [sys_remark] + child_codes

                        #         sql_update_child = f"""
                        #             UPDATE orders_df
                        #             SET status = -2, sys_remark = %s
                        #             WHERE code IN ({placeholders}) AND status NOT IN (-1, -2)
                        #         """
                        #         self.logger.info(f"正在更新子订单状态: {child_codes}")
                        #         self.logger.info("执行 SQL：%s", sql_update_child)
                        #         self.logger.info("SQL 参数：%s", params)

                        #         if not await cur.execute(sql_update_child, params):
                        #             self.logger.error(f"批量更新子订单 {child_codes} 状态失败。")
                        #             await conn.rollback()
                        #             return self.msg[10017]
                        #         self.logger.info(f"母订单 {code} 的 {len(child_orders)} 个子订单已成功级联取消。")
                        #     else:
                        #         self.logger.info(f"母订单 {code} 没有找到活跃子订单需要级联取消。")
                        # else:
                        #     await conn.commit()
                        #     return await self.json_response(dict(code=20000, msg='驳回成功'))
                        # endregion
                    # endregion
                    
                    # region 状态更新
                    if order_type == 3 and status in [0, 1]:
                        if not await cur.execute(sql_update_cancel_n_del, (sys_remark, code)):
                            await conn.rollback()
                            self.logger.warning(cur._last_executed)
                            return await self.json_response(msg[10017])
                    else:
                        if not await cur.execute(sql_update_cancel, (sys_remark, code)):
                            await conn.rollback()
                            self.logger.warning(cur._last_executed)
                            return await self.json_response(msg[10017])
                    # endregion
                    
                    self.logger.info('更新订单状态%s' % cur._last_executed)
                except Exception as e:
                    self.logger.warning('回退订单失败,code={code},异常={e}'.format(code=code, e=e))
                    await conn.rollback()
                    return await self.json_response(msg[10017])
                else:
                    await conn.commit()
                    # 重新接单
                    if order['payment_id'] and order_type in [1, 3] and not (order_type == 3 and status in [0, 1]):
                        await self.redis.lrem('payment_active_df', 0, order['payment_id'])
                        await self.redis.rpush('payment_active_df', order['payment_id'])
                    # 代付回调-驳回
                    if order_type in [1, 2]:
                        await self.redis.publish('order_df_notify', code)
                    return await self.json_response(dict(code=20000, msg='驳回成功'))



# 代付手动回调
class handleNotifydf(BaseHandler):
    @tornado.web.authenticated
    async def post(self):
        data = json.loads(self.request.body)
        
        # #328 由前端控制了
        # if '_' in data['code']:
        #     self.logger.warning(f"订单 {data['code']} 包含下划线，不允许直接操作。")
        #     return await self.json_response(data=msg[10007])
        
        await self.redis.publish('order_df_notify', data['code'])
        result = dict(code=20000, msg='操作成功')
        return await self.json_response(result)

# 获取三方代付列表
class getThirdPays(BaseHandler):
    @tornado.web.authenticated
    async def post(self):
        data = json.loads(self.request.body)
        if await self.is_null(data, ['size', 'page']):
            return await self.json_response(data=msg[10007])
        keys = {'id', 'pay_name', 'pay_name_zh'}
        data_r, total = await self.get_result('third_pay_df', keys, None, data['serchData'], None, data['size'], data['page'])
        result = dict(code=20000, data=data_r, total=total, msg='获取成功')
        return await self.json_response(result)

# 批量三方代付
class handleBatchThirdpay(BaseHandler):
    @tornado.web.authenticated
    async def post(self):
        data = json.loads(self.request.body)
        if await self.is_null(data, ['id', 'codes']):
            return await self.json_response(data=msg[10007])
        thirdPay = await self.get_result_by_condition('third_pay_df', {"*"}, {'id': data['id']})
        if not thirdPay:
            return await self.json_response(msg[10218])
        if not thirdPay['status']:
            return await self.json_response(msg[10219])

        for i in data['codes']:
            # 查询订单状态
            order = await self.get_result_by_condition('orders_df', {'status', 'is_split'}, {'code': i})
            is_split_value = order['is_split']
            # 变更后: 添加日志，指示将进行跳过逻辑判断
            self.logger.info(f"[{i}] 检查订单 {i} 是否需要跳过派单：状态={order['status']}，is_split={is_split_value}。")
            # 只有当订单状态为1或2 并且 is_split不为1 (即is_split为0或None) 时才跳过
            if order['status'] in (1, 2) and is_split_value != 1:
                self.logger.info(f"[{i}] 跳过订单：当前状态为 {order['status']} 且为非拆单订单 (is_split={is_split_value})，不满足派单条件。")
                continue

            await self.redis.publish('batch_Thirdpay_df', '{id}_{code}'.format(id=data['id'], code=i))
            # 检查是否发布成功
            self.logger.info('publish代付订单{code} {id} {name}'.format(code=i, id=thirdPay['id'], name=thirdPay['pay_name_zh']))
        return await self.json_response(msg[20000])

# 上传凭证
class upload(BaseHandler):
    @tornado.web.authenticated
    async def post(self):
        try:
            files = self.request.files
            arguments = self.request.arguments
            if not self.request.files or len(files) > 1 or not arguments:
                return await self.json_response(msg[10007])
            if files.get('image'):
                files = files['image'][0]
                if not files.filename.split('.')[1].lower() in ['jpg', 'jpeg', 'png', 'bmp', 'tif', 'tiff']:
                    return await self.json_response(msg[10007])
                filename = arguments['code'][0].decode('utf-8')
                with open("static/upload/{}.jpg".format(filename), 'wb') as f:
                    f.write(files['body'])
                # Define the full file path
                file_path = "static/upload/{}.jpg".format(filename)
                result = dict(code=20000, path=file_path, msg='上传成功')
                return await self.json_response(result)
            elif files.get('file'):
                files = files['file'][0]
                filename = files.filename.rsplit('.', 1)[0].lower()
                file_format = files.filename.rsplit('.', 1)[1].lower()
                if file_format in ['csv']:
                    filename = '{}_{}'.format(filename, arguments['code'][0].decode('utf-8')).lower()
                    with open("static/upload/df/{}.csv".format(filename), 'wb') as f:
                        f.write(files['body'])
                    return await self.json_response(msg[20000])
                if file_format in ['xls']:
                    filename = '{}_{}'.format(filename, arguments['code'][0].decode('utf-8')).lower()
                    with open("static/upload/df/{}.xls".format(filename), 'wb') as f:
                        f.write(files['body'])
                    return await self.json_response(msg[20000])
                if file_format in ['xlsx']:
                    filename = '{}_{}'.format(filename, arguments['code'][0].decode('utf-8')).lower()
                    with open("static/upload/df/{}.xlsx".format(filename), 'wb') as f:
                        f.write(files['body'])
                    return await self.json_response(msg[20000])
            else:
                return await self.json_response(msg[10007])
        except Exception as e:
            self.logger.exception(e)
            return await self.json_response(msg[10007])


# 批量上传凭证
class UploadReceiptBatch(BaseHandler):
    def __init__(self, *args, **kwargs):
        # 后续需要解析的方法可在此添加：
        # '银行名': '解析方法',
        self.bank_handle_map = {
            'IOB BANK': self.extract_iob_data,
            'FINANA BANK': self.extract_finana_data,
            'YES BANK': self.extract_yes_data,  # 新增YES BANK的映射
            'CANARA BANK': self.extract_canara_data,
            'ICICI BANK': self.extract_icici_data,
            'BANDHAN BANK': self.extract_bandhan_data,
            'IDFC BANK': self.extract_idfc_data,
            'haoda': self.extract_haoda_data,
            'KOTAK BANK': self.extract_kotak_data,
            'AXIS BANK': self.extract_axit_data,
            'FEDERAL BANK': self.extract_federal_data,
            'ESAF SMALL BANK': self.extract_esaf_data,
            'IOB NEW BANK': self.extract_iob_new_data,
            'KVB BANK': self.extract_kvb_data,
            'KVB IMPS': self.extract_kvb_imps_data,
            'YES BUSINESS': self.extract_yes_business_data,
            'AXIS NEO BANK': self.extract_axit_neo_data,
            'AXIS NEO NEFT BANK': self.extract_axit_neo_neft_data,
            'AU BANK': self.extract_au_bank_data,
            'marspay': self.extract_marspay_data,
        }
        super().__init__(*args, **kwargs)

    # _update_child_orders_status 
    @staticmethod
    async def _update_child_orders_status(self, parent_order_identifier: str, bank_name: str):
        """
        更新指定母订单标识符下的未完成子订单状态为 4。

        Args:
            parent_order_identifier: code
            bank_name: 银行名称，用于日志记录。
        """
        if not parent_order_identifier:
            self.logger.warning(f"尝试更新子订单但父订单标识符为空，银行: {bank_name}")
            return

        sql_update_sub_orders = f"""
            UPDATE orders_df
            SET status = 4
            WHERE parent_id = '{parent_order_identifier}'
            AND is_split = 0  -- 确保是子订单
            AND status != 4   -- 只更新未完成的子订单
        """
        # 1. 查询当前订单（即 parent_order_identifier 对应的订单）的类型信息
        #   - is_split: 标记是否为拆分订单（母订单通常设为1，子订单为0）
        #   - parent_id: 如果是子订单，则指向其母订单的 code
        sql_check_order_type = f"""
            SELECT is_split, parent_id
            FROM orders_df
            WHERE code = '{parent_order_identifier}'
            LIMIT 1
        """
        order_info_result = await self.execute(sql_check_order_type)

        # 检查查询结果是否存在且有效
        if order_info_result and hasattr(order_info_result, 'rows') and len(order_info_result.rows) > 0:
            order_details = order_info_result.rows[0]
            is_split_flag = order_details.get('is_split') # 获取 is_split 标记
            actual_parent_id_of_found_order = order_details.get('parent_id') # 如果是子订单，获取其母订单 ID

            # 情况 A: 当前找到的 'code' (parent_order_identifier) 是一个母订单
            # 假设 is_split = 1 标识一个原始的、已创建子订单的母订单
            if is_split_flag == 1:
                self.logger.info(f"订单 {parent_order_identifier} 被识别为母订单，将更新其子订单。")
                # 执行更新子订单的操作
                sub_update_result = await self.execute(sql_update_sub_orders)
                if sub_update_result:
                    self.logger.info(f"{bank_name} 子订单更新成功, 母订单标识: {parent_order_identifier}")
                else:
                    self.logger.error(f"{bank_name} 子订单更新失败, 母订单标识: {parent_order_identifier}, 操作人: {self.current_user['id']}")

            # 情况 B: 当前找到的 'code' (parent_order_identifier) 是一个子订单
            # 假设 is_split = 0 且 parent_id 存在标识一个子订单
            elif is_split_flag == 0 and actual_parent_id_of_found_order:
                self.logger.info(f"订单 {parent_order_identifier} 被识别为子订单。")
                # 检查其同级子订单中是否有未完成的
                # 统计所有与当前子订单同母订单下，且状态不是4（未完成），且不是当前子订单本身的子订单数量
                sql_count_unfinished_siblings = f"""
                    SELECT COUNT(*) AS unfinished_count
                    FROM orders_df
                    WHERE parent_id = '{actual_parent_id_of_found_order}'
                    AND is_split = 0   -- 确保只计算子订单
                    AND status != 4    -- 统计未完成的订单
                    AND code != '{parent_order_identifier}' -- 排除当前已完成的子订单本身
                """
                siblings_count_result = await self.execute(sql_count_unfinished_siblings)

                if siblings_count_result and hasattr(siblings_count_result, 'rows') and len(siblings_count_result.rows) > 0:
                    unfinished_siblings_count = siblings_count_result.rows[0].get('unfinished_count')

                    # 如果没有其他未完成的兄弟订单，说明当前子订单是最后一单
                    if unfinished_siblings_count == 0:
                        self.logger.info(f"子订单 {parent_order_identifier} 完成，其母订单 {actual_parent_id_of_found_order} 的所有子订单均已完成，将更新母订单状态。")
                        # 调用 BaseOrderHandler 的方法来更新母订单状态
                        # 这里的 actual_parent_id_of_found_order 应该是母订单的 code
                        await BaseOrderHandler.order_success(self, actual_parent_id_of_found_order, 1)
                    else:
                        self.logger.info(f"子订单 {parent_order_identifier} 完成，但仍有 {unfinished_siblings_count} 个兄弟订单未完成。")
                else:
                    self.logger.error(f"无法获取母订单 {actual_parent_id_of_found_order} 的兄弟订单完成情况。")
            else:
                # 如果 is_split 既不是1也不是0，或者 parent_id 不明确
                self.logger.warning(f"订单 {parent_order_identifier} 类型不明确 (is_split: {is_split_flag}, parent_id: {actual_parent_id_of_found_order})，未执行子订单或母订单更新逻辑。")
        else:
            self.logger.error(f"无法查询订单 {parent_order_identifier} 的类型信息，可能订单不存在。")      
    
    
    @staticmethod
    async def extract_finana_data(uploaded_name):
        """
        最终返回内容为字典组成的列表:
        [
            {'account': 账号, 'ifsc': ifsc, 'debit_account': 回执金额, 'utr': 回执utr}),
            {'account': 账号, 'ifsc': ifsc, 'debit_account': 回执金额, 'utr': 回执utr}),
            ......
        ]
        """
        order_remark_field = 1
        debit_account_field = 2
        results = []
        uploaded_name = uploaded_name.lower()
        
        # 读取 Excel 文件
        workbook = xlrd.open_workbook("static/upload/df/{}".format(uploaded_name), ignore_workbook_corruption=True)  
        df = pd.read_excel(workbook)

        # 定义两个正则表达式，分别处理 WITHDRAWAL 和 NEFT 格式的数据，正则表达式用于提取 ifsc, account, utr
        pattern_imps = r"WITHDRAWAL IMPS-RRN:(\d+)-To:.*?-A/c:(\d+)-(\w+)"
        pattern_neft = r"NEFT OUT-UTR:(\w+)-To:.*?-A/c:(\d+)-(\w+)NEFT"
        
        for index, row in df.iterrows():
            order_remark = row[order_remark_field];
            # 判断order_remark是否为nan
            if pd.isna(order_remark):
                continue
            
            # 提取 ifsc, account, utrd
            match = re.search(pattern_imps, order_remark)
            
            if not match:
                # 如果WITHDRAWAL格式没有匹配，再尝试匹配NEFT格式
                match = re.search(pattern_neft, order_remark)
            
            if match:
                utr = match.group(1)  # 提取 UTR
                account = match.group(2)  # 提取 Account
                ifsc = match.group(3)  # 提取 IFSC

                # 取出 Balance 列的值作为 debit_account
                debit_account = row[debit_account_field]
                results.append(
                            {'account': account,
                             'ifsc': ifsc,
                             'debit_account': debit_account,
                             'utr': utr})
        return results

    @staticmethod
    async def extract_iob_data(uploaded_name):
        """
        最终返回内容为字典组成的列表:
        [
        {'code': '订单code1', 'debit_account': '回执账号1', 'utr': '回执utr1'},
        {'code': '订单code2', 'debit_account': '回执账号2', 'utr': '回执utr2'},
        ......
        ]
        """
        order_code_field = 'Remarks'
        debit_account_field = 'A/c Number'
        utr_field = 'UTR No'
        results = []
        uploaded_name = uploaded_name.lower()
        with open("static/upload/df/{}".format(uploaded_name), 'rb') as f:
            # data = f.read()
            line = 1
            while True:
                line_data = f.readline()
                if line_data:
                    line_data = line_data.decode('utf-8')
                    if line == 1:
                        field_list = line_data.split(',')
                        print(field_list)
                        order_code_index = field_list.index(order_code_field)
                        debit_account_index = field_list.index(debit_account_field)
                        utr_index = field_list.index(utr_field)
                    else:
                        line_data = line_data.split(',')
                        results.append(
                            {'code': line_data[order_code_index],
                             'debit_account': line_data[debit_account_index],
                             'utr': line_data[utr_index]})
                    line += 1
                else:
                    break
        return results

    @staticmethod
    async def extract_yes_databak(uploaded_name):
        """
        解析 YES BANK 上传的回执 Excel 文件，返回字典列表：
        [
            {'debit_account': '回执账号', 'utr': '回执utr', 'account': '收款账号', 'ifsc': '银行IFSC'},
            ...
        ]
        """
        # 定义列名称，使用图片中的实际列名
        debit_account_field = 'Debit Account Number'
        utr_field = 'RBI/UTR Reference Number'
        account_field = 'Beneficiary Account Number'
        ifsc_field = 'Benficiary Bank IFSC'
        amount_field = 'Amount'
        code_field = 'Customer Reference Number'
       
        # 文件路径
        file_path = f"static/upload/df/{uploaded_name}"

        try:
            # 读取 Excel 文件，跳过前5行，并只选择需要的列（从第二列开始）
            df = pd.read_excel(file_path, header=5, usecols="B:J")
            results = []
            for _, row in df.iterrows():
                result = {
                    'debit_account': row.get(debit_account_field),
                    'utr': row.get(utr_field),
                    'account': row.get(account_field),
                    'ifsc': row.get(ifsc_field),
                    'amount': row.get(amount_field),
                    'code': row.get(code_field)
                }
                results.append(result)

            # print('最终yes bank的导入信息列表', results)
            return results

        except Exception as e:
            print(f"读取文件时发生错误: {e}")
            return []
            
    @staticmethod
    async def extract_yes_data(uploaded_name):
        """
        解析 YES BANK 上传的回执 Excel 文件，返回字典列表：
        [
            {'utr': '回执utr', 'code': '交易编码'},
            ...
        ]
        如果 Transaction Type 字段存在，仅筛选 Transaction Type != 'CREDIT' 的记录。
        如果字段不存在，根据新的规则解析数据。
        """
        # 定义列名称
        description_field = 'Description'  # 包含 utr 和 code 的字段
        amount_field = 'Amount'  # 金额字段
        transaction_type_field = 'Transaction Type'  # 交易类型字段
        utr_field = 'RBI/UTR Reference Number'  # UTR 字段
        code_field = 'Customer Reference Number'  # Code 字段

        # 文件路径
        file_path = f"static/upload/df/{uploaded_name}"

        try:
            # 读取 Excel 文件
            df = pd.read_excel(file_path, header=4)

            # 检查 Transaction Type 字段是否存在
            if transaction_type_field in df.columns:
                # 筛选出 Transaction Type != 'CREDIT' 的记录
                filtered_df = df[df[transaction_type_field] != 'CREDIT']

                # 结果列表
                results = []
                for _, row in filtered_df.iterrows():
                    description = row.get(description_field, '')
                    amount_raw = row.get(amount_field, '')
                    # 检查是否是以 INR 开头的金额
                    if not amount_raw or not str(amount_raw).strip():
                        # 提取 Description 中的相关信息
                        if 'NEFT' not in description:
                            # 如果 description 不包含 'NEFT'，跳过当前循环
                            continue
                        utr_match = re.match(r'NEFT O/W-(YESBN\d+)-([A-Z0-9]+)-([\w\s]+)', description)
                        amount_raw = row.get(transaction_type_field, '')
                        # Regular expression to extract numeric part
                        amount_raw = re.sub(r'[^\d.]', '', amount_raw)  # This removes everything that's not a digit or a decimal point
                        amount_raw = amount_raw.replace(',', '').replace(' ', '')
                        if utr_match:
                            account = utr_match.group(3).strip()  # 提取账户名并去除多余空格
                            ifsc = utr_match.group(2)  # 提取 IFSC
                            utr = utr_match.group(1)  # 提取 UTR
                            # 构建结果字典
                            result = {
                                'account': account,
                                'ifsc': ifsc,
                                'amount': amount_raw,
                                'utr': utr
                            }

                            # 添加到结果列表
                            results.append(result)
                    # 如果不符合 INR 条件，则执行其他逻辑
                    else:
                        utr_match = re.search(r'RRN:(\d+)', description)  # 提取 UTR
                        code_match = re.search(r'/([A-Z0-9]+)$', description)  # 提取 code
                        result = {
                            'utr': utr_match.group(1) if utr_match else None,
                            'code': code_match.group(1) if code_match else None,
                        }
                        results.append(result)

            else:
                # 读取 Excel 文件
                df = pd.read_excel(file_path, header=5)
                # 如果 Transaction Type 字段不存在，直接按照图中规则解析
                print(f"警告: '{transaction_type_field}' 字段不存在，按新规则解析数据")

                # 结果列表
                results = []
                for _, row in df.iterrows():
                    utr = row.get(utr_field, None)
                    code = row.get(code_field, None)
                    if not code:  # 检查 code 是否为空或 None
                        account = row.get('Beneficiary Name', '')
                        ifsc = row.get('Benficiary Bank IFSC', '')
                        amount = row.get('Amount', '')
                        utr = row.get('RBI/UTR Reference Number', '')
                        # 构建结果字典
                        result = {
                            'account': account,
                            'ifsc': ifsc,
                            'amount': str(amount).replace(',', '').replace(' ', ''),
                            'utr': utr
                        }

                        # 添加到结果列表
                        results.append(result)
                    else:
                        result = {
                            'utr': utr,
                            'code': code,
                        }
                        results.append(result)

            # 打印 results
            print("解析结果 (results):", results)
            print("解析结果 (results) 记录数:", len(results))
            return results

        except Exception as e:
            print(f"读取文件时发生错误: {e}")
            return []

    @staticmethod
    async def extract_canara_data(uploaded_name):
        """
        最终返回内容为字典组成的列表:
        [
        {'name': '姓名', 'account': '回执账号1', 'amount': '金额', 'utr': '回执utr1'},
        ......
        ]
        """
        detail_field = 'Credit Account Details'
        amount_field = 'Amount'
        utr_field = 'External Reference IdRecord Status'
        results = []
        uploaded_name = uploaded_name.lower()
        detail_field_index = amount_field_index = utr_field_index= None
        with open("static/upload/df/{}".format(uploaded_name), 'rb') as f:
            # data = f.read()
            line = 1
            while True:
                line_data = f.readline()
                if line_data:
                    line_data = line_data.decode('utf-8')
                    # 找对应列的位置
                    if detail_field_index is None or amount_field_index is None or utr_field_index is None :
                        field_list = line_data.split(',')
                        print(field_list)
                        try:
                            detail_field_index = field_list.index(detail_field)
                            amount_field_index = field_list.index(amount_field)
                            utr_field_index = field_list.index(utr_field)
                        except ValueError:
                            continue
                    
                    line_data = line_data.split(',')
                    # 匹配detail，提取卡号姓名
                    pattern = r"(\d+)\s+([A-Za-z\s]+)"
                    match = re.match(pattern, line_data[detail_field_index])
                    if match:
                        account = match.group(1)  # 提取卡号
                        name = match.group(2)         # 提取姓名
                        
                        results.append(
                            {'account': account,
                            'name': name,
                            'amount': line_data[amount_field_index],
                            'utr': line_data[utr_field_index]})
                else:
                    break
        return results

    @staticmethod
    async def extract_icici_data(uploaded_name):
        """
        最终返回内容为字典组成的列表:
        [
            {'code': 订单号, 'utr': 回执utr}),
            {'code': 订单号, 'utr': 回执utr}),
            ......
        ]
        """
        remark_field = 'Remark'
        utr_field = 'UTR NO'
        credit_narration_field = 'Credit narration'  # 新增加的字段
        results = []
        uploaded_name = uploaded_name.lower()

        # 读取 Excel 文件
        try:
            df = pd.read_excel("static/upload/df/{}".format(uploaded_name), engine='openpyxl')
            logging.info(f"成功读取文件: {uploaded_name}")
        except Exception as e:
            logging.error(f"读取文件失败: {e}")
            return []

        for index, row in df.iterrows():
            # 获取订单编号的优先顺序，首先检查 Credit narration 是否有值，如果没有再使用 Remark
            code = row.get(credit_narration_field) or row.get(remark_field)
            
            # 判断如果 code 是 nan 则跳过
            if pd.isna(code):
                logging.info(f"第 {index + 1} 行的订单编号为空，跳过该行")
                continue
            
            # 获取 UTR
            utr = row.get(utr_field)
            if pd.isna(utr):
                logging.info(f"第 {index + 1} 行的UTR号为空，跳过该行")
                continue
            
            # 添加到结果中
            results.append({'code': code, 'utr': utr})
            logging.info(f"第 {index + 1} 行 - 订单号: {code}, UTR: {utr}")

        logging.info(f"提取完成，共处理 {len(results)} 条数据")
        return results
    
    @staticmethod
    async def extract_bandhan_data(uploaded_name):
        """
        最终返回内容为字典组成的列表:
        [
        {'code': '代付订单号', 'status': '代付订单状态', 'amount': '金额', 'utr': '回执utr'},
        ......
        ]
        """
        amount_field = 'Amount'
        utr_field = 'CoreRefNumber'
        orderNo_field = 'DestinationNarration'
        status_field = 'Status'
        results = []
        uploaded_name = uploaded_name.lower()
        orderNo_field_index = amount_field_index = utr_field_index = status_field_index = None
        with open("static/upload/df/{}".format(uploaded_name), 'rb') as f:
            # data = f.read()
            line = 1
            while True:
                line_data = f.readline()
                if line_data:
                    line_data = line_data.decode('utf-8')
                    # 找对应列的位置
                    if orderNo_field_index is None or amount_field_index is None or utr_field_index is None or status_field_index is None :
                        field_list = line_data.split(',')
                        try:
                            orderNo_field_index = field_list.index(orderNo_field)
                            amount_field_index = field_list.index(amount_field)
                            utr_field_index = field_list.index(utr_field)
                            status_field_index = field_list.index(status_field)
                        except ValueError:
                            continue
                    elif line > 1:          
                        line_data = line_data.split(',')
                        results.append(
                            {
                                'code': line_data[orderNo_field_index],
                                'amount': line_data[amount_field_index],
                                'status': line_data[status_field_index],
                                'utr': line_data[utr_field_index]
                            }
                        )
                    line += 1
                else:
                    break
        return results

    @staticmethod
    async def extract_idfc_data(uploaded_name):
        """
        最终返回内容为字典组成的列表:
        [
            {'code': '订单号', 'utr': '回执utr'},
            {'code': '订单号', 'utr': '回执utr'},
            ......
        ]
        """
        remarks_field = 'Remarks'  # Remarks字段对应的列（订单号）
        utr_field = 'UTR Number'  # UTR Number字段对应的列（回执UTR）
        results = []  # 用于存放结果的列表
        uploaded_name = uploaded_name.lower()  # 文件名转小写，确保一致性

        # 读取 Excel 文件
        file_path = f"static/upload/df/{uploaded_name}"  # 构造文件路径
        try:
            # 使用 pandas 读取 Excel 文件
            df = pd.read_excel(file_path, engine='openpyxl')
        except Exception as e:
            print(f"读取文件时出错: {e}")
            return []  # 如果读取失败，返回空列表

        # 遍历每一行数据，提取需要的字段
        for index, row in df.iterrows():
            # 获取 'Remarks' 列的值，即订单号
            code = row[remarks_field]

            # 判断 'Remarks' 是否为 NaN
            if pd.isna(code):
                continue  # 如果为空，则跳过此行

            # 提取 'UTR Number' 字段的值
            utr = row[utr_field]

            # 将提取到的数据添加到结果列表
            results.append({'code': code, 'utr': utr})

        return results  # 返回最终结果


    @staticmethod
    async def extract_haoda_data(uploaded_name):
        """
        最终返回内容为字典组成的列表:
        [
        {'code': '订单code1', 'utr': '回执utr1'},
        {'code': '订单code2', 'utr': '回执utr2'},
        ......
        ]
        """
        order_code_field = 'Payout Reference'
        utr_field = 'Transaction Reference'
        results = []
        uploaded_name = uploaded_name.lower()
        with open("static/upload/df/{}".format(uploaded_name), 'rb') as f:
            line = 1
            while True:
                line_data = f.readline()
                if line_data:
                    line_data = line_data.decode('utf-8')
                    if line == 1:
                        field_list = line_data.split(',')
                        print(field_list)
                        order_code_index = field_list.index(order_code_field)
                        utr_index = field_list.index(utr_field)
                    else:
                        line_data = line_data.split(',')
                        results.append(
                            {'code': line_data[order_code_index],
                             'utr': line_data[utr_index]})
                    line += 1
                else:
                    break
        return results
    @staticmethod
    async def extract_kotak_data(uploaded_name):
        """
        最终返回内容为字典组成的列表:
        [
            {'code': '订单号', 'utr': '回执utr'},
            {'code': '订单号', 'utr': '回执utr'},
            ...
        ]
        """
        # 新的字段名对应表格中的列名
        remarks_field = 'Instrument Payment Ref No'  # 'Instrument Payment Ref No'对应订单号
        utr_field = 'UTR SrNo'  # 'UTR SrNo'对应回执UTR
        results = []  # 用于存放结果的列表
        uploaded_name = uploaded_name.lower()  # 文件名转小写，确保一致性

        # 读取 Excel 文件
        file_path = f"static/upload/df/{uploaded_name}"  # 构造文件路径
        try:
            # 使用 pandas 读取 Excel 文件
            df = pd.read_excel(file_path, engine='xlrd')
        except Exception as e:
            print(f"读取文件时出错: {e}")
            return []  # 如果读取失败，返回空列表

        # 遍历每一行数据，提取需要的字段
        for index, row in df.iterrows():
            # 获取 'Batch Payment Ref No' 列的值，即订单号
            code = row.get(remarks_field)

            # 判断 'Batch Payment Ref No' 是否为 NaN
            if pd.isna(code):
                continue  # 如果为空，则跳过此行

            # 提取 'UTR SrNo' 字段的值
            utr = row.get(utr_field)

            # 将提取到的数据添加到结果列表
            results.append({'code': code, 'utr': utr})

        return results  # 返回最终结果

    @staticmethod
    async def extract_axit_data(uploaded_name):
        """
        从上传的 Excel 文件中提取数据并返回一个字典列表:
        [
            {'utr_no': 'UTR值', 'Beneficiary': '账号', 'amount': '金额'},
            {'utr_no': 'UTR值', 'Beneficiary': '账号', 'amount': '金额'},
            ...
        ]
        """
        # 定义 Excel 文件中的列名
        utr_field = 'UTR No'  # UTR 列
        beneficiary_field = 'Beneficiary'  # 账号
        amount_field = 'Amount'  # 金额列
        results = []  # 初始化结果列表

        # 文件路径构造
        uploaded_name = uploaded_name.lower()  # 文件名转小写，确保一致性

        # 读取 Excel 文件
        file_path = f"static/upload/df/{uploaded_name}"  # 构造文件路径

        try:
            # 使用 pandas 读取 Excel 文件
            df = pd.read_excel(file_path, engine='openpyxl')  # 修改为 openpyxl 以兼容 .xlsx 文件
        except FileNotFoundError:
            logging.error(f"文件未找到: {file_path}")
            return []
        except Exception as e:
            logging.error(f"读取文件时出错: {e}")
            return []

        # 检查字段是否存在
        missing_columns = [col for col in [utr_field, beneficiary_field, amount_field] if col not in df.columns]
        if missing_columns:
            logging.error(f"文件中缺少必要的列: {', '.join(missing_columns)}")
            return []

        # 遍历每一行数据
        for index, row in df.iterrows():
            # 提取字段数据
            utr_no = row.get(utr_field, "")
            Beneficiary = row.get(beneficiary_field, "")
            amount = row.get(amount_field, 0)

            # 检查金额是否为有效数字
            try:
                # 尝试将金额转换为浮点数
                amount = float(amount)
            except (ValueError, TypeError):
                # 如果金额字段无效（不是数字或为 None），跳过此行
                continue

            # 如果所有字段为空，跳过此行
            if pd.isna(utr_no) and pd.isna(Beneficiary) and pd.isna(amount):
                logging.debug(f"第 {index} 行数据为空，跳过")
                continue

            # 添加到结果列表
            results.append({
                'utr': str(utr_no).strip(),
                'beneficiary': str(Beneficiary).strip(),
                'amount': str(amount).strip()
            })

        logging.info(f"成功从文件提取 {len(results)} 条记录")
        return results

    @staticmethod
    async def extract_federal_data(uploaded_name):
        """
        从上传的 Excel 文件中提取数据并返回一个字典列表:
        [
            {'utr_no': 'UTR值', 'beneficiary': '账号', 'amount': '金额'},
            {'utr_no': 'UTR值', 'beneficiary': '账号', 'amount': '金额'},
            ...
        ]
        """
        # 定义 Excel 文件中的列名
        utr_field = 'RBI/UTR Reference Number'  # UTR 列
        beneficiary_field = 'Beneficiary Account Number'  # 受益人账号列
        amount_field = 'Amount'  # 金额列
        results = []  # 初始化结果列表

        # 文件路径构造
        uploaded_name = uploaded_name.lower()  # 文件名转小写，确保一致性
        file_path = f"static/upload/df/{uploaded_name}"  # 构造文件路径

        # print(f"正在读取文件: {file_path}")

        try:
            # 根据文件扩展名选择引擎
            if file_path.endswith('.xls'):
                # df = pd.read_excel(file_path, engine='xlrd')  # 使用 xlrd 处理旧格式 .xls
                 # 读取 Excel 文件
                df = pd.read_excel(file_path, engine='xlrd', header=5)
            elif file_path.endswith('.xlsx'):
                df = pd.read_excel(file_path, engine='openpyxl')  # 使用 openpyxl 处理 .xlsx
            elif file_path.endswith('.csv'):
                # 读取 CSV 文件
                df = pd.read_csv(file_path)  # 你可以修改 encoding 适配不同编码格式
            else:
                raise ValueError("不支持的文件格式")
            # print(f"成功读取文件，共有 {len(df)} 行数据")
        except FileNotFoundError:
            print(f"错误: 文件未找到 {file_path}")
        except ValueError as ve:
            print(f"错误: {ve}")
        except Exception as e:
            print(f"错误: 无法读取文件，原因: {e}")

        # 检查字段是否存在
        # print(f"检查列是否存在: {utr_field}, {beneficiary_field}, {amount_field}")
        missing_columns = [col for col in [utr_field, beneficiary_field, amount_field] if col not in df.columns]
        if missing_columns:
            logging.error(f"文件中缺少必要的列: {', '.join(missing_columns)}")
            # print(f"错误: 缺少必要列 {missing_columns}")
            # return []
            # 尝试从 header=10 开始读取
            # 备用列名（如果上面缺失）
            particulars_field = 'Particulars'
            withdrawal_field = 'Withdrawal'
            try:
                # df = pd.read_excel(file_path, engine='xlrd', header=10)
                df = pd.read_csv(file_path, skiprows=10)  # 使用 pandas 读取 CSV 文件
            except Exception as e:
                logging.error(f"错误: 备用方式读取 Excel 失败，原因: {e}")
                return []
            
            # 确保备用列存在
            if particulars_field not in df.columns or withdrawal_field not in df.columns:
                logging.error("错误: 备用列 'Particulars' 或 'Withdrawal' 也不存在，无法解析")
                return []
            
            # 遍历数据行
            for index, row in df.iterrows():
                particulars = row.get(particulars_field, "")
                amount = row.get(withdrawal_field, 0)

                # 提取 UTR 和 Beneficiary
                utr = ""
                beneficiary = ""

                if pd.notna(particulars):
                    # 提取 UTR (匹配 FDRLM+数字)
                    utr_match = re.search(r'FDRLM\d+', particulars)
                    utr = utr_match.group(0) if utr_match else ""

                    # 去掉 "NFT/" (可能存在) 和 提取 Beneficiary 名称
                    particulars_cleaned = re.sub(r'NFT/', '', particulars).strip()  # 移除 NFT/
                    particulars_cleaned = re.sub(r'FDRLM\d+', '', particulars_cleaned).strip()  # 移除 UTR 编号
                    particulars_cleaned = re.sub(r'\b[A-Z]{4,}\d+[A-Z]*\b', '', particulars_cleaned).strip()  # 移除银行代码
                    
                    beneficiary = ' '.join(particulars_cleaned.split())  # 规范空格格式

                # 确保金额有效
                try:
                    amount = float(str(amount).replace(",", "").replace(" ", "").strip())
                    if amount <= 0 or pd.isna(amount):  # 如果金额无效或 NaN，则跳过
                        continue
                except ValueError:
                    continue
                
                # 如果 UTR 和 Beneficiary 都为空，则跳过
                if not utr and not beneficiary:
                    return None
                
                results.append({
                    'utr': utr,
                    'payment_name': beneficiary,
                    'beneficiary': '',
                    'amount': round(amount, 2)
                })

            logging.info(f"成功从文件提取 {len(results)} 条记录")
            return results

        # print(f"所有必要列存在: {utr_field}, {beneficiary_field}, {amount_field}")

        # 遍历每一行数据
        for index, row in df.iterrows():
            # print(f"处理第 {index + 1} 行数据: {row.to_dict()}")

            # 提取字段数据
            utr_no = row.get(utr_field, "")
            beneficiary = row.get(beneficiary_field, "")
            # amount = row.get(amount_field, 0)
            amount = str(row.get(amount_field, 0)).replace(",", "").replace(" ", "")  # 去除逗号

            # print(f"提取数据 - UTR: {utr_no}, Beneficiary: {beneficiary}, Amount: {amount}")

            # 检查金额是否为有效数字
            try:
                # 尝试将金额转换为浮点数
                amount = float(amount)
            except (ValueError, TypeError):
                # print(f"跳过第 {index + 1} 行: 无效的金额 {amount}")
                continue
            
            # 检查金额是否小于等于 0
            if pd.isna(amount) or amount <= 0:
                # print(f"跳过第 {index + 1} 行: 金额无效或不大于 0 -> {amount}")
                continue

            # 如果所有字段为空，跳过此行
            if pd.isna(utr_no) and pd.isna(beneficiary) and pd.isna(amount):
                logging.debug(f"第 {index} 行数据为空，跳过")
                # print(f"跳过第 {index + 1} 行: 所有字段均为空")
                continue

            # 添加到结果列表
            results.append({
                'utr': str(utr_no).strip(),
                'payment_name': '',
                'beneficiary': str(beneficiary).strip(),
                'amount': round(amount, 2)  # 保留两位小数
            })
            # print(f"第 {index + 1} 行数据已添加: UTR={utr_no}, Beneficiary={beneficiary}, Amount={amount}")

        # print(f"文件处理完成，共提取 {len(results)} 条有效记录")
        logging.info(f"成功从文件提取 {len(results)} 条记录")
        return results
    
    @staticmethod
    async def extract_esaf_data(uploaded_name):
        """
        从上传的 Excel 文件中提取数据并返回一个字典列表:
        [
            {'utr': 'UTR值', 'beneficiary': '账号', 'amount': '金额', 'code': '账户类型'},
            {'utr': 'UTR值', 'beneficiary': '账号', 'amount': '金额', 'code': '账户类型'},
            ...
        ]
        """
        # 定义 Excel 文件中的列名
        utr_field = 'coreRefNumber'  # UTR 列 (corresponding to coreRefNumber)
        beneficiary_field = 'beneficiaryName'  # 受益人账户列 (corresponding to beneficiaryName)
        amount_field = 'amount'  # 金额列
        code_field = 'beneficiaryAccountType'  # 账户类型列 (corresponding to beneficiaryAccountType)

        results = []  # 初始化结果列表

        # 文件路径构造
        uploaded_name = uploaded_name.lower()  # 文件名转小写，确保一致性

        # 读取 Excel 文件
        file_path = f"static/upload/df/{uploaded_name}"  # 构造文件路径

        try:
            # 使用 pandas 读取 Excel 文件
            # df = pd.read_excel(file_path, engine='openpyxl')  # 使用 openpyxl 以兼容 .xlsx 文件
            # 使用 pandas 读取 CSV 文件
            df = pd.read_csv(file_path)  # 使用 pandas 读取 CSV 文件
        except FileNotFoundError:
            logging.error(f"文件未找到: {file_path}")
            return []
        except Exception as e:
            logging.error(f"读取文件时出错: {e}")
            return []

        # 检查字段是否存在
        missing_columns = [col for col in [utr_field, beneficiary_field, amount_field, code_field] if col not in df.columns]
        if missing_columns:
            logging.error(f"文件中缺少必要的列: {', '.join(missing_columns)}")
            return []

        # 遍历每一行数据
        for index, row in df.iterrows():
            # 提取字段数据
            utr = row.get(utr_field, "")
            beneficiary = row.get(beneficiary_field, "")
            # amount = row.get(amount_field, 0)
            code = row.get(code_field, "")
            amount = str(row.get(amount_field, "")).replace(",", "").replace(" ", "").strip()

            # 检查金额是否为有效数字
            try:
                # 尝试将金额转换为浮点数
                amount = float(amount)
            except (ValueError, TypeError):
                # 如果金额字段无效（不是数字或为 None），跳过此行
                continue

            # 如果所有字段为空，跳过此行
            if pd.isna(utr) and pd.isna(beneficiary) and pd.isna(amount) and pd.isna(code):
                logging.debug(f"第 {index} 行数据为空，跳过")
                continue

            # 添加到结果列表
            results.append({
                'utr': str(utr).strip(),
                'beneficiary': str(beneficiary).strip(),
                'amount': str(amount).strip(),
                'code': str(code).strip()  # 账户类型
            })

        logging.info(f"成功从文件提取 {len(results)} 条记录")
        return results

    @staticmethod
    async def extract_iob_new_data(uploaded_name):
        """
        最终返回内容为字典组成的列表:
        [
            {'payment_name': '名称', 'ifsc': 'Ifsc', 'debit_account': '回执账号', 'utr': 'Utr', 'amount': '金额'},
            ......
        ]
        """
        local_to_table_map = {
            'payment_name': 'Name',
            'ifsc': 'Ifsc Code',
            'debit_account': 'A/c Number',
            'utr': 'Utr No',
            'amount': 'Amount',
        }

        results = []
        uploaded_name = uploaded_name.lower()
        file_path = f"static/upload/df/{uploaded_name}"
        try:
            df = pd.read_excel(file_path, engine='openpyxl')
        except FileNotFoundError:
            logging.error(f"文件未找到: {file_path}")
            return []
        except Exception as e:
            logging.error(f"读取文件时出错: {e}")
            return []
        # 检查字段是否存在
        missing_columns = [table_field for local_field, table_field in local_to_table_map.items() if table_field not in df.columns]
        if missing_columns:
            logging.error(f"文件中缺少必要的列: {', '.join(missing_columns)}")
            return []
        for index, row in df.iterrows():
            result = {}
            for local_field, table_field in local_to_table_map.items():
                if local_field == 'amount':
                    try:
                        result[local_field] = float(row.get(table_field, 0)) or None
                    except (ValueError, TypeError):
                        continue
                else:
                    result[local_field] = row.get(table_field, "")
            # 添加到结果列表
            results.append(result)

        logging.info(f"成功从文件提取 {len(results)} 条记录")
        return results

    @staticmethod
    async def extract_kvb_data(uploaded_name):
        """
        从上传的 CSV 文件中提取 KVB BANK 数据并返回一个字典列表:
        [
            {'utr': 'UTR值', 'code': '账户类型', 'amount': '金额'},
            {'utr': 'UTR值', 'code': '账户类型', 'amount': '金额'},
            ...
        ]
        """
        # 定义 CSV 文件中的列名
        description_field = 'Description'  # 描述字段 (corresponding to Description)
        debit_field = 'Debit'  # Debit 列
        credit_field = 'Credit'  # Credit 列

        results = []  # 初始化结果列表

        # 文件路径构造
        uploaded_name = uploaded_name.lower()  # 文件名转小写，确保一致性

        # 读取 CSV 文件
        file_path = f"static/upload/df/{uploaded_name}"  # 构造文件路径

        try:
            cleaned_file = f"static/upload/df/{uploaded_name}new"
            header_row_index = None
            row_index = 0  # 记录当前行号
            with open(file_path, "r", encoding="utf-8") as infile, open(cleaned_file, "w", encoding="utf-8", newline="") as outfile:
                reader = csv.reader(infile)
                writer = csv.writer(outfile)
                for row in reader:
                    row_index += 1  # 当前行号
                    if len(row) == 8:  # 你的表头有 8 列
                        writer.writerow(row)  # 只保留格式正确的行
                        if header_row_index is None:  # 记录第一行符合格式的索引
                            header_row_index = row_index
            
            # 处理 header_row_index 没有被设置的情况
            if header_row_index is None:
                header_row_index = 1  # 如果没有有效数据，默认设为 1

            logging.info(f"row_index: {row_index}")
            logging.info(f"Header row index: {header_row_index}")

            logging.info("清理完成，已保存为 cleaned_file.csv")
            # 使用 pandas 读取 CSV 文件
            # 如果没有写入任何有效行，设置 header_row_index = 9
            if header_row_index == 1:
                df = pd.read_csv(cleaned_file, header=10)
            else:
                df = pd.read_csv(cleaned_file)
        except FileNotFoundError:
            logging.error(f"文件未找到: {file_path}")
            return []
        except Exception as e:
            logging.error(f"读取文件时出错: {e}")
            return []

        # 检查字段是否存在
        missing_columns = [col for col in [description_field, debit_field, credit_field] if col not in df.columns]
        if missing_columns:
            logging.error(f"文件中缺少必要的列: {', '.join(missing_columns)}")
            return []

        # 遍历每一行数据
        for index, row in df.iterrows():
            # 提取字段数据
            description = row.get(description_field, "")
            debit = str(row.get(debit_field, "")).replace(",", "").replace(" ", "").strip()

            # 检查金额是否为有效数字且 Debit 大于 0
            try:
                # 如果没有 Debit 值，跳过此行
                debit = float(debit) if debit else 0
                if math.isnan(debit) or debit <= 0:
                    continue
            except (ValueError, TypeError):
                continue
            
            utr = ''
            code = ''
            # 提取 UTR 和 code
            if isinstance(description, str):
                parts = description.split('-')
                if len(parts) >= 3:
                    utr = parts[0]  # UTR value is the second part
                    code = parts[3]  # Code is the third part
                    # 检查 code 是否匹配订单号的格式 (以 F 开头，后跟数字)
                    if not code.startswith("F"):
                        continue  # 如果订单号格式不匹配，跳过此行

            # 如果 UTR、code 或 amount 无效，跳过此行
            if not utr or not code or not debit:
                continue

            # 添加到结果列表
            results.append({
                'utr': utr.strip(),
                'code': code.strip(),
                'amount': str(debit).strip()
            })

        logging.info(f"成功从文件提取 {len(results)} 条记录")
        return results

    @staticmethod
    async def extract_kvb_imps_data(uploaded_name):
        """
        从 KVB IMPS 的 CSV 文件中提取交易数据，并返回格式化的列表:
        [
            {'transaction_date': '日期', 'description': '交易描述', 'amount': '金额', 'type': 'debit/credit'},
            ...
        ]
        """
        # 定义 CSV 文件中的列名
        transaction_date_field = 'Transaction Date'  # 交易日期
        value_date_field = 'Value Date'  # 记账日期
        description_field = 'Description'  # 交易描述
        debit_field = 'Debit'  # 借方金额
        credit_field = 'Credit'  # 贷方金额
        balance_field = 'Balance'  # 账户余额
        
        results = []  # 初始化结果列表

        # 构造文件路径
        uploaded_name = uploaded_name.lower()
        file_path = f"static/upload/df/{uploaded_name}"

        # 确保只处理 CSV 文件
        if not file_path.endswith('.csv'):
            logging.error("错误: 仅支持 CSV 文件")
            return []

        try:
            cleaned_file = f"static/upload/df/{uploaded_name}new"
            header_row_index = None
            row_index = 0  # 记录当前行号
            with open(file_path, "r", encoding="utf-8") as infile, open(cleaned_file, "w", encoding="utf-8", newline="") as outfile:
                reader = csv.reader(infile)
                writer = csv.writer(outfile)
                for row in reader:
                    row_index += 1  # 当前行号
                    if len(row) == 8:  # 你的表头有 8 列
                        writer.writerow(row)  # 只保留格式正确的行
                        if header_row_index is None:  # 记录第一行符合格式的索引
                            header_row_index = row_index
            
            # 处理 header_row_index 没有被设置的情况
            if header_row_index is None:
                header_row_index = 1  # 如果没有有效数据，默认设为 1

            logging.info(f"row_index: {row_index}")
            logging.info(f"Header row index: {header_row_index}")

            logging.info("清理完成，已保存为 cleaned_file.csv")
            # 使用 pandas 读取 CSV 文件
            # 如果没有写入任何有效行，设置 header_row_index = 9
            if header_row_index == 1:
                df = pd.read_csv(cleaned_file, header=10)
            else:
                df = pd.read_csv(cleaned_file)
        except FileNotFoundError:
            logging.error(f"文件未找到: {file_path}")
            return []
        except Exception as e:
            logging.error(f"读取文件时出错: {e}")
            return []

        # 确保 CSV 具有所需的列
        missing_columns = [col for col in [transaction_date_field, description_field, debit_field, credit_field] if col not in df.columns]
        if missing_columns:
            logging.error(f"错误: 文件缺少必要的列 {missing_columns}")
            return []

        # 遍历数据行
        for _, row in df.iterrows():
            description = row.get(description_field, "").strip()
            debit = str(row.get(debit_field, "0")).replace(",", "").replace(" ", "").strip()
            credit = str(row.get(credit_field, "0")).replace(",", "").replace(" ", "").strip()

            # 处理借方 (debit) / 贷方 (credit)
            try:
                debit_amount = float(debit) if debit else 0
                credit_amount = float(credit) if credit else 0
            except ValueError:
                continue  # 遇到无效金额时跳过

            # 只保留有效交易
            if debit_amount > 0:
                amount = debit_amount
            # elif credit_amount > 0:
            #     transaction_type = "credit"
            #     amount = credit_amount
            else:
                continue  # 如果金额为 0，则跳过

            # **使用正则表达式提取 UTR 和 Code**
            utr_match = re.search(r'IMPS-(\d+)', description)  # 提取 UTR 编号
            utr = utr_match.group(1) if utr_match else ""

            code_match = re.search(r'(F\d+)', description)  # 提取 Code
            code = code_match.group(1) if code_match else ""

            # 添加到结果列表
            results.append({
                'amount': round(amount, 2),
                'utr': utr,
                'code': code,
            })

        # **打印所有数据明细**
        for record in results:
            logging.info(f"utr: {record['utr']}, code: {record['code']}, "
                    f"金额: {record['amount']}")
        logging.info(f"成功从 CSV 文件提取 {len(results)} 条 KVB IMPS 交易记录")
        return results
            
    @staticmethod
    async def extract_yes_business_data(uploaded_name):
        """
        从 YES Business CSV 文件中提取交易数据，并解析 UTR 和 Code:
        [
            {'transaction_date': '日期', 'description': '交易描述', 'utr': 'UTR编号', 'code': '交易代码', 'amount': '金额', 'type': 'debit/credit'},
            ...
        ]
        """
        # 定义 CSV 文件中的列名
        transaction_date_field = 'Transaction Date'  # 交易日期
        description_field = 'Description'  # 交易描述
        reference_number_field = 'Reference Number'  # 交易参考号
        debit_field = 'Withdrawals'  # 借方金额
        credit_field = 'Deposits'  # 贷方金额
        
        results = []  # 初始化结果列表

        # 构造文件路径
        uploaded_name = uploaded_name.lower()
        file_path = f"static/upload/df/{uploaded_name}"

        # 确保只处理 CSV 文件
        if not file_path.endswith('.csv'):
            logging.error("错误: 仅支持 CSV 文件")
            return []

        try:
            cleaned_file = f"static/upload/df/{uploaded_name}new"
            header_row_index = None
            row_index = 0  # 记录当前行号
            with open(file_path, "r", encoding="utf-8") as infile, open(cleaned_file, "w", encoding="utf-8", newline="") as outfile:
                reader = csv.reader(infile)
                writer = csv.writer(outfile)
                for row in reader:
                    row_index += 1  # 当前行号
                    if len(row) == 7:  # 你的表头有 8 列
                        writer.writerow(row)  # 只保留格式正确的行
                        if header_row_index is None:  # 记录第一行符合格式的索引
                            header_row_index = row_index
            
            # 处理 header_row_index 没有被设置的情况
            if header_row_index is None:
                header_row_index = 1  # 如果没有有效数据，默认设为 1

            logging.info(f"row_index: {row_index}")
            logging.info(f"Header row index: {header_row_index}")

            logging.info("清理完成，已保存为 cleaned_file.csv")
            # 使用 pandas 读取 CSV 文件
            # 如果没有写入任何有效行，设置 header_row_index = 9
            if header_row_index == 1:
                df = pd.read_csv(cleaned_file, header=5)
            else:
                df = pd.read_csv(cleaned_file)
        except FileNotFoundError:
            logging.error(f"文件未找到: {file_path}")
            return []
        except Exception as e:
            logging.error(f"读取文件时出错: {e}")
            return []

        # 确保 CSV 具有所需的列
        missing_columns = [col for col in [transaction_date_field, description_field, reference_number_field, debit_field, credit_field] if col not in df.columns]
        if missing_columns:
            logging.error(f"错误: 文件缺少必要的列 {missing_columns}")
            return []

        # 遍历数据行
        for _, row in df.iterrows():
            description = row.get(description_field, "").strip()
            debit = str(row.get(debit_field, "0")).replace(",", "").replace(" ", "").strip()
            credit = str(row.get(credit_field, "0")).replace(",", "").replace(" ", "").strip()

            # 处理借方 (debit) / 贷方 (credit)
            try:
                debit_amount = float(debit) if debit else 0
                credit_amount = float(credit) if credit else 0
            except ValueError:
                continue  # 遇到无效金额时跳过

            # 确定交易类型和金额
            if debit_amount > 0:
                amount = debit_amount
            # elif credit_amount > 0:
            #     transaction_type = "credit"
            #     amount = credit_amount
            else:
                continue  # 如果金额为 0，则跳过

            # **提取 UTR（NEFT/RTGS 编号）**
            utr_match = re.search(r'YESBN\d+', description)  # 匹配 `YESBN12025022705194404`
            utr = utr_match.group(0) if utr_match else ""

            if not utr:
                continue

            # **提取付款人姓名（Payment Name）**
            name_match = re.search(r'-(\D+?)-BULKPAYMENT', description)
            payment_name = name_match.group(1).strip() if name_match else ""

            # 添加到结果列表
            results.append({
                'utr': utr,
                'payment_name': payment_name,
                'amount': round(amount, 2)
            })

        # **打印所有数据明细**
        for record in results:
            print(
                f"UTR: {record['utr']}, payment_name: {record['payment_name']}, "
                f"金额: {record['amount']}")

        logging.info(f"成功从 CSV 文件提取 {len(results)} 条 YES Business 交易记录")
        return results

    @staticmethod
    async def extract_axit_neo_data(uploaded_name):
        """
        提取 Axit Neo 银行对账单中的支出（DR）记录，包含：UTR 和 Beneficiary 名称。
        """
        results = []
        file_path = f"static/upload/df/{uploaded_name.lower()}"

        try:
            df = pd.read_excel(file_path, header=14)
            # 默认字段
            beneficiary_field = 'Transaction Particulars'
            amount_field = 'Amount'
            transaction_field = 'Transaction Type'

            # 如果字段不存在，尝试 header=21 和新字段名
            if not all(col in df.columns for col in [beneficiary_field, amount_field, transaction_field]):
                df = pd.read_excel(file_path, header=19)
                beneficiary_field = 'PARTICULARS'
                amount_field = 'Amount(INR)'
                transaction_field = 'DR/CR'

                # 再次检查字段是否存在
                if not all(col in df.columns for col in [beneficiary_field, amount_field, transaction_field]):
                    logging.error(f"文件中缺少必要的列: {beneficiary_field}, {amount_field}, {transaction_field}")
                    return []

        except FileNotFoundError:
            logging.error(f"文件未找到: {file_path}")
            return []
        except Exception as e:
            logging.error(f"读取文件时出错: {e}")
            return []


        # # 检查字段是否存在
        # beneficiary_field = 'Transaction Particulars'  # Beneficiary 字段名称（请确保字段名称与实际匹配）
        # amount_field = 'Amount'  # Amount 字段名称（请确保字段名称与实际匹配）
        # transaction_field = 'Transaction Type'

        missing_columns = [col for col in [beneficiary_field, amount_field, transaction_field] if col not in df.columns]
        if missing_columns:
            logging.error(f"文件中缺少必要的列: {', '.join(missing_columns)}")
            return []

        # 遍历每一行数据
        for index, row in df.iterrows():
            # 提取字段数据
            particulars = str(row.get(beneficiary_field, '')).strip()
            amount = str(row.get(amount_field, 0)).replace("INR", "").replace(" ", "").replace(",", "").strip()  # 获取 Amount 字段
            txn_type = str(row.get(transaction_field, '')).strip().upper()  # 获取 Transaction Type 字段
            
            logging.info(f"[Info] Row {index}: particulars='{particulars}', amount_raw='{amount}', cleaned_amount='{amount}', txn_type='{txn_type}'")

            # 只处理 DR 类型（支出）
            if txn_type != 'DR':
                continue

            # 提取 UTR 和 Beneficiary
            utr_no = ''  # 获取 UTR 字段
            beneficiary = row.get(beneficiary_field, "")  # 获取 Beneficiary 字段

            # 检查金额是否为有效数字
            try:
                amount = float(amount)
            except (ValueError, TypeError):
                # 如果金额字段无效（不是数字或为 None），跳过此行
                logging.debug(f"第 {index} 行金额无效，跳过")
                continue

            # 如果所有字段为空，跳过此行
            if pd.isna(beneficiary) and pd.isna(amount):
                logging.debug(f"第 {index} 行数据为空，跳过")
                continue

            # 拆分 Transaction Particulars 字段并提取 UTR 和 Beneficiary
            parts = particulars.split('/')
            if len(parts) >= 5:
                utr_no = parts[2].strip()
                beneficiary = parts[3].strip()

            # 添加到结果列表
            results.append({
                'utr': str(utr_no).strip(),
                'beneficiary': str(beneficiary).strip(),
                'amount': str(amount).strip()
            })

        logging.info(f"成功从文件提取 {len(results)} 条 DR（支出）记录")
        return results


    @staticmethod
    async def extract_axit_neo_neft_data(uploaded_name):
        """
        提取 Axit Neo Neft 银行对账单中的支出（DR）记录，包含：UTR 和 Beneficiary 名称。
        """
        results = []
        file_path = f"static/upload/df/{uploaded_name.lower()}"

        try:
            # 使用 pandas 读取 Excel 文件
            # df = pd.read_excel(file_path, engine='openpyxl')  # 修改为 openpyxl 以兼容 .xlsx 文件
            df = pd.read_excel(file_path, header=15)
        except FileNotFoundError:
            logging.error(f"文件未找到: {file_path}")
            return []
        except Exception as e:
            logging.error(f"读取文件时出错: {e}")
            return []

        # 检查字段是否存在
        beneficiary_field = 'Beneficiary Code'  # Beneficiary 字段名称（请确保字段名称与实际匹配）
        amount_field = 'Amount'  # Amount 字段名称（请确保字段名称与实际匹配）
        transaction_field = 'UTR'

        missing_columns = [col for col in [beneficiary_field, amount_field, transaction_field] if col not in df.columns]
        if missing_columns:
            logging.error(f"文件中缺少必要的列: {', '.join(missing_columns)}")
            return []

        # 遍历每一行数据
        for index, row in df.iterrows():
            # 提取字段数据
            code = str(row.get('Beneficiary Code', '')).strip()
            amount = str(row.get(amount_field, 0)).replace("INR", "").replace(" ", "").replace(",", "").strip()  # 获取 Amount 字段
            utr = str(row.get('UTR', '')).strip().upper()  # 获取 Transaction Type 字段
            
            # 检查金额是否为有效数字
            try:
                amount = float(amount)
            except (ValueError, TypeError):
                # 如果金额字段无效（不是数字或为 None），跳过此行
                logging.debug(f"第 {index} 行金额无效，跳过")
                continue

            # 检查是否为空（NaN）
            if code == 'INDIAN' or utr == 'NAN' or amount == 'nan' or code == 'nan':
                logging.debug(f"第 {index} 行数据为空，跳过")
                continue

            # 添加到结果列表
            results.append({
                'utr': str(utr).strip(),
                'code': str(code).strip(),
                'amount': str(amount).strip()
            })
        print(results)
        logging.info(f"成功从文件提取 {len(results)} 条 DR（支出）记录")
        return results
    
    @staticmethod
    async def extract_au_bank_data(uploaded_name):
        """
        提取 AU Bank 对账单中的支出（DR）记录，包含：UTR（交易参考号）、订单编号（Payment Instructions1）、金额（Amount Payable）。
        """

        results = []
        file_path = f"static/upload/df/{uploaded_name.lower()}"

        try:
            # 读取 Excel 文件，第14行是表头（从索引 14 开始，header=14）
            df = pd.read_excel(file_path, header=13)
        except FileNotFoundError:
            logging.error(f"文件未找到: {file_path}")
            return []
        except Exception as e:
            logging.error(f"读取文件时出错: {e}")
            return []

        # 指定 AU Bank 对应的字段
        utr_field = 'UTR no./ Transaction Reference no.'
        amount_field = 'Amount Payable'
        code_field = 'Payment Instructions1'

        # 校验必需字段是否存在
        missing_columns = [col for col in [utr_field, amount_field, code_field] if col not in df.columns]
        if missing_columns:
            logging.error(f"文件中缺少必要的列: {', '.join(missing_columns)}")
            return []

        # 遍历每一行数据
        for index, row in df.iterrows():
            # 提取并清洗字段数据
            utr = str(row.get(utr_field, '')).strip().upper()
            amount = str(row.get(amount_field, 0)).replace("INR", "").replace(" ", "").replace(",", "").strip()
            code = str(row.get(code_field, '')).strip()

            # 跳过无效金额
            try:
                amount = float(amount)
            except (ValueError, TypeError):
                logging.debug(f"第 {index} 行金额无效，跳过")
                continue

            # 跳过无效记录
            if not utr or utr.lower() == 'nan' or not code or code.lower() == 'nan':
                logging.debug(f"第 {index} 行数据不完整，跳过")
                continue

            # 添加到结果
            results.append({
                'utr': utr,
                'code': code,
                'amount': amount
            })

        logging.info(f"成功从文件提取 {len(results)} 条 AU Bank DR（支出）记录")
        return results

    async def extract_marspay_data(self, uploaded_name):
        """
        从上传的 Excel 文件中提取数据并返回一个字典列表:
        [
            {'utr': 'Txnid值', 'amount': '金额', 'code': 'Client ID值'},
            {'utr': 'Txnid值', 'amount': '金额', 'code': 'Client ID值'},
            ...
        ]
        """
        # 定义 Excel 文件中的列名
        utr_field = 'Txnid'  # UTR 列 (corresponding to Txnid)
        amount_field = 'Amount'  # 金额列 (corresponding to Amount)
        code_field = 'Client ID'  # 账户类型列 (corresponding to Client ID)

        results = []  # 初始化结果列表

        # 文件路径构造
        uploaded_name = uploaded_name.lower()  # 文件名转小写，确保一致性

        # 读取 Excel 文件
        file_path = f"static/upload/df/{uploaded_name}"  # 构造文件路径

        try:
            # 使用 pandas 读取 CSV 文件
            df = pd.read_csv(file_path)  # 使用 pandas 读取 CSV 文件
        except FileNotFoundError:
            logging.error(f"文件未找到: {file_path}")
            return []
        except Exception as e:
            logging.error(f"读取文件时出错: {e}")
            return []

        # 检查字段是否存在
        missing_columns = [col for col in [utr_field, amount_field, code_field] if col not in df.columns]
        if missing_columns:
            logging.error(f"文件中缺少必要的列: {', '.join(missing_columns)}")
            return []

        # 遍历每一行数据
        for index, row in df.iterrows():
            # 提取字段数据
            utr = str(row.get(utr_field, "")).replace(".0", "").strip()
            amount = str(row.get(amount_field, "")).replace(",", "").replace(" ", "").strip()
            code = row.get(code_field, "")
            if pd.isna(utr) or (isinstance(utr, str) and utr.strip().lower() == 'nan'):
                logging.debug(f"第 {index} 行 UTR '{utr}' 为 nan，跳过。")
                continue
            # 检查金额是否为有效数字
            try:
                # 尝试将金额转换为浮点数
                amount = float(amount)
            except (ValueError, TypeError):
                # 如果金额字段无效（不是数字或为 None），跳过此行
                continue

            # 如果所有字段为空，跳过此行
            if pd.isna(utr) and pd.isna(amount) and pd.isna(code):
                logging.info(f"第 {index} 行数据为空，跳过")
                continue
            logging.info(f"添加记录: UTR = {str(utr).strip()}, Amount = {str(amount).strip()}, Client ID = {str(code).strip()}")

            # 添加到结果列表
            results.append({
                'utr': str(utr).strip(),
                'amount': str(amount).strip(),
                'code': str(code).strip()  # Client ID
            })

        logging.info(f"成功从文件提取 {len(results)} 条记录")
        return results


    @tornado.web.authenticated
    async def post(self):
        try:
            # 获取参数
            data = json.loads(self.request.body)
            filename = data.get('filename')
            random_code = data.get('random_code')
            bank_name = data.get('bank_name')
            receipt_type = data.get('receipt_type')
            third_pay_name = data.get('third_pay_name')
            self.logger.info('批量上传回执参数: {}'.format(str(data)))
            name, file_format = filename.rsplit('.', 1)
            uploaded_file_name = '{}_{}.{}'.format(name, random_code, file_format)
            # 根据银行名解析文件内容
            if(receipt_type == '2'):
                if third_pay_name.startswith('haoda'):
                    bank_name = 'haoda'

            handle_file_func = self.bank_handle_map.get(bank_name)
            if handle_file_func:
                results = await handle_file_func(uploaded_file_name)
            else:
                self.logger.error('此银行没有添加解析批量上传回执的方法： {}'.format(bank_name))
                return await self.json_response(msg[10007])
            # 遍历文件内容，并同步数据库
            update_success_order = []   
            if bank_name == 'FINANA BANK':             
                for result in results:
                    update_data = {'utr': result.get('utr')}
                    condition = {'payment_account': result.get('account'), 'amount': result.get('debit_account'), 'ifsc': result.get('ifsc'), 'status': 4}
                    if await self.update_result('orders_df', update_data, condition):
                        update_success_order.append(result)

                        condition = {
                            'payment_account': result.get('account'),
                            'amount': result.get('debit_account'), # Note: This 'amount' is from 'debit_account'
                            'ifsc': result.get('ifsc'),
                            'status': 4
                        }
                        # 2. 构建查询 'code' 的 SQL 语句
                        sql_select_code = """
                            SELECT code
                            FROM orders_df
                            WHERE payment_account = '{payment_account}'
                            AND amount = {amount}
                            AND ifsc = '{ifsc}'
                            AND status = {status}
                            LIMIT 1
                        """.format(
                            payment_account=condition['payment_account'],
                            amount=condition['amount'],
                            ifsc=condition['ifsc'],
                            status=condition['status']
                        )
                        # 3. 执行查询 'code' 的操作
                        code_result = await self.execute(sql_select_code)

                        if code_result and hasattr(code_result, 'rows') and len(code_result.rows) > 0:
                            parent_order_identifier = code_result.rows[0].get('code')
                            if parent_order_identifier:
                                self.logger.info(f"成功获取到 code: {parent_order_identifier}")
                                # ==== 调用新的方法处理子订单逻辑====
                                await self._update_child_orders_status(parent_order_identifier, bank_name)
                                # ======================================================
                        
                    else:
                        self.logger.error(
                            'FINANA BANK回执更新异常, 数据：{data} 操作人{admin}'.format(data=result, admin=self.current_user['id']))
            elif bank_name == 'YES BANK':  # 新增YES BANK的处理逻辑
                for result in results:
                    # 检查是否存在 'ifsc' 数据
                    if 'ifsc' in result and result.get('ifsc'):
                        # 如果存在 'ifsc' 数据，执行更新逻辑
                        # 准备更新的数据
                        update_data = {'utr': result.get('utr')}

                        # 获取当前时间并减去 24 小时
                        time_24_hours_ago = datetime.now() - timedelta(days=1)

                        # 将时间格式化为 MySQL 接受的格式
                        time_24_hours_ago_str = time_24_hours_ago.strftime('%Y-%m-%d %H:%M:%S')

                        condition = {
                            'payment_name': result.get('account'),  # 收款卡号
                            'amount': result.get('amount'),  # 金额
                            'ifsc': result.get('ifsc'),  # ifsc
                            'status': 4,  # 固定状态为 4
                            'time_success': time_24_hours_ago_str  # 使用时间范围条件
                        }

                        # 构建 SQL 更新语句
                        sql_update = """
                            UPDATE orders_df 
                            SET utr = '{utr}'
                            WHERE payment_name = '{payment_name}' 
                            AND amount = {amount} 
                            AND ifsc = '{ifsc}' 
                            AND status = {status} 
                            AND time_success >= '{time_success}' 
                            LIMIT 1
                        """.format(
                            utr=result.get('utr'),
                            payment_name=condition['payment_name'],
                            amount=condition['amount'],
                            ifsc=condition['ifsc'],
                            status=condition['status'],
                            time_success=condition['time_success']
                        )

                        # 执行更新语句
                        update_result = await self.execute(sql_update)
                        if update_result:
                            update_success_order.append(result)  # 成功时将结果加入成功列表

                            sql_select_code = """
                                SELECT code
                                FROM orders_df
                                WHERE payment_name = '{payment_name}'
                                AND amount = {amount}
                                AND ifsc = '{ifsc}'
                                AND status = {status}
                                AND time_success >= '{time_success}'
                                LIMIT 1
                            """.format(
                                # 直接使用和 sql_update 相同的 condition 字典来格式化
                                payment_name=condition['payment_name'],
                                amount=condition['amount'],
                                ifsc=condition['ifsc'],
                                status=condition['status'],
                                time_success=condition['time_success']
                            )

                            # 3. 执行查询 'code' 的操作
                            code_result = await self.execute(sql_select_code)

                            if code_result and hasattr(code_result, 'rows') and len(code_result.rows) > 0:
                                parent_order_identifier = code_result.rows[0].get('code')
                                if parent_order_identifier:
                                    self.logger.info(f"成功获取到 code: {parent_order_identifier}")
                                    # ==== 调用新的方法处理子订单逻辑====
                                    await self._update_child_orders_status(parent_order_identifier, bank_name)
                                    # ======================================================
                        else:
                            error_message = f"YES BANK 回执更新异常, 数据: {result}, 操作人: {self.current_user['id']}"
                            self.logger.error(error_message)
                    else:
                        # 如果不存在 'ifsc' 数据，执行更新逻辑
                        update_data = {'utr': result.get('utr')}
                        condition = {'code': result.get('code'), 'status': 4}
                        if await self.update_result('orders_df', update_data, condition):
                            update_success_order.append(result)

                            # ==== 调用新的方法处理子订单逻辑====
                            parent_order_identifier = result.get('code')
                            self.logger.info(f"成功获取到 code: {parent_order_identifier}")
                            await self._update_child_orders_status(parent_order_identifier, bank_name)
                            # ======================================================
                        else:
                            self.logger.error(
                                'YES BANK回执更新异常, 数据：{data} 操作人{admin}'.format(data=result, admin=self.current_user['id'])
                            )

            elif bank_name =='CANARA BANK':     # 新增CANARA BANK的处理逻辑
                for result in results:
                    update_data = {'utr': result.get('utr')}
                    condition = {'payment_account': result.get('account'), 'payment_name': result.get('name'), 'amount': result.get('amount'),  'status': 4}
                    if await self.update_result('orders_df', update_data, condition):
                        update_success_order.append(result)

                        # 这是你提供的新的 condition 字典，它将作为后续查询的基准
                        condition = {
                            'payment_account': result.get('account'),
                            'payment_name': result.get('name'),
                            'amount': result.get('amount'),
                            'status': 4
                        }

                        # 2. 构建查询 'code' 的 SQL 语句
                        # 根据上面新的 condition 字典，修改 WHERE 子句和 .format() 中的参数
                        sql_select_code = """
                            SELECT code
                            FROM orders_df
                            WHERE payment_account = '{payment_account}'
                            AND payment_name = '{payment_name}'
                            AND amount = {amount}
                            AND status = {status}
                            LIMIT 1
                        """.format(
                            payment_account=condition['payment_account'],
                            payment_name=condition['payment_name'], # 新增：从 condition 获取 payment_name
                            amount=condition['amount'],
                            status=condition['status']
                        )
                        # 3. 执行查询 'code' 的操作
                        code_result = await self.execute(sql_select_code)

                        if code_result and hasattr(code_result, 'rows') and len(code_result.rows) > 0:
                            parent_order_identifier = code_result.rows[0].get('code')
                            if parent_order_identifier:
                                self.logger.info(f"成功获取到 code: {parent_order_identifier}")
                                # ==== 调用新的方法处理子订单逻辑====
                                await self._update_child_orders_status(parent_order_identifier, bank_name)
                                # ======================================================
                    else:
                        self.logger.error(
                            'CANARA BANK回执更新异常, 数据：{data} 操作人{admin}'.format(data=result, admin=self.current_user['id']))
            elif bank_name =='ICICI BANK':     # ICIC BANK的处理逻辑
                for result in results:
                    update_data = {'utr': result.get('utr')}
                    condition = {'code': result.get('code'), 'status': 4}
                    if await self.update_result('orders_df', update_data, condition):
                        update_success_order.append(result)

                        # ==== 调用新的方法处理子订单逻辑====
                        parent_order_identifier = result.get('code')
                        self.logger.info(f"成功获取到 code: {parent_order_identifier}")
                        await self._update_child_orders_status(parent_order_identifier, bank_name)
                        # ======================================================
                    else:
                        self.logger.error(
                            'ICIC BANK回执更新异常, 数据：{data} 操作人{admin}'.format(data=result, admin=self.current_user['id']))
            elif bank_name == 'BANDHAN BANK':  # 新增BANDHAN BANK的处理逻辑
                for result in results:
                    # 'code' 'debit_account' 'utr' 'account' ifsc'
                    update_data = {'utr': result.get('utr')}
                    condition = {'code': result.get('code'), 'status': 4}
                    if await self.update_result('orders_df', update_data, condition):
                        update_success_order.append(result)

                        # ==== 调用新的方法处理子订单逻辑====
                        parent_order_identifier = result.get('code')
                        self.logger.info(f"成功获取到 code: {parent_order_identifier}")
                        await self._update_child_orders_status(parent_order_identifier, bank_name)
                        # ======================================================
                    else:
                        self.logger.error(
                            'BANDHAN BANK回执更新异常, 数据：{data} 操作人{admin}'.format(data=result, admin=self.current_user['id']))
            elif bank_name == 'IDFC BANK':  # 新增IDFC BANK的处理逻辑
                # print("处理银行: IDFC BANK")  # 打印银行名称
                for result in results:
                    # print("当前处理结果:", result)  # 打印当前的 result 数据

                    # 准备更新的数据和条件
                    update_data = {'utr': result.get('utr')}
                    # print("准备更新的数据:", update_data)  # 打印 update_data

                    condition = {'code': result.get('code'), 'status': 4}
                    # print("更新条件:", condition)  # 打印条件

                    # 尝试更新数据并记录成功或失败
                    update_result = await self.update_result('orders_df', update_data, condition)
                    # print("更新结果:", update_result)  # 打印更新结果是否成功

                    if update_result:
                        update_success_order.append(result)
                        # print("更新成功，添加到 update_success_order:", result)  # 成功时打印结果

                        # ==== 调用新的方法处理子订单逻辑====
                        parent_order_identifier = result.get('code')
                        self.logger.info(f"成功获取到 code: {parent_order_identifier}")
                        await self._update_child_orders_status(parent_order_identifier, bank_name)
                        # ======================================================
                    else:
                        error_message = 'IDFC BANK回执更新异常, 数据：{data} 操作人{admin}'.format(
                            data=result,
                            admin=self.current_user['id']
                        )
                        self.logger.error(error_message)
                        print("更新失败:", error_message)  # 失败时打印错误信息


            elif bank_name =='haoda':     # haoda的处理逻辑
                for result in results:
                    update_data = {'utr': result.get('utr')}
                    condition = {'code': result.get('code'), 'status': 4}
                    if await self.update_result('orders_df', update_data, condition):
                        update_success_order.append(result)

                        # ==== 调用新的方法处理子订单逻辑====
                        parent_order_identifier = result.get('code')
                        self.logger.info(f"成功获取到 code: {parent_order_identifier}")
                        await self._update_child_orders_status(parent_order_identifier, bank_name)
                        # ======================================================
                    else:
                        self.logger.error(
                            'haoda回执更新异常, 数据：{data} 操作人{admin}'.format(data=result, admin=self.current_user['id']))
            elif bank_name == 'KOTAK BANK':  # 新增KOTAK BANK的处理逻辑
                # print("处理银行: KOTAK BANK")  # 打印银行名称
                for result in results:
                    # print("当前处理结果:", result)  # 打印当前的 result 数据

                    # 准备更新的数据和条件
                    update_data = {'utr': result.get('utr')}
                    # print("准备更新的数据:", update_data)  # 打印 update_data

                    condition = {'code': result.get('code'), 'status': 4}
                    # print("更新条件:", condition)  # 打印条件

                    # 尝试更新数据并记录成功或失败
                    update_result = await self.update_result('orders_df', update_data, condition)
                    # print("更新结果:", "成功" if update_result else "失败")  # 打印更新结果是否成功

                    if update_result:
                        update_success_order.append(result)
                        # print("更新成功，添加到 update_success_order:", result)  # 成功时打印结果

                        # ==== 调用新的方法处理子订单逻辑====
                        parent_order_identifier = result.get('code')
                        self.logger.info(f"成功获取到 code: {parent_order_identifier}")
                        await self._update_child_orders_status(parent_order_identifier, bank_name)
                        # ======================================================
                    else:
                        error_message = 'KOTAK BANK回执更新异常, 数据：{data} 操作人{admin}'.format(
                            data=result,
                            admin=self.current_user['id']
                        )
                        self.logger.error(error_message)
                        # print("更新失败:", error_message)  # 失败时打印错误信息
            elif bank_name == 'AXIS BANK':  # 新增 AXIS BANK 的处理逻辑
                for result in results:
                    # 打印当前处理的 result 数据（调试用）
                    # print("当前处理结果:", result)

                    # 准备更新的数据
                    update_data = {'utr': result.get('utr')}
                    # print("准备更新的数据:", update_data)

                    # 条件修改为通过收款姓名、卡号和金额匹配代付订单
                    # 获取当前时间并减去 24 小时
                    time_24_hours_ago = datetime.now() - timedelta(days=1)

                    # 将时间格式化为 MySQL 接受的格式
                    time_24_hours_ago_str = time_24_hours_ago.strftime('%Y-%m-%d %H:%M:%S')

                    condition = {
                        'payment_name': result.get('beneficiary'),  # 收款卡号
                        'amount': result.get('amount'),  # 金额
                        'status': 4,  # 固定状态为 4
                        'time_success': time_24_hours_ago_str  # 使用时间范围条件
                    }
                    # print("更新条件:", condition)

                    # Execute the query with parameters
                    sql_update = """
                        UPDATE orders_df 
                        SET utr = '{utr}'
                        WHERE payment_name = '{payment_name}' 
                        AND amount = {amount} 
                        AND status = {status} 
                        AND time_success >= '{time_success}' 
                        LIMIT 1
                    """.format(
                        utr=result.get('utr'),
                        payment_name=condition['payment_name'],
                        amount=condition['amount'],
                        status=condition['status'],
                        time_success=condition['time_success']
                    )
                    # Call the execute method to run the update query
                    update_result = await self.execute(sql_update)
                    if update_result:
                        update_success_order.append(result)  # 成功时将结果加入成功列表
                        # print("更新成功:", result)

                        sql_select_code = """
                            SELECT code
                            FROM orders_df
                            WHERE payment_name = '{payment_name}'
                            AND amount = {amount}
                            AND status = {status}
                            AND time_success >= '{time_success}'
                            LIMIT 1
                        """.format(
                            # These values come from the 'condition' dictionary that aligns with the UPDATE
                            payment_name=condition['payment_name'],
                            amount=condition['amount'],
                            status=condition['status'],
                            time_success=condition['time_success']
                        )
                        # 3. 执行查询 'code' 的操作
                        code_result = await self.execute(sql_select_code)

                        if code_result and hasattr(code_result, 'rows') and len(code_result.rows) > 0:
                            parent_order_identifier = code_result.rows[0].get('code')
                            if parent_order_identifier:
                                self.logger.info(f"成功获取到 code: {parent_order_identifier}")
                                # ==== 调用新的方法处理子订单逻辑====
                                await self._update_child_orders_status(parent_order_identifier, bank_name)
                                # ======================================================
                    else:
                        error_message = f"AXIS BANK 回执更新异常, 数据: {result}, 操作人: {self.current_user['id']}"
                        self.logger.error(error_message)
                        # print("更新失败:", error_message)  # 记录错误信息
            

            elif bank_name == 'AXIS NEO BANK':  # 新增 AXIS NEO BANK 的处理逻辑
                for result in results:
                    # 打印当前处理的 result 数据（调试用）
                    # print("当前处理结果:", result)

                    # 准备更新的数据
                    update_data = {'utr': result.get('utr')}
                    # print("准备更新的数据:", update_data)

                    # 条件修改为通过收款姓名、卡号和金额匹配代付订单
                    # 获取当前时间并减去 24 小时
                    time_24_hours_ago = datetime.now() - timedelta(days=1)

                    # 将时间格式化为 MySQL 接受的格式
                    time_24_hours_ago_str = time_24_hours_ago.strftime('%Y-%m-%d %H:%M:%S')

                    condition = {
                        'payment_name': result.get('beneficiary'),  # 收款卡号
                        'amount': result.get('amount'),  # 金额
                        'status': 4,  # 固定状态为 4
                        'time_success': time_24_hours_ago_str  # 使用时间范围条件
                    }

                    # 打印更新条件，确认参数值
                    # print("更新条件:", condition)
                    async with self.application.db.acquire() as conn:
                        async with conn.cursor(DictCursor) as cur:
                            # 构建 SQL 更新语句
                            sql_update_bank_record = """
                            UPDATE orders_df
                            SET utr = %s
                            WHERE payment_name LIKE %s
                            AND amount = %s
                            AND status = %s
                            AND time_success >= %s
                            LIMIT 1
                            """
                            # 准备要传递的参数
                            params = (
                                result.get('utr'),  # UTR 字段
                                f"%{condition['payment_name']}%",  # 使用 LIKE 来模糊匹配支付名称
                                condition['amount'],  # 金额
                                condition['status'],  # 状态
                                condition['time_success']  # 时间条件
                            )
                            try:
                                # 执行 SQL 更新操作
                                update_result = await cur.execute(sql_update_bank_record, params)
                                # 如果更新成功（影响了至少一行）
                                if update_result:
                                    update_success_order.append(result)  # 将成功的记录加入成功列表
                                    logging.info(f"更新成功: {result}")  # 记录成功信息 
                                    logging.info(f"成功更新记录，更新的 UTR 为: {result.get('utr')}, 操作人: {self.current_user['id']}")
                                    # 如果执行成功，提交事务
                                    await conn.commit()
                                    
                                    # 2. 构建查询 'code' 的 SQL 语句 (使用参数化查询)
                                    sql_select_code = """
                                        SELECT code
                                        FROM orders_df
                                        WHERE payment_name LIKE %s
                                        AND amount = %s
                                        AND status = %s
                                        AND time_success >= %s
                                        LIMIT 1
                                    """
                                    # 准备要传递给 SELECT 查询的参数，与 UPDATE 查询的条件参数一致
                                    select_params = (
                                        f"%{condition['payment_name']}%",  # Matches the LIKE condition
                                        condition['amount'],
                                        condition['status'],
                                        condition['time_success']
                                    )

                                    # 3. 执行查询 'code' 的操作 (将参数作为单独的参数传递给 execute 方法)
                                    code_result = await self.execute(sql_select_code, select_params)

                                    if code_result and hasattr(code_result, 'rows') and len(code_result.rows) > 0:
                                        parent_order_identifier = code_result.rows[0].get('code')
                                        if parent_order_identifier:
                                            self.logger.info(f"成功获取到 code: {parent_order_identifier}")
                                            # ==== 调用新的方法处理子订单逻辑====
                                            await self._update_child_orders_status(parent_order_identifier, bank_name)
                                            # ======================================================
                                else:
                                    # 如果没有行被更新，记录失败信息
                                    error_message = f"AXIS NEO BANK 回执更新异常, 数据: {result}, 操作人: {self.current_user['id']}"
                                    self.logger.error(error_message)
                                    await conn.rollback()  # 回滚事务
                                    # print("更新失败:", error_message)  # 记录错误信息
                            except Exception as e:
                                # 发生异常时，回滚事务
                                await conn.rollback()
                                error_message = f"AXIS NEO BANK 回执更新异常, 数据: {result}, 操作人: {self.current_user['id']}"
                                self.logger.error(error_message)

            elif bank_name == 'AXIS NEO NEFT BANK':  # 新增IDFC BANK的处理逻辑
                # print("处理银行: IDFC BANK")  # 打印银行名称
                for result in results:
                    # print("当前处理结果:", result)  # 打印当前的 result 数据

                    # 准备更新的数据和条件
                    update_data = {'utr': result.get('utr')}
                    # print("准备更新的数据:", update_data)  # 打印 update_data

                    condition = {'code': result.get('code'), 'status': 4}
                    # print("更新条件:", condition)  # 打印条件

                    # 尝试更新数据并记录成功或失败
                    update_result = await self.update_result('orders_df', update_data, condition)
                    # print("更新结果:", update_result)  # 打印更新结果是否成功

                    if update_result:
                        update_success_order.append(result)
                        logging.info("更新成功，添加到 update_success_order:", result)  # 成功时打印结果
                        
                        # ==== 调用新的方法处理子订单逻辑====
                        parent_order_identifier = result.get('code')
                        self.logger.info(f"成功获取到 code: {parent_order_identifier}")
                        await self._update_child_orders_status(parent_order_identifier, bank_name)
                        # ======================================================
                    else:
                        error_message = 'AXIS NEO NEFT BANK回执更新异常, 数据：{data} 操作人{admin}'.format(
                            data=result,
                            admin=self.current_user['id']
                        )
                        # self.logger.error(error_message)
                        # logging.info("更新失败:", error_message)  # 失败时打印错误信息


            elif bank_name == 'FEDERAL BANK':  # 新增 FEDERAL BANK 的处理逻辑
                for result in results:
                    # 打印当前处理的 result 数据（调试用）
                    # print("当前处理结果:", result)

                    # 准备更新的数据
                    update_data = {'utr': result.get('utr')}
                    # print("准备更新的数据:", update_data)

                    # 条件修改为通过收款姓名、卡号和金额匹配代付订单
                    # 获取当前时间并减去 24 小时
                    time_24_hours_ago = datetime.now() - timedelta(days=1)

                    # 将时间格式化为 MySQL 接受的格式
                    time_24_hours_ago_str = time_24_hours_ago.strftime('%Y-%m-%d %H:%M:%S')

                    condition = {
                        'payment_name': result.get('payment_name'),  # 收款卡号
                        'payment_account': result.get('beneficiary'),  # 收款卡号
                        'amount': result.get('amount'),  # 金额
                        'status': 4,  # 固定状态为 4
                        'time_success': time_24_hours_ago_str  # 使用时间范围条件
                    }
                    # print("更新条件:", condition)

                    # Execute the query with parameters
                    if not condition['payment_name']:  
                        sql_update = """
                            UPDATE orders_df 
                            SET utr = '{utr}'
                            WHERE payment_account = '{payment_account}' 
                            AND amount = {amount} 
                            AND status = {status} 
                            AND time_success >= '{time_success}' 
                            LIMIT 1
                        """.format(
                            utr=result.get('utr'),
                            payment_account=condition['payment_account'],
                            amount=condition['amount'],
                            status=condition['status'],
                            time_success=condition['time_success']
                        )
                    else:  
                        sql_update = """
                            UPDATE orders_df 
                            SET utr = '{utr}'
                            WHERE payment_name = '{payment_name}' 
                            AND amount = {amount} 
                            AND status = {status} 
                            AND time_success >= '{time_success}' 
                            LIMIT 1
                        """.format(
                            utr=result.get('utr'),
                            payment_name=condition['payment_name'],
                            amount=condition['amount'],
                            status=condition['status'],
                            time_success=condition['time_success']
                        )
                    # Call the execute method to run the update query
                    update_result = await self.execute(sql_update)
                    if update_result:
                        update_success_order.append(result)  # 成功时将结果加入成功列表
                        # print("更新成功:", result)

                        # 2. 构建查询 'code' 的 SQL 语句，根据 'payment_name' 是否为空来决定 WHERE 子句
                        if not condition['payment_name']:
                            # Match Branch 1 of UPDATE
                            sql_select_code = """
                                SELECT code
                                FROM orders_df
                                WHERE payment_account = '{payment_account}'
                                AND amount = {amount}
                                AND status = {status}
                                AND time_success >= '{time_success}'
                                LIMIT 1
                            """.format(
                                payment_account=condition['payment_account'],
                                amount=condition['amount'],
                                status=condition['status'],
                                time_success=condition['time_success']
                            )
                        else:
                            # Match Branch 2 of UPDATE
                            sql_select_code = """
                                SELECT code
                                FROM orders_df
                                WHERE payment_name = '{payment_name}'
                                AND amount = {amount}
                                AND status = {status}
                                AND time_success >= '{time_success}'
                                LIMIT 1
                            """.format(
                                payment_name=condition['payment_name'],
                                amount=condition['amount'],
                                status=condition['status'],
                                time_success=condition['time_success']
                            )

                        # 3. 执行查询 'code' 的操作
                        code_result = await self.execute(sql_select_code)

                        if code_result and hasattr(code_result, 'rows') and len(code_result.rows) > 0:
                            parent_order_identifier = code_result.rows[0].get('code')
                            if parent_order_identifier:
                                self.logger.info(f"成功获取到 code: {parent_order_identifier}")
                                # ==== 调用新的方法处理子订单逻辑====
                                # Make sure 'bank_name' is available in this scope if used by _update_child_orders_status
                                await self._update_child_orders_status(parent_order_identifier, bank_name)
                                # ======================================================
                    else:
                        error_message = f"FEDERAL BANK 回执更新异常, 数据: {result}, 操作人: {self.current_user['id']}"
                        self.logger.error(error_message)
                        # print("更新失败:", error_message)  # 记录错误信息

            elif bank_name == 'ESAF SMALL BANK':  # 新增 AXIT BANK 的处理逻辑
                for result in results:
                    # 打印当前处理的 result 数据（调试用）
                    # print("当前处理结果:", result)

                    # 准备更新的数据
                    update_data = {'utr': result.get('utr')}
                    # print("准备更新的数据:", update_data)

                    # 条件修改为通过收款姓名、卡号和金额匹配代付订单
                    # 获取当前时间并减去 24 小时
                    time_24_hours_ago = datetime.now() - timedelta(days=1)

                    # 将时间格式化为 MySQL 接受的格式
                    time_24_hours_ago_str = time_24_hours_ago.strftime('%Y-%m-%d %H:%M:%S')

                    condition = {
                        'code': result.get('code'),  # 收款卡号
                        'amount': result.get('amount'),  # 金额
                        'status': 4,  # 固定状态为 4
                        'time_success': time_24_hours_ago_str  # 使用时间范围条件
                    }
                    # print("更新条件:", condition)

                    # Execute the query with parameters
                    sql_update = """
                        UPDATE orders_df 
                        SET utr = '{utr}'
                        WHERE code = '{code}' 
                        AND amount = {amount} 
                        AND status = {status} 
                        AND time_success >= '{time_success}' 
                        LIMIT 1
                    """.format(
                        utr=result.get('utr'),
                        code=condition['code'],
                        amount=condition['amount'],
                        status=condition['status'],
                        time_success=condition['time_success']
                    )
                    # Call the execute method to run the update query
                    update_result = await self.execute(sql_update)
                    if update_result:
                        update_success_order.append(result)  # 成功时将结果加入成功列表
                        # print("更新成功:", result)

                        # ==== 调用新的方法处理子订单逻辑====
                        parent_order_identifier = result.get('code')
                        self.logger.info(f"成功获取到 code: {parent_order_identifier}")
                        await self._update_child_orders_status(parent_order_identifier, bank_name)
                        # ======================================================
                    else:
                        error_message = f"ESAF SMALL BANK 回执更新异常, 数据: {result}, 操作人: {self.current_user['id']}"
                        self.logger.error(error_message)
                        # print("更新失败:", error_message)  # 记录错误信息

            elif bank_name == 'KVB BANK':  # 新增 KVB BANK 的处理逻辑
                for result in results:
                    # 打印当前处理的 result 数据（调试用）
                    # print("当前处理结果:", result)

                    # 准备更新的数据
                    update_data = {'utr': result.get('utr')}
                    # print("准备更新的数据:", update_data)

                    # 条件修改为通过收款姓名、卡号和金额匹配代付订单
                    # 获取当前时间并减去 24 小时
                    time_24_hours_ago = datetime.now() - timedelta(days=1)

                    # 将时间格式化为 MySQL 接受的格式
                    time_24_hours_ago_str = time_24_hours_ago.strftime('%Y-%m-%d %H:%M:%S')

                    condition = {
                        'code': result.get('code'),  # code
                        'amount': result.get('amount'),  # 金额
                        'status': 4,  # 固定状态为 4
                        'time_success': time_24_hours_ago_str  # 使用时间范围条件
                    }
                    # print("更新条件:", condition)

                    # Execute the query with parameters
                    sql_update = """
                        UPDATE orders_df 
                        SET utr = '{utr}'
                        WHERE code = '{code}' 
                        AND amount = {amount} 
                        AND status = {status} 
                        AND time_success >= '{time_success}' 
                        LIMIT 1
                    """.format(
                        utr=result.get('utr'),
                        code=condition['code'],
                        amount=condition['amount'],
                        status=condition['status'],
                        time_success=condition['time_success']
                    )
                    # Call the execute method to run the update query
                    update_result = await self.execute(sql_update)
                    if update_result:
                        update_success_order.append(result)  # 成功时将结果加入成功列表
                        # print("更新成功:", result)

                        # ==== 调用新的方法处理子订单逻辑====
                        parent_order_identifier = result.get('code')
                        self.logger.info(f"成功获取到 code: {parent_order_identifier}")
                        await self._update_child_orders_status(parent_order_identifier, bank_name)
                        # ======================================================
                    else:
                        error_message = f"KVB BANK 回执更新异常, 数据: {result}, 操作人: {self.current_user['id']}"
                        self.logger.error(error_message)
                        # print("更新失败:", error_message)  # 记录错误信息
            elif bank_name == 'KVB IMPS':  # 新增 KVB IMPS 的处理逻辑
                for result in results:
                    # 打印当前处理的 result 数据（调试用）
                    # print("当前处理结果:", result)

                    # 准备更新的数据
                    update_data = {'utr': result.get('utr')}
                    # print("准备更新的数据:", update_data)

                    # 条件修改为通过收款姓名、卡号和金额匹配代付订单
                    # 获取当前时间并减去 24 小时
                    time_24_hours_ago = datetime.now() - timedelta(days=1)

                    # 将时间格式化为 MySQL 接受的格式
                    time_24_hours_ago_str = time_24_hours_ago.strftime('%Y-%m-%d %H:%M:%S')

                    condition = {
                        'code': result.get('code'),  # code
                        'amount': result.get('amount'),  # 金额
                        'status': 4,  # 固定状态为 4
                        'time_success': time_24_hours_ago_str  # 使用时间范围条件
                    }
                    # print("更新条件:", condition)

                    # Execute the query with parameters
                    sql_update = """
                        UPDATE orders_df 
                        SET utr = '{utr}'
                        WHERE code = '{code}' 
                        AND amount = {amount} 
                        AND status = {status} 
                        AND time_success >= '{time_success}' 
                        LIMIT 1
                    """.format(
                        utr=result.get('utr'),
                        code=condition['code'],
                        amount=condition['amount'],
                        status=condition['status'],
                        time_success=condition['time_success']
                    )
                    # Call the execute method to run the update query
                    update_result = await self.execute(sql_update)
                    if update_result:
                        update_success_order.append(result)  # 成功时将结果加入成功列表
                        # print("更新成功:", result)

                        # ==== 调用新的方法处理子订单逻辑====
                        parent_order_identifier = result.get('code')
                        self.logger.info(f"成功获取到 code: {parent_order_identifier}")
                        await self._update_child_orders_status(parent_order_identifier, bank_name)
                        # ======================================================
                    else:
                        error_message = f"KVB IMPS 回执更新异常, 数据: {result}, 操作人: {self.current_user['id']}"
                        self.logger.error(error_message)
                        # print("更新失败:", error_message)  # 记录错误信息
            elif bank_name == 'YES BUSINESS':  # 新增 YES Business 的处理逻辑
                for result in results:
                    # 打印当前处理的 result 数据（调试用）
                    # print("当前处理结果:", result)

                    # 准备更新的数据
                    update_data = {'utr': result.get('utr')}
                    # print("准备更新的数据:", update_data)

                    # 条件修改为通过收款姓名、卡号和金额匹配代付订单
                    # 获取当前时间并减去 24 小时
                    time_24_hours_ago = datetime.now() - timedelta(days=1)

                    # 将时间格式化为 MySQL 接受的格式
                    time_24_hours_ago_str = time_24_hours_ago.strftime('%Y-%m-%d %H:%M:%S')

                    condition = {
                        'payment_name': result.get('payment_name'),  # payment_name
                        'amount': result.get('amount'),  # 金额
                        'status': 4,  # 固定状态为 4
                        'time_success': time_24_hours_ago_str  # 使用时间范围条件
                    }
                    # print("更新条件:", condition)

                    # Execute the query with parameters
                    sql_update = """
                        UPDATE orders_df 
                        SET utr = '{utr}'
                        WHERE payment_name = '{payment_name}' 
                        AND amount = {amount} 
                        AND status = {status} 
                        AND time_success >= '{time_success}' 
                        LIMIT 1
                    """.format(
                        utr=result.get('utr'),
                        payment_name=condition['payment_name'],
                        amount=condition['amount'],
                        status=condition['status'],
                        time_success=condition['time_success']
                    )
                    # Call the execute method to run the update query
                    update_result = await self.execute(sql_update)
                    if update_result:
                        update_success_order.append(result)  # 成功时将结果加入成功列表
                        # print("更新成功:", result)

                        # 2. 构建查询 'code' 的 SQL 语句
                        sql_select_code = """
                            SELECT code
                            FROM orders_df
                            WHERE payment_name = '{payment_name}'
                            AND amount = {amount}
                            AND status = {status}
                            AND time_success >= '{time_success}'
                            LIMIT 1
                        """.format(
                            # Use the same 'condition' dictionary that supplied parameters for sql_update
                            payment_name=condition['payment_name'],
                            amount=condition['amount'],
                            status=condition['status'],
                            time_success=condition['time_success']
                        )

                        # 3. 执行查询 'code' 的操作
                        code_result = await self.execute(sql_select_code)

                        if code_result and hasattr(code_result, 'rows') and len(code_result.rows) > 0:
                            parent_order_identifier = code_result.rows[0].get('code')
                            if parent_order_identifier:
                                self.logger.info(f"成功获取到 code: {parent_order_identifier}")
                                # ==== 调用新的方法处理子订单逻辑====
                                await self._update_child_orders_status(parent_order_identifier, bank_name)
                                # ======================================================
                    else:
                        error_message = f"YES Business 回执更新异常, 数据: {result}, 操作人: {self.current_user['id']}"
                        self.logger.error(error_message)
                        # print("更新失败:", error_message)  # 记录错误信息

            elif bank_name == 'IOB NEW BANK':
                condition_fields = ['payment_name', 'ifsc', 'amount']
                update_data_fields = ['debit_account', 'utr']
                await self.update_data_by_fields(results, update_success_order, condition_fields, update_data_fields)
            elif bank_name == 'AU BANK':  # 新增 AU Bank 的处理逻辑
                for result in results:
                    # 准备更新的数据
                    update_data = {'utr': result.get('utr')}

                    # 获取当前时间减去 24 小时，用于匹配近24小时的订单
                    time_24_hours_ago = datetime.now() - timedelta(days=1)
                    time_24_hours_ago_str = time_24_hours_ago.strftime('%Y-%m-%d %H:%M:%S')

                    # 构建更新条件
                    condition = {
                        'code': result.get('code'),       # 订单编号
                        'amount': result.get('amount'),   # 金额
                        'status': 4,                      # 状态固定为 4
                        'time_success': time_24_hours_ago_str
                    }

                    # 拼接 SQL 更新语句
                    sql_update = """
                        UPDATE orders_df 
                        SET utr = '{utr}'
                        WHERE code = '{code}' 
                        AND amount = {amount} 
                        AND status = {status} 
                        AND time_success >= '{time_success}' 
                        LIMIT 1
                    """.format(
                        utr=update_data['utr'],
                        code=condition['code'],
                        amount=condition['amount'],
                        status=condition['status'],
                        time_success=condition['time_success']
                    )

                    # 执行更新
                    update_result = await self.execute(sql_update)
                    if update_result:
                        update_success_order.append(result)

                        # ==== 调用新的方法处理子订单逻辑====
                        parent_order_identifier = result.get('code')
                        self.logger.info(f"成功获取到 code: {parent_order_identifier}")
                        await self._update_child_orders_status(parent_order_identifier, bank_name)
                        # ======================================================
                    else:
                        error_message = f"AU Bank 回执更新异常, 数据: {result}, 操作人: {self.current_user['id']}"
                        self.logger.error(error_message)

            elif bank_name == 'marspay':  # 新增 marspay 的处理逻辑
                for result in results:
                    # 准备更新的数据
                    update_data = {'utr': result.get('utr')}

                    # 获取当前时间减去 24 小时，用于匹配近24小时的订单
                    time_24_hours_ago = datetime.now() - timedelta(days=1)
                    time_24_hours_ago_str = time_24_hours_ago.strftime('%Y-%m-%d %H:%M:%S')

                    # 构建更新条件
                    condition = {
                        'code': result.get('code'),       # 订单编号
                        'amount': result.get('amount'),   # 金额
                        'status': 4,                      # 状态固定为 4
                        'time_success': time_24_hours_ago_str
                    }

                    # 拼接 SQL 更新语句
                    sql_update = """
                        UPDATE orders_df 
                        SET utr = '{utr}'
                        WHERE code = '{code}' 
                        AND amount = {amount} 
                        AND status = {status} 
                        AND time_success >= '{time_success}' 
                        LIMIT 1
                    """.format(
                        utr=update_data['utr'],
                        code=condition['code'],
                        amount=condition['amount'],
                        status=condition['status'],
                        time_success=condition['time_success']
                    )

                    # 执行更新
                    update_result = await self.execute(sql_update)
                    if update_result:
                        update_success_order.append(result)

                        # ==== 调用新的方法处理子订单逻辑====
                        parent_order_identifier = result.get('code')
                        self.logger.info(f"成功获取到 code: {parent_order_identifier}")
                        await self._update_child_orders_status(self, parent_order_identifier, bank_name)
                        # ======================================================
                    else:
                        error_message = f"marspay 回执更新异常, 数据: {result}, 操作人: {self.current_user['id']}"
                        self.logger.error(error_message)
            else:   # 默认IOB BANK的处理逻辑
                for result in results:
                    update_data = {'debit_account': result.get('debit_account'), 'utr': result.get('utr')}
                    condition = {'code': result.get('code'), 'status': 4}
                    if await self.update_result('orders_df', update_data, condition):
                        update_success_order.append(result)

                        # ==== 调用新的方法处理子订单逻辑====
                        parent_order_identifier = result.get('code')
                        self.logger.info(f"成功获取到 code: {parent_order_identifier}")
                        await self._update_child_orders_status(parent_order_identifier, bank_name)
                        # ======================================================
                    else:
                        self.logger.error(
                            '回执更新异常，单号{code}, 数据：{data} 操作人{admin}'.format(code=result.get('code'), data=result, admin=self.current_user['id']))
                
            self.logger.warning('回执更新完成，上传信息共{}条, 成功更新{}条, 成功更新数据： {}, 操作人{}'.format(
                        len(results), len(update_success_order), update_success_order, self.current_user['id']))
            return await self.json_response(msg[20000])        
                
        except Exception as e:
            self.logger.exception(e)
            return await self.json_response(msg[10007])

    async def update_data_by_fields(self, results, update_success_order, condition_fields, update_data_fields):
        """
        condition_fields: 条件字段
        update_data_fields: 更新数据字段
        指定条件字段和更新数据字段，传入后会根据条件更新数据(在条件中额外添加了status = 4和time_success在24小时内)
        """
        for result in results:
            time_24_hours_ago = datetime.now() - timedelta(days=1)
            time_24_hours_ago_str = time_24_hours_ago.strftime('%Y-%m-%d %H:%M:%S')
            update_data = ', '.join(["{} = '{}'".format(field, result.get(field)) for field in update_data_fields])
            condition_list = ["{} = '{}'".format(field, result.get(field)) for field in condition_fields]
            condition_list.extend(["status = 4", "time_success > '{}'".format(time_24_hours_ago_str)])
            condition = ' AND '.join(condition_list)

            sql_update = "UPDATE orders_df SET {} WHERE {} LIMIT 1".format(update_data, condition)
            print("sql_update: ", sql_update)
            update_result = await self.execute(sql_update)
            if update_result:
                update_success_order.append(result)
                # print("更新成功:", result)

                condition_fields = ['payment_name', 'ifsc', 'amount']

                # 1. 构建 condition 字典
                condition = {}
                for field in condition_fields:
                    # 根据字段名从 result 中获取对应的值
                    # 假设 'amount' 字段对应 result.get('amount')，而不是 result.get('debit_account')
                    # 如果 'amount' 确实来自 'debit_account'，你需要在这里添加一个特例处理
                    if field == 'amount':
                        condition[field] = result.get('amount') # 或者 result.get('debit_account') 如果是这个源
                    else:
                        condition[field] = result.get(field)

                condition['status'] = 4 # 'status' 是固定条件，不随 condition_fields 变化

                # 2. 构建查询 'code' 的 SQL 语句
                # 动态构建 WHERE 子句
                where_clauses = []
                format_params = {}

                for field in condition_fields:
                    if field in condition and condition[field] is not None:
                        # 对于字符串类型的字段，需要加引号；对于数字类型，不需要
                        # 这里假设 payment_name 和 ifsc 是字符串，amount 是数字
                        if field in ['payment_name', 'ifsc']:
                            where_clauses.append(f"{field} = '{{{field}}}'")
                            format_params[field] = condition[field]
                        elif field == 'amount':
                            where_clauses.append(f"{field} = {{{field}}}")
                            format_params[field] = condition[field]

                # 添加固定的 status 条件
                where_clauses.append("status = {status}")
                format_params['status'] = condition['status']

                # 将所有 WHERE 子句用 AND 连接起来
                where_string = " AND ".join(where_clauses)

                sql_select_code = f"""
                    SELECT code
                    FROM orders_df
                    WHERE {where_string}
                    LIMIT 1
                """

                # 使用构建好的参数字典进行格式化
                sql_select_code = sql_select_code.format(**format_params)


                # 3. 执行查询 'code' 的操作
                code_result = await self.execute(sql_select_code)

                if code_result and hasattr(code_result, 'rows') and len(code_result.rows) > 0:
                    parent_order_identifier = code_result.rows[0].get('code')
                    if parent_order_identifier:
                        self.logger.info(f"成功获取到 code: {parent_order_identifier}")
                        # ==== 调用新的方法处理子订单逻辑====
                        await self._update_child_orders_status(parent_order_identifier, 'IOB NEW BANK')
                        # ======================================================
            else:
                error_message = f"ESAF SMALL BANK 回执更新异常, 数据: {result}, 操作人: {self.current_user['id']}"
                self.logger.error(error_message)


class getBankTypeByPaymentId(BaseHandler):
    @tornado.web.authenticated
    async def post(self):
        data = json.loads(self.request.body)
        payment_id = data.get('payment_id') 
        bank_sql = 'select bank.id,bank.name from bank_type bank \
                left join payment p on p.bank_type_id = bank.id \
                where p.id = {payment_id} limit 1'.format(payment_id = payment_id)
        bank_type = await self.query(bank_sql)
        result_data = {}
        if len(bank_type) > 0: 
            result_data = bank_type[0]
        result = dict(code=20000, data=result_data, msg='获取成功')
        return await self.json_response(result)
        
# 导出未支付的代付订单
class exportOrderDfList(BaseHandler):
    @tornado.web.authenticated
    async def get(self):
        # 获取参数
        payment_id = self.get_argument("payment_id")
        self.logger.info('导出未支付的代付订单参数: {}'.format(str(payment_id)))
        bank_sql = 'select bank.id,bank.name from bank_type bank \
                left join payment p on p.bank_type_id = bank.id \
                where p.id = {payment_id} limit 1'.format(payment_id = payment_id)
        results = await self.query(bank_sql)
        if len(results) > 0: 
            bank_type = results[0]
            
        if bank_type:
            # 查询代付订单
            sql = 'select df.* from orders_df df \
                left join payment p on p.id = df.payment_id\
                where df.status = 1 and p.id= {payment_id} \
                order by df.time_create desc'.format(payment_id = payment_id)
            orders = await self.query(sql)
            
            bank_name = bank_type.get('name')
            if bank_name =='IOB BANK':
                await self.export_iob_df_data(self,orders)
            elif bank_name =='YES BANK':
                await self.export_yes_df_data(self,orders)
            elif bank_name =='ICICI BANK':
                await self.export_icici_df_data(self,orders)
                
    @staticmethod
    async def export_iob_df_data(self, orders):
        # 创建 CSV 文件
        output = io.StringIO()
        writer = csv.writer(output)

        # 写入 CSV 头部
        writer.writerow(['IOB or OTHER BANK Transaction indicator','Account Number','Amount','Narration','IFSC Code','Account type','Name of the Beneficiary','Address of the Beneficiary','Remarks'])

        # 写入数据行
        for data in orders:
            row_data = ['OTH',data['payment_account'],data['amount'],'Salary Credit',data['ifsc'],'10',data['payment_name'],'india',data['code']]
            writer.writerow(row_data)

        # 设置 HTTP 头部信息，让浏览器下载 CSV 文件
        current_date = datetime.now()
        output_filename = 'iob_bank_order_df_{date}{random}.csv'.format(date=current_date.strftime('%Y%m%d'),random= random.randint(100, 999))
        self.set_header('Content-Type', 'text/csv')
        self.set_header('Content-Disposition', 'attachment; filename={filename}'.format(filename=output_filename))

        # 将 CSV 内容返回给前端
        self.write(output.getvalue())
        self.finish()
        
    @staticmethod
    async def export_yes_df_data(self, orders):
        # 读取 Excel 模板
        # template_file = 'application/order/template/yes_bank_export_template.xls'  # 模板文件路径
        # template_df = pd.read_excel(template_file)

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            
            # 计算数据写入起始行
            start_row = 1  # 从第二行开始写入
            current_date = datetime.now()
            # 创建一个 DataFrame，提取所需字段
            df_export = pd.DataFrame({
                "A": "D",  # 固定值
                "B": "IMPS",  # 固定值
                "C": "138427000000039",  # 固定值
                "D": "SHOURYANG FASHIONS",  # 固定值
                "E": "",  # 固定值
                "F": "",  # 固定值
                "G": "",  # 固定值
                "ifsc": [order['ifsc'] for order in orders],
                "payment_account": [order['payment_account'] for order in orders],
                "payment_name": [order['payment_name'] for order in orders],
                "K": "",  # 固定值
                "L": "",  # 固定值
                "M": "",  # 固定值
                "N": "",  # 固定值
                "code": [order['code'] for order in orders],
                "time_create": [order['time_create'].strftime('%Y-%m-%d') for order in orders],
                "amount": [order['amount'] for order in orders],
                "R": "SUP OF HSD",
            })

            # 将数据写入模板的第二行
            df_export.to_excel(writer, sheet_name='Sheet1', startrow=start_row, index=False, header=False)

            worksheet = writer.sheets['Sheet1']

            #  写入首行数据
            # 获取当前日期
            
            worksheet.cell(row=1, column=1, value='H')  # 单元格 1
            worksheet.cell(row=1, column=2, value=current_date.strftime('%d/%m/%Y'))  # 单元格 2
            worksheet.cell(row=1, column=3, value='34700626')  # 单元格 3

            # 插入总条数行
            total_count = len(orders)
            last_row = start_row + total_count  # 计算最后一行的索引
            
            total_amount = sum(order["amount"] for order in orders)
            worksheet.cell(row=last_row + 1, column=1, value='F')  # 单元格 1
            worksheet.cell(row=last_row + 1, column=2, value=len(orders))  # 单元格 2
            worksheet.cell(row=last_row + 1, column=3, value=total_amount)  # 单元格 3

        # 返回文件下载
        output_file_name = 'yes_bank_order_df_{date}{random}.xls'.format(date=current_date.strftime('%Y%m%d'),random= random.randint(100, 999))
        self.set_header('Content-Type', 'application/vnd.ms-excel')
        self.set_header('Content-Disposition', 'attachment; filename="{filename}"'.format(filename=output_file_name))
        self.write(output.getvalue())
        output.close()  # 关闭内存流
            
    @staticmethod
    async def export_icici_df_data(self, orders):
        # 读取 Excel 模板
        template_file = 'application/order/template/icici_bank_export_template.xlsx'  # 模板文件路径
        workbook = load_workbook(template_file)
        worksheet = workbook.active  # 获取活动工作表

        # 计算数据写入起始行
        start_row = 2  # 从第二行开始写入

        # 将数据写入模板的第二行
        for index, order in enumerate(orders):
            worksheet.cell(row=start_row + index, column=1, value='PAB_VENDOR')  # 固定值
            worksheet.cell(row=start_row + index, column=2, value='IMPS')  # 固定值
            worksheet.cell(row=start_row + index, column=3, value='252805002004')  # 固定值
            worksheet.cell(row=start_row + index, column=4, value=order['payment_name'])  # 姓名
            worksheet.cell(row=start_row + index, column=5, value=order['payment_account'])  # 账号
            worksheet.cell(row=start_row + index, column=6, value=order['ifsc'])  # ifsc
            worksheet.cell(row=start_row + index, column=7, value=order['amount'])  # 金额
            worksheet.cell(row=start_row + index, column=8, value=order['time_create'].strftime('%Y-%m-%d'))  # 时间
            worksheet.cell(row=start_row + index, column=9, value=order['code'])  # 订单号

        # 保存导出的 Excel 文件
        current_date = datetime.now()
        output_file = 'icici_bank_order_df_{date}{random}.xls'.format(date=current_date.strftime('%Y%m%d'),random= random.randint(100, 999))
        output = BytesIO()
        workbook.save(output)

        # 返回文件下载
        self.set_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.set_header('Content-Disposition', 'attachment; filename="{filename}"'.format(filename=output_file))
        self.write(output.getvalue())
        output.close()  # 关闭内存流
